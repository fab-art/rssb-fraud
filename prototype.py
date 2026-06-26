import streamlit as st
import pandas as pd
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Page Configuration
st.set_page_config(page_title="PharmaScan CV Workbench", layout="wide", page_icon="💊")

# ─── CUSTOM CSS FOR CARD UI & THEME ───────────────────────────────────────
st.markdown("""
<style>
    .stApp { background-color: #0B1120; color: #E2E8F0; }
    .stSidebar > div { background-color: #0F172A; }
    [data-testid="stHeader"] { background: transparent; }
    
    /* Card Styling */
    .verify-card {
        background: #1E293B; border: 1px solid #334155; border-radius: 12px;
        padding: 16px; margin-bottom: 12px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.3);
    }
    .verify-card:hover { border-color: #00E5A0; transition: 0.2s; }
    .card-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px; }
    .card-title { font-size: 1.1rem; font-weight: 700; color: #00E5A0; }
    .card-badge { padding: 4px 10px; border-radius: 6px; font-size: 0.75rem; font-weight: 600; font-family: monospace; }
    .badge-clean { background: #064E3B; color: #34D399; }
    .badge-review { background: #78350F; color: #FBBF24; }
    .badge-ghost { background: #7F1D1D; color: #FCA5A5; }
    
    .card-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-bottom: 12px; }
    .grid-item { background: #0F172A; padding: 10px; border-radius: 8px; border: 1px solid #1E293B; }
    .grid-label { font-size: 0.7rem; color: #64748B; text-transform: uppercase; font-family: monospace; }
    .grid-value { font-size: 0.95rem; font-weight: 600; color: #F8FAFC; margin-top: 4px; }
    
    /* Streamlit Overrides */
    .stSelectbox label, .stNumberInput label, .stTextArea label { color: #94A3B8; font-size: 0.8rem; }
    .stButton button { background: #00E5A0; color: #0B1120; font-weight: 600; border: none; }
    .stButton button:hover { background: #10B981; }
</style>
""", unsafe_allow_html=True)

# ─── SESSION STATE INIT ───────────────────────────────────────────────────
if 'claims' not in st.session_state: st.session_state.claims = pd.DataFrame()
if 'reviews' not in st.session_state: st.session_state.reviews = {}
if 'hospital_data' not in st.session_state: st.session_state.hospital_data = pd.DataFrame()

# ─── SIDEBAR ──────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("💊 PharmaScan CV")
    st.caption("RSSB Counter-Verification Workbench")
    st.divider()
    
    uploaded_pharma = st.file_uploader("Upload Pharmacy Vouchers", type=["xlsx", "csv"], key="pharma_file")
    uploaded_hospital = st.file_uploader("Upload Hospital Records (Optional)", type=["xlsx", "csv"], key="hospital_file")
    
    if uploaded_pharma is not None:
        if st.button("🚀 Load & Prepare Data", type="primary"):
            try:
                df = pd.read_excel(uploaded_pharma)
                df.columns = [c.strip() for c in df.columns]
                
                # Smart column mapping (adjust if your file headers differ slightly)
                col_map = {
                    'Paper Code': 'paper_code', 'Patient Name': 'patient_name',
                    'Affiliation/RAMA No': 'rama_number', 'Dispensing Date': 'dispensing_date',
                    'Practitioner Name': 'doctor_name', 'Department': 'department',
                    'Medicine Cost': 'med_cost', 'Insurance Co-payment': 'insurance_copay',
                    'Total Cost': 'total_cost'
                }
                df = df.rename(columns=col_map)
                df['dispensing_date'] = pd.to_datetime(df['dispensing_date'], errors='coerce')
                st.session_state.claims = df
                st.success(f"✅ Loaded {len(df)} claims. Ready to verify.")
            except Exception as e:
                st.error(f"Error loading file: {e}")

    if uploaded_hospital is not None:
        if st.button("🏥 Load Hospital Records"):
            try:
                h_df = pd.read_excel(uploaded_hospital)
                h_df.columns = [c.strip() for c in h_df.columns]
                if 'Affiliation/RAMA No' in h_df.columns: h_df.rename(columns={'Affiliation/RAMA No': 'rama_number'}, inplace=True)
                if 'Visit Date' in h_df.columns: h_df.rename(columns={'Visit Date': 'visit_date'}, inplace=True)
                st.session_state.hospital_data = h_df[['rama_number', 'visit_date']].dropna()
                st.success("✅ Hospital data loaded for cross-matching.")
            except Exception as e:
                st.error(f"Error: {e}")

# ─── MAIN APP ─────────────────────────────────────────────────────────────
st.title("📝 Counter-Verification Workbench")
if st.session_state.claims.empty:
    st.info("👈 Upload a pharmacy voucher file in the sidebar to begin.")
    st.stop()

df = st.session_state.claims
reviews = st.session_state.reviews
h_data = st.session_state.hospital_data

# 1. PRE-PROCESSING FLAGS
if 'status_flag' not in df.columns: df['status_flag'] = 'CLEAN'

# Repeat Visit Detection (>1 unique visit per RAMA)
if 'rama_number' in df.columns and 'dispensing_date' in df.columns:
    visit_counts = df.groupby('rama_number')['dispensing_date'].nunique()
    repeats = visit_counts[visit_counts > 1].index.tolist()
    df.loc[df['rama_number'].isin(repeats), 'status_flag'] = 'REPEAT_VISIT'

# Cross-Facility Ghost Detection (RAMA match + ±7 day window)
if not h_data.empty and 'rama_number' in df.columns:
    def check_ghost(row):
        rama = str(row.get('rama_number', '')).strip()
        d_date = row.get('dispensing_date')
        if pd.isna(d_date): return False
        match = h_data[h_data['rama_number'].astype(str).str.strip() == rama]
        if match.empty: return True
        for h_date in match['visit_date']:
            if abs((d_date - pd.to_datetime(h_date)).days) <= 7: return False
        return True
    df['is_ghost'] = df.apply(check_ghost, axis=1)
    df.loc[df['is_ghost'], 'status_flag'] = 'GHOST_PRESCRIPTION'

# 2. DASHBOARD METRICS
c1, c2, c3, c4 = st.columns(4)
total = len(df)
reviewed = len(reviews)
pending = total - reviewed
deductions = sum(r.get('deduction', 0) for r in reviews.values())

c1.metric("Total Claims", total)
c2.metric("Pending Review", pending)
c3.metric("Reviewed", reviewed, f"{(reviewed/total*100):.1f}%")
c4.metric("Total Deductions", f"{deductions:,.0f} RWF", delta_color="inverse")
st.divider()

# 3. VERIFICATION CARD QUEUE
st.subheader("🔍 Claim Verification Queue")
f1, f2, f3 = st.columns([1,1,2])
filter_status = f1.selectbox("Filter by System Flag", ["ALL", "REPEAT_VISIT", "GHOST_PRESCRIPTION", "CLEAN"])
filter_reviewed = f2.radio("Show", ["Pending", "Reviewed", "All"], horizontal=True)
search_term = f3.text_input("🔍 Search (RAMA, Name, Paper Code)", key="search")

# Apply Filters
display_df = df.copy()
if filter_status != "ALL": display_df = display_df[display_df['status_flag'] == filter_status]
if filter_reviewed == "Pending": display_df = display_df[~display_df['paper_code'].astype(str).isin(reviews.keys())]
elif filter_reviewed == "Reviewed": display_df = display_df[display_df['paper_code'].astype(str).isin(reviews.keys())]
if search_term:
    display_df = display_df[
        display_df['patient_name'].astype(str).str.contains(search_term, case=False, na=False) |
        display_df['rama_number'].astype(str).str.contains(search_term, case=False, na=False) |
        display_df['paper_code'].astype(str).str.contains(search_term, case=False, na=False)
    ]

# Pagination (prevents UI freeze on large files)
page_size = 10
pages = max(1, (len(display_df) + page_size - 1) // page_size)
current_page = st.number_input("Page", 1, pages, 1)
page_df = display_df.iloc[(current_page - 1) * page_size : current_page * page_size]

# Render Cards
for _, row in page_df.iterrows():
    code = str(row.get('paper_code', 'N/A'))
    rev = reviews.get(code, {})
    
    flag = row.get('status_flag', 'CLEAN')
    if flag == 'REPEAT_VISIT': badge_cls, badge_txt = 'badge-review', '🔁 REPEAT VISIT'
    elif flag == 'GHOST_PRESCRIPTION': badge_cls, badge_txt = 'badge-ghost', '👻 GHOST PRESCRIPTION'
    else: badge_cls, badge_txt = 'badge-clean', '✅ CLEAN'

    st.markdown(f"""
    <div class="verify-card">
        <div class="card-header">
            <div class="card-title">Paper Code: {code}</div>
            <span class="card-badge {badge_cls}">{badge_txt}</span>
        </div>
        <div class="card-grid">
            <div class="grid-item"><div class="grid-label">Patient</div><div class="grid-value">{row.get('patient_name', '-')}</div></div>
            <div class="grid-item"><div class="grid-label">RAMA / Affiliation</div><div class="grid-value">{row.get('rama_number', '-')}</div></div>
            <div class="grid-item"><div class="grid-label">Date</div><div class="grid-value">{pd.to_datetime(row.get('dispensing_date')).strftime('%Y-%m-%d') if pd.notna(row.get('dispensing_date')) else '-'}</div></div>
            <div class="grid-item"><div class="grid-label">Practitioner</div><div class="grid-value">{row.get('doctor_name', '-')}</div></div>
            <div class="grid-item"><div class="grid-label">Medicine Cost</div><div class="grid-value">{row.get('med_cost', 0):,.0f} RWF</div></div>
            <div class="grid-item"><div class="grid-label">Total Cost</div><div class="grid-value">{row.get('total_cost', 0):,.0f} RWF</div></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    c1, c2, c3, c4, c5 = st.columns([2, 2, 2, 3, 2])
    with c1:
        status = st.selectbox("Status", ["Verified", "Deducted", "Ghost/Rejected"], key=f"st_{code}", 
                              index=["Verified", "Deducted", "Ghost/Rejected"].index(rev.get('status', 'Verified')))
    with c2:
        deduction = st.number_input("Deduction (RWF)", min_value=0.0, value=float(rev.get('deduction', 0.0)), step=500.0, key=f"ded_{code}")
    with c3:
        st.number_input("Co-pay %", value=85.0, disabled=True, key=f"cp_{code}")
    with c4:
        reason = st.text_area("Reason / Notes", value=rev.get('reason', ''), key=f"re_{code}", placeholder="e.g., Qty exceeds limit, Signature mismatch...")
    with c5:
        if st.button("💾 Save", key=f"save_{code}", use_container_width=True):
            reviews[code] = {'status': status, 'deduction': deduction, 'reason': reason}
            st.session_state.reviews = reviews
            st.rerun()
    st.divider()

# 4. DRAFT REPORT & EXPORT
st.subheader("📊 Draft Counter-Verification Report")
report_data = []
for _, row in df.iterrows():
    code = str(row['paper_code'])
    rev = reviews.get(code, {'status': 'Verified', 'deduction': 0.0, 'reason': ''})
    total = float(row.get('total_cost', 0))
    deduction = float(rev['deduction'])
    ins_copay = float(row.get('insurance_copay', total * 0.85))
    
    report_data.append({
        'Paper Code': code, 'Patient Name': row.get('patient_name', ''), 'RAMA No': row.get('rama_number', ''),
        'Date': row.get('dispensing_date', ''), 'Practitioner': row.get('doctor_name', ''),
        'Total Cost': total, 'Status': rev['status'], 'Deduction Amount': deduction,
        '100% After CV': total - deduction, '85% After CV': max(0, ins_copay - deduction), 'Reason/Notes': rev['reason']
    })

report_df = pd.DataFrame(report_data)
st.dataframe(report_df, use_container_width=True, height=400)

if st.button("📥 Export Final CV Report (Excel)", type="primary"):
    wb = Workbook()
    ws = wb.active
    ws.title = "CV Report"
    
    headers = list(report_df.columns)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal='center')
        
    for r, row_data in enumerate(report_df.itertuples(index=False), 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if isinstance(val, float): cell.number_format = '#,##0.00'
            
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    st.download_button(label="✅ Download Ready-to-Submit Excel", data=buf.getvalue(), 
                       file_name=f"RSSB_CV_Report_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx", 
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")