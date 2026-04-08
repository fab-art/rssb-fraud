"""
Counter-Verification Report Generator for PharmaScan.

This module handles:
- Generating two-sheet Excel counter-verification reports
- Matching annotated voucher data with original records
- Formatting cells according to reference specifications
"""

import io
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_counter_verification_xlsx(
    df: pd.DataFrame,
    deductions: list[dict],
    meta: dict,
    prepared_by: str,
    verified_by: str,
    approved_by: str,
    pc_col: Optional[str] = None,
    ins_col: Optional[str] = None,
    tot_col: Optional[str] = None,
    obs_col: Optional[str] = None,
    dif_col: Optional[str] = None,
) -> bytes:
    """
    Generate counter-verification report.
    
    Sheet 1 — "After counter verification"
    Sheet 2 — "Counter verification report"
    
    Args:
        df: Original DataFrame with voucher records
        deductions: List of deduction dictionaries from annotated file
        meta: Metadata dictionary (date range, facility info, etc.)
        prepared_by: Name of person who prepared the report
        verified_by: Name of person who verified the report
        approved_by: Name of person who approved the report
        pc_col: Paper code column name
        ins_col: Insurance column name
        tot_col: Total amount column name
        obs_col: Observation column name
        dif_col: Difference column name
        
    Returns:
        Bytes of the generated Excel file
    """
    # ── Palette (exact hex from reference file) ───────────────────────────────
    C_BLUE = "003366"  # Primary Blue
    C_GOLD = "FFCC00"  # Accent Gold
    C_GREY = "F4F4F4"  # Zebra / metadata bg
    C_WHITE = "FFFFFF"
    C_TEXT = "333333"  # Body text
    C_RED = "C0392B"  # Deduction amounts / NO amounts
    C_GREEN = "1E7E34"  # Verified YES text
    C_AMBER = "B8860B"  # Verified NO text
    C_FILL_GREEN = "D4EDDA"  # YES background
    C_FILL_AMBER = "FFF3CD"  # NO background
    C_TITLE_BG = "E8F0F7"  # Title banner background

    # ── Style helpers ─────────────────────────────────────────────────────────
    def fill(hex_col):
        return PatternFill("solid", fgColor=hex_col)

    def _font(name="Calibri", bold=False, size=11.0, color=C_TEXT):
        return Font(name=name, bold=bold, size=size, color=color)

    def side(style, color="000000"):
        return Side(border_style=style, color=color)

    THIN_GREY = side("thin", "AAAAAA")
    MED_BLUE = side("medium", C_BLUE)
    MED_GOLD = side("medium", C_GOLD)
    THIN_ANY = side("thin")  # thin with default colour
    NONE_S = Side(border_style=None)

    def border_hdr_mid():
        """Header cell middle — medium blue all, gold bottom."""
        return Border(left=MED_BLUE, right=MED_BLUE, top=MED_BLUE, bottom=MED_GOLD)

    def border_hdr_first():
        """Header cell column A/first — thin left, medium rest."""
        return Border(left=THIN_ANY, right=MED_BLUE, top=MED_BLUE, bottom=MED_GOLD)

    def border_hdr_last():
        """Header cell last column — medium left, thin right."""
        return Border(left=MED_BLUE, right=THIN_ANY, top=MED_BLUE, bottom=MED_GOLD)

    def border_data():
        """Regular data cell borders."""
        return Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

    def border_data_first():
        """Data cell col A — thin/? left."""
        return Border(left=THIN_ANY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

    def border_data_last():
        """Data cell last col E — thin/? right."""
        return Border(left=THIN_GREY, right=THIN_ANY, top=THIN_GREY, bottom=THIN_GREY)

    # ── Alignments ────────────────────────────────────────────────────────────
    A_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    A_CENTER_NW = Alignment(horizontal="center", vertical="center", wrap_text=False)
    A_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    A_LEFT_NW = Alignment(horizontal="left", vertical="center", wrap_text=False)
    A_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
    A_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)

    # ── Numeric helpers ───────────────────────────────────────────────────────
    def _safe_float(v):
        if v is None:
            return 0.0
        try:
            if pd.isna(v):
                return 0.0
        except Exception:
            pass
        try:
            return float(str(v).replace(",", "").replace(" ", ""))
        except ValueError:
            return 0.0

    def _safe_date(v):
        if v is None:
            return None
        try:
            if pd.isna(v):
                return None
        except Exception:
            pass
        return v

    def _get_col(row_, *keys, default=""):
        for k in keys:
            if k and k in row_.index:
                v = row_[k]
                try:
                    if pd.notna(v):
                        return v
                except Exception:
                    if v is not None:
                        return v
        return default

    # ── Deduction lookup ──────────────────────────────────────────────────────
    ded_map = {str(d["paper_code"]).strip(): d for d in deductions}

    wb = Workbook()

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 1 — "After counter verification"
    # ═════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "After counter verification"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A2"

    # Column widths — exact from reference file
    col_widths_s1 = [
        13.54,  # A  Paper Code
        17.91,  # B  Dispensing Date
        22.54,  # C  Patient Name
        24.82,  # D  RAMA Number
        28.36,  # E  Practitioner Name
        18.18,  # F  Health Facility
        19.73,  # G  Date of Treatment
        13.63,  # H  Verified
        23.45,  # I  Total Before Counter-V
        35.82,  # J  85% After Counter-V
        23.36,  # K  After Counter-V 100%
        20.82,  # L  After Counter-V 85%
        19.73,  # M  Amount Deducted
        28.00,  # N  Explanation
    ]
    for ci, w in enumerate(col_widths_s1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Row 1: Header ─────────────────────────────────────────────────────────
    headers_s1 = [
        "Paper Code",
        "Dispensing Date",
        "Patient Name",
        "RAMA Number",
        "Practitioner Name",
        "Health Facility",
        "Date of Treatment",
        "Verified",
        "Total Before Counter-V (RWF)",
        "85% After Counter-V (RWF)",
        "After Counter-V 100%",
        "After Counter-V 85%",
        "Amount Deducted (RWF)",
        "Explanation",
    ]
    for ci, h in enumerate(headers_s1, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = _font(bold=True, size=11.0, color=C_WHITE)
        c.fill = fill(C_BLUE)
        c.alignment = A_CENTER
        if ci == 1:
            c.border = border_hdr_first()
        elif ci == len(headers_s1):
            c.border = border_hdr_last()
        else:
            c.border = border_hdr_mid()

    # ── Data rows ─────────────────────────────────────────────────────────────
    start_row = 2
    total_before_sum = 0.0
    total_deducted_sum = 0.0
    verified_yes = 0
    verified_no = 0

    for ri, (_, row) in enumerate(df.iterrows(), start=start_row):
        paper_code = _get_col(row, "voucher_id", "paper_code", pc_col, default="")
        ded_info = ded_map.get(str(paper_code).strip())

        dispensing_date = _safe_date(_get_col(row, "visit_date", "dispensing_date", default=None))
        patient_name = _get_col(row, "patient_name", default="")
        rama_number = _get_col(row, "patient_id", "rama_number", default="")
        practitioner = _get_col(row, "doctor_name", "practitioner_name", default="")
        facility = _get_col(row, "facility", "health_facility", default="")
        treatment_date = _safe_date(_get_col(row, "visit_date", "date_of_treatment", default=None))

        total_before = _safe_float(_get_col(row, "amount", "total_cost", tot_col, default=0.0))
        total_before_sum += total_before

        if ded_info:
            diff_val = _safe_float(ded_info.get("difference", 0.0))
            observation = ded_info.get("observation", "")
            verified_flag = "NO"
            verified_no += 1
            total_deducted_sum += diff_val
            after_100 = total_before - diff_val
        else:
            diff_val = 0.0
            observation = ""
            verified_flag = "YES"
            verified_yes += 1
            after_100 = total_before

        after_85 = after_100 * 0.85
        after_85_label = after_85

        # Write data cells
        data_vals = [
            paper_code,  # A
            dispensing_date.date() if hasattr(dispensing_date, "date") else dispensing_date,  # B
            patient_name,  # C
            rama_number,  # D
            practitioner,  # E
            facility,  # F
            treatment_date.date() if hasattr(treatment_date, "date") else treatment_date,  # G
            verified_flag,  # H
            total_before if total_before else "",  # I
            after_85_label if after_85_label else "",  # J
            after_100 if after_100 else "",  # K
            after_85 if after_85 else "",  # L
            diff_val if diff_val else "",  # M
            observation,  # N
        ]

        for ci, val in enumerate(data_vals, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font = _font(size=10.5)
            c.alignment = A_CENTER if ci in [1, 7, 8] else A_LEFT
            if ci in [9, 10, 11, 12, 13]:
                c.alignment = A_RIGHT
                c.number_format = "#,##0.00"
            if ci == 8:  # Verified column
                if verified_flag == "YES":
                    c.font = _font(bold=True, color=C_GREEN)
                    c.fill = fill(C_FILL_GREEN)
                else:
                    c.font = _font(bold=True, color=C_AMBER)
                    c.fill = fill(C_FILL_AMBER)
            if ci == 13 and diff_val:  # Amount Deducted
                c.font = _font(bold=True, color=C_RED)
            if ci == 1:
                c.border = border_data_first()
            elif ci == len(data_vals):
                c.border = border_data_last()
            else:
                c.border = border_data()

    # ── Footer totals row ─────────────────────────────────────────────────────
    footer_row = len(df) + 2
    footer_vals = [
        "TOTALS",
        "",
        "",
        "",
        "",
        "",
        "",
        f"{verified_yes} YES / {verified_no} NO",
        total_before_sum,
        "",
        "",
        "",
        total_deducted_sum,
        "",
    ]
    for ci, val in enumerate(footer_vals, 1):
        c = ws1.cell(row=footer_row, column=ci, value=val)
        c.font = _font(bold=True, size=11.0)
        c.fill = fill(C_TITLE_BG)
        c.alignment = A_RIGHT if ci in [1, 8, 9, 13] else A_CENTER
        if ci in [9, 13]:
            c.number_format = "#,##0.00"
        if ci == 1:
            c.border = Border(left=MED_BLUE, right=THIN_GREY, top=MED_GOLD, bottom=MED_BLUE)
        elif ci == len(footer_vals):
            c.border = Border(left=THIN_GREY, right=MED_BLUE, top=MED_GOLD, bottom=MED_BLUE)
        else:
            c.border = Border(left=THIN_GREY, right=THIN_GREY, top=MED_GOLD, bottom=MED_BLUE)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 2 — "Counter verification report"
    # ═════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="Counter verification report")
    ws2.sheet_view.showGridLines = False

    # Column widths for sheet 2
    col_widths_s2 = [
        13.54,  # A  Paper Code
        17.91,  # B  Dispensing Date
        22.54,  # C  Patient Name
        24.82,  # D  RAMA Number
        28.36,  # E  Practitioner Name
        18.18,  # F  Health Facility
        19.73,  # G  Date of Treatment
        23.45,  # H  Total Cost
        19.73,  # I  Amount Deducted
        28.00,  # J  Observation
    ]
    for ci, w in enumerate(col_widths_s2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # Header row
    headers_s2 = [
        "Paper Code",
        "Dispensing Date",
        "Patient Name",
        "RAMA Number",
        "Practitioner Name",
        "Health Facility",
        "Date of Treatment",
        "Total Cost (RWF)",
        "Amount Deducted (RWF)",
        "Observation",
    ]
    for ci, h in enumerate(headers_s2, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = _font(bold=True, size=11.0, color=C_WHITE)
        c.fill = fill(C_BLUE)
        c.alignment = A_CENTER
        if ci == 1:
            c.border = border_hdr_first()
        elif ci == len(headers_s2):
            c.border = border_hdr_last()
        else:
            c.border = border_hdr_mid()

    # Deduction-only rows
    ded_row = 2
    for ded in deductions:
        paper_code = ded.get("paper_code", "")
        row_data = df[df["voucher_id"].astype(str).str.strip() == str(paper_code).strip()]

        if len(row_data) > 0:
            row = row_data.iloc[0]
            vals = [
                paper_code,
                _get_col(row, "visit_date", "dispensing_date"),
                _get_col(row, "patient_name", default=""),
                _get_col(row, "patient_id", "rama_number", default=""),
                _get_col(row, "doctor_name", "practitioner_name", default=""),
                _get_col(row, "facility", "health_facility", default=""),
                _get_col(row, "visit_date", "date_of_treatment"),
                _safe_float(_get_col(row, "amount", "total_cost", default=0.0)),
                _safe_float(ded.get("difference", 0.0)),
                ded.get("observation", ""),
            ]
        else:
            vals = [
                paper_code,
                ded.get("dispensing_date", ""),
                ded.get("patient_name", ""),
                ded.get("rama_number", ""),
                ded.get("practitioner_name", ""),
                ded.get("facility", ""),
                ded.get("treatment_date", ""),
                0.0,
                _safe_float(ded.get("difference", 0.0)),
                ded.get("observation", ""),
            ]

        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=ded_row, column=ci, value=val)
            c.font = _font(size=10.5)
            c.alignment = A_CENTER if ci in [1, 7] else A_LEFT
            if ci in [8, 9]:
                c.alignment = A_RIGHT
                c.number_format = "#,##0.00"
            if ci == 9 and val:
                c.font = _font(bold=True, color=C_RED)
            if ci == 1:
                c.border = border_data_first()
            elif ci == len(vals):
                c.border = border_data_last()
            else:
                c.border = border_data()

        ded_row += 1

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
