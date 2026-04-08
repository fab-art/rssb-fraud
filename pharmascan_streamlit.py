"""
PharmaScan — Pharmacy Voucher Intelligence (Streamlit Edition)

Install:
    pip install streamlit pandas matplotlib networkx openpyxl odfpy

Run:
    streamlit run pharmascan_streamlit.py
"""

import difflib
import io
import re
import warnings
from collections import defaultdict as _dd

import matplotlib
matplotlib.use("Agg")
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
import pandas as pd
import streamlit as st

warnings.filterwarnings("ignore")

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PharmaScan",
    page_icon="💊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Colours ───────────────────────────────────────────────────────────────────
ACCENT  = "#00e5a0"
ACCENT2 = "#0ea5e9"
PURPLE  = "#a78bfa"
WARN    = "#f59e0b"
DANGER  = "#ef4444"
MUTED   = "#64748b"
TEXT    = "#e2e8f0"
DARK    = "#0d1117"
BG      = DARK   # alias for matplotlib axes
CARD    = "#111720"
BORDER  = "#1e2a38"

plt.rcParams.update({
    "figure.facecolor": CARD,  "axes.facecolor":  DARK,
    "axes.edgecolor":   BORDER,"axes.labelcolor": MUTED,
    "axes.titlecolor":  TEXT,  "xtick.color":     MUTED,
    "ytick.color":      MUTED, "text.color":      TEXT,
    "grid.color":       BORDER,"grid.linewidth":  0.5,
    "font.family":      "monospace", "font.size": 9,
})

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Mono&display=swap');

.stApp { background: #080c10; }
section[data-testid="stSidebar"] { background: #0d1117 !important; border-right: 1px solid #1e2a38; }

[data-testid="stMetric"] {
    background: #111720; border: 1px solid #1e2a38;
    border-radius: 12px; padding: 16px 20px !important;
}
[data-testid="stMetricLabel"] { color: #64748b !important; font-size: 11px !important; text-transform: uppercase; letter-spacing: .5px; }
[data-testid="stMetricValue"] { color: #e2e8f0 !important; font-size: 26px !important; font-weight: 800 !important; font-family: 'Syne', sans-serif !important; }

.stTabs [data-baseweb="tab-list"] { background: #0d1117; border-bottom: 1px solid #1e2a38; gap: 4px; }
.stTabs [data-baseweb="tab"] { background: transparent; color: #64748b; font-weight: 600; border-radius: 0; border-bottom: 2px solid transparent; padding: 10px 18px; }
.stTabs [aria-selected="true"] { color: #00e5a0 !important; border-bottom: 2px solid #00e5a0 !important; background: transparent !important; }

h1, h2, h3 { font-family: 'Syne', sans-serif !important; }

.sidebar-title { font-family: 'Syne', sans-serif; font-size: 22px; font-weight: 800; color: #e2e8f0; margin-bottom: 4px; }
.sidebar-sub   { font-size: 12px; color: #64748b; margin-bottom: 20px; }

.chip-row { display: flex; flex-wrap: wrap; gap: 8px; margin-top: 8px; }
.chip {
    background: rgba(14,165,233,.08); border: 1px solid rgba(14,165,233,.2);
    border-radius: 6px; padding: 3px 10px; font-size: 11px;
    font-family: 'DM Mono', monospace; color: #0ea5e9;
}
.info-banner {
    background: rgba(14,165,233,.06); border: 1px solid rgba(14,165,233,.2);
    border-radius: 10px; padding: 14px 18px; margin-bottom: 16px;
}
.info-banner b { color: #0ea5e9; }

.sec-head {
    font-family: 'Syne', sans-serif; font-size: 15px; font-weight: 700;
    color: #e2e8f0; padding-left: 10px; border-left: 3px solid #00e5a0;
    margin: 20px 0 12px;
}

.rapid-card {
    background: #111720; border: 1px solid #1e2a38;
    border-left: 3px solid #f59e0b; border-radius: 10px; padding: 12px 14px;
}
.rapid-card.crit { border-left-color: #ef4444; }
.rc-head { display: flex; justify-content: space-between; align-items: flex-start; }
.rc-name { font-size: 13px; font-weight: 600; color: #e2e8f0; }
.rc-id   { font-size: 11px; color: #64748b; font-family: 'DM Mono', monospace; margin-top: 2px; }
.rc-days { font-size: 24px; font-weight: 800; font-family: 'Syne', sans-serif; color: #f59e0b; line-height: 1; }
.rapid-card.crit .rc-days { color: #ef4444; }
.rc-days small { font-size: 11px; font-weight: 400; }
.rc-meta { font-size: 11px; color: #64748b; font-family: 'DM Mono', monospace; margin-top: 5px; }
</style>
""", unsafe_allow_html=True)

# ── Column normalisation — tuned to real Pharmacie Vinca / RAMA vouchers ──────
COLUMN_MAP = {
    # ── Exact columns from the real data ──────────────────────────────────────
    r"#":                                               "row_number",
    r"paper.?code":                                     "voucher_id",
    r"dispensing.?date":                                "visit_date",
    r"patient.?name":                                   "patient_name",
    r"patient.?type":                                   "patient_type",
    r"gender":                                          "gender",
    r"is.?newborn":                                     "is_newborn",
    r"rama.?number":                                    "patient_id",
    r"practitioner.?name":                              "doctor_name",
    r"practitioner.?type":                              "doctor_type",
    r"total.?cost":                                     "amount",
    r"patient.?co.?payment":                            "patient_copay",
    r"insurance.?co.?payment":                          "insurance_copay",
    r"medicine.?cost":                                  "medicine_cost",
    # ── Generic fallbacks ─────────────────────────────────────────────────────
    r"patient.?(id|no|num|number|code)?":               "patient_id",
    r"pat.?id|pid":                                     "patient_id",
    r"doctor.?(id|no|num|code)?":                       "doctor_id",
    r"(doctor|dr|physician|prescriber).?name":          "doctor_name",
    r"doc.?id|did":                                     "doctor_id",
    r"prescriber":                                      "doctor_name",
    r"(visit|service|rx|voucher).?date":                "visit_date",
    r"date.?(of.?)?(visit|service|dispensing)?":        "visit_date",
    r"date":                                            "visit_date",
    r"(pharmacy|facility|clinic|hospital|branch).?(name|id|code)?": "facility",
    r"(drug|medicine|medication|item|product).?(name|description|desc)?": "drug_name",
    r"(drug|medicine|medication|item|product).?(code|id)?":              "drug_code",
    r"(amount|cost|price|value|total|charge)":          "amount",
    r"quantity|qty":                                    "quantity",
    r"(diagnosis|diag|icd|condition)":                  "diagnosis",
    r"(voucher|claim|ref|reference).?(no|number|id|code)?": "voucher_id",
}


@st.cache_data(show_spinner=False)
def load_and_process(file_bytes: bytes, filename: str, rapid_days: int):
    fname = filename.lower()
    if fname.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes), encoding="utf-8", on_bad_lines="skip")
    elif fname.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(file_bytes))
    elif fname.endswith(".ods"):
        df = pd.read_excel(io.BytesIO(file_bytes), engine="odf")
    else:
        raise ValueError("Unsupported file type. Use CSV, XLSX, XLS, or ODS.")

    # Normalise column names
    renamed, used = {}, {}
    for col in df.columns:
        key = re.sub(r"[^a-z0-9]", "_", col.lower().strip())
        key = re.sub(r"_+", "_", key).strip("_")
        matched = False
        for pattern, target in COLUMN_MAP.items():
            if re.fullmatch(pattern, key):
                if target not in used:
                    renamed[col] = target
                    used[target] = col
                    matched = True
                break
        if not matched:
            renamed[col] = key
    df = df.rename(columns=renamed)

    # Parse dates
    if "visit_date" in df.columns:
        df["visit_date"] = pd.to_datetime(df["visit_date"], errors="coerce")
    else:
        for col in df.columns:
            if df[col].dtype == object:
                try:
                    parsed = pd.to_datetime(df[col], errors="coerce")
                    if parsed.notna().sum() > len(df) * 0.5:
                        df["visit_date"] = parsed
                        break
                except Exception:
                    pass

    # Summary stats
    s = {"total_rows": len(df), "columns": list(df.columns)}
    id_col = "patient_id" if "patient_id" in df.columns else "patient_name" if "patient_name" in df.columns else None
    if id_col:
        vc = df[id_col].value_counts()
        s["patient_col"]     = id_col
        s["unique_patients"] = int(df[id_col].nunique())
        s["repeat_patients"] = int((vc > 1).sum())
        s["max_visits"]      = int(vc.max())
        s["top_patients"]    = vc.head(15).rename_axis("id").reset_index(name="visits")
    dcol = "doctor_name" if "doctor_name" in df.columns else "doctor_id" if "doctor_id" in df.columns else None
    if dcol:
        dvc = df[dcol].value_counts()
        s["unique_doctors"] = int(df[dcol].nunique())
        s["top_doctors"]    = dvc.head(15).rename_axis("doctor").reset_index(name="visits")
        s["doctor_col"]     = dcol
    if "visit_date" in df.columns:
        v = df["visit_date"].dropna()
        if len(v):
            s["date_min"] = str(v.min().date())
            s["date_max"] = str(v.max().date())
    if "facility" in df.columns:
        fvc = df["facility"].value_counts()
        s["unique_facilities"] = int(df["facility"].nunique())
        s["top_facilities"]    = fvc.head(10).rename_axis("name").reset_index(name="visits")
    for amt_col in ["amount", "medicine_cost", "insurance_copay", "patient_copay"]:
        if amt_col in df.columns:
            df[amt_col] = pd.to_numeric(df[amt_col], errors="coerce")
    if "amount" in df.columns:
        s["total_amount"] = round(float(df["amount"].sum()), 2)
        s["avg_amount"]   = round(float(df["amount"].mean()), 2)

    # Repeat visits
    repeat_groups, repeat_detail = [], pd.DataFrame()
    if id_col:
        vc2 = df[id_col].value_counts()
        repeat_ids = vc2[vc2 > 1].index.tolist()
        rdf = df[df[id_col].isin(repeat_ids)].copy()
        if "visit_date" in rdf.columns:
            rdf = rdf.sort_values([id_col, "visit_date"])
        repeat_detail = rdf.head(500)
        for pid in repeat_ids[:300]:
            grp = df[df[id_col] == pid]
            entry = {id_col: str(pid), "visits": int(len(grp))}
            if "patient_name" in grp.columns and id_col != "patient_name":
                entry["patient_name"] = str(grp["patient_name"].iloc[0])
            if "visit_date" in grp.columns:
                dates = grp["visit_date"].dropna().sort_values()
                entry["dates"] = ", ".join(str(d.date()) for d in dates if pd.notna(d))
            repeat_groups.append(entry)
        repeat_groups.sort(key=lambda x: x["visits"], reverse=True)

    # Rapid revisits
    rapid = []
    if id_col and "visit_date" in df.columns:
        cols = [id_col, "visit_date"]
        if "patient_name" in df.columns and id_col != "patient_name":
            cols.append("patient_name")
        if dcol:
            cols.append(dcol)
        sub = df[cols].dropna(subset=[id_col, "visit_date"]).sort_values([id_col, "visit_date"])
        for pid, grp in sub.groupby(id_col):
            dates = grp["visit_date"].tolist()
            for i in range(len(dates) - 1):
                diff = (dates[i + 1] - dates[i]).days
                if 0 < diff <= rapid_days:
                    name = str(grp["patient_name"].iloc[0]) if "patient_name" in grp.columns else str(pid)
                    rapid.append({
                        "patient_id":   str(pid),
                        "patient_name": name,
                        "visit_1":      str(dates[i].date()),
                        "visit_2":      str(dates[i + 1].date()),
                        "days_apart":   diff,
                        "doctor":       str(grp[dcol].iloc[0]) if dcol else "—",
                    })
        rapid.sort(key=lambda x: x["days_apart"])

    return df, renamed, s, repeat_groups, repeat_detail, rapid


# ── Chart helpers ─────────────────────────────────────────────────────────────

def hbar_chart(labels, values, color, title, xlabel):
    fig, ax = plt.subplots(figsize=(7, max(2.5, len(labels) * 0.42)))
    bars = ax.barh(labels[::-1], values[::-1],
                   color=color if isinstance(color, list) else [color] * len(labels),
                   height=0.65)
    for bar, val in zip(bars, values[::-1]):
        ax.text(bar.get_width() + max(values) * 0.01,
                bar.get_y() + bar.get_height() / 2,
                str(val), va="center", color=TEXT, fontsize=8)
    ax.set_xlabel(xlabel)
    ax.set_title(title, fontsize=11, fontweight="bold", color=TEXT, pad=10)
    ax.set_xlim(0, max(values) * 1.2)
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(axis="x", alpha=0.3)
    fig.tight_layout()
    return fig


def time_series_chart(df):
    if "visit_date" not in df.columns:
        return None
    s = df["visit_date"].dropna()
    if len(s) < 2:
        return None
    monthly = s.dt.to_period("M").value_counts().sort_index()
    dates, vals = [str(p) for p in monthly.index], monthly.values
    fig, ax = plt.subplots(figsize=(10, 3))
    ax.fill_between(range(len(vals)), vals, alpha=0.2, color=ACCENT)
    ax.plot(range(len(vals)), vals, color=ACCENT, linewidth=2, marker="o", markersize=4)
    step = max(1, len(dates) // 12)
    ax.set_xticks(range(0, len(dates), step))
    ax.set_xticklabels(dates[::step], rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Visits")
    ax.set_title("Monthly Visit Volume", fontsize=11, fontweight="bold", color=TEXT, pad=10)
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    return fig


def rapid_histogram(rapid):
    if not rapid:
        return None
    days = [r["days_apart"] for r in rapid]
    fig, ax = plt.subplots(figsize=(6, 3))
    bins = list(range(1, max(days) + 2))
    _, bins_out, patches = ax.hist(days, bins=bins, color=WARN, edgecolor=CARD, rwidth=0.8)
    for patch, left in zip(patches, bins_out):
        if left <= 2:
            patch.set_facecolor(DANGER)
    ax.set_xlabel("Days Between Visits")
    ax.set_ylabel("Cases")
    ax.set_title("Rapid Revisit Distribution", fontsize=11, fontweight="bold", color=TEXT, pad=10)
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(axis="y", alpha=0.3)
    ax.legend(handles=[mpatches.Patch(color=DANGER, label="≤2 days"),
                        mpatches.Patch(color=WARN,   label="3+ days")],
              fontsize=8, facecolor=CARD, edgecolor=BORDER, labelcolor=TEXT)
    fig.tight_layout()
    return fig


def build_network_data(df: pd.DataFrame, col_a: str, col_b: str,
                       max_nodes: int, min_edge_weight: int):
    """
    Build graph data for vis.js interactive rendering.
    Returns (vis_nodes, vis_edges, stats) or (None, None, {}).
    """
    sub = df[[col_a, col_b]].dropna()
    if len(sub) == 0:
        return None, None, {}

    G = nx.Graph()
    edge_w: dict = {}
    for _, row in sub.iterrows():
        a, b = str(row[col_a]), str(row[col_b])
        G.add_node(a, side="A")
        G.add_node(b, side="B")
        key = (a, b)
        edge_w[key] = edge_w.get(key, 0) + 1

    for (a, b), w in edge_w.items():
        if w >= min_edge_weight:
            G.add_edge(a, b, weight=w)

    G.remove_nodes_from(list(nx.isolates(G)))

    if len(G.nodes) == 0:
        return None, None, {}

    # Prune to top nodes by degree
    if len(G.nodes) > max_nodes:
        top = sorted(G.degree(), key=lambda x: x[1], reverse=True)[:max_nodes]
        G = G.subgraph([n for n, _ in top]).copy()

    nodes_a = [n for n, d in G.nodes(data=True) if d.get("side") == "A"]
    nodes_b = [n for n, d in G.nodes(data=True) if d.get("side") == "B"]
    degrees = dict(G.degree())
    max_deg = max(degrees.values()) if degrees else 1

    col_a_lbl = col_a.replace("_", " ").title()
    col_b_lbl = col_b.replace("_", " ").title()

    # Build vis.js nodes list
    vis_nodes = []
    for n, data in G.nodes(data=True):
        deg = degrees.get(n, 1)
        is_a = data.get("side") == "A"
        size = max(14, min(50, 10 + deg * 4)) if is_a else max(8, min(30, 6 + deg * 2))
        vis_nodes.append({
            "id":    n,
            "label": str(n)[:20] + ("…" if len(str(n)) > 20 else ""),
            "title": f"<b>{n}</b><br>Type: {col_a_lbl if is_a else col_b_lbl}<br>Connections: {deg}",
            "color": {
                "background":  "#00e5a0" if is_a else "#0ea5e9",
                "border":      "#00b87a" if is_a else "#0284c7",
                "highlight":   {"background": "#ffffff", "border": "#00e5a0" if is_a else "#0ea5e9"},
                "hover":       {"background": "#f0fdf4" if is_a else "#e0f2fe",
                                "border":     "#00e5a0" if is_a else "#0ea5e9"},
            },
            "shape": "diamond" if is_a else "dot",
            "size":  size,
            "font":  {"color": "#e2e8f0", "size": 11, "face": "DM Mono, monospace"},
            "group": "A" if is_a else "B",
            "degree": deg,
        })

    # Build vis.js edges list
    max_w = max((G[u][v].get("weight", 1) for u, v in G.edges()), default=1)
    vis_edges = []
    for i, (u, v, data) in enumerate(G.edges(data=True)):
        w = data.get("weight", 1)
        vis_edges.append({
            "id":     i,
            "from":   u,
            "to":     v,
            "weight": w,
            "width":  max(0.5, min(6, 0.5 + 4 * (w / max_w))),
            "color":  {
                "color":     "rgba(100,116,139,0.35)",
                "highlight": "#00e5a0",
                "hover":     "#f59e0b",
            },
            "title":  f"Co-occurrences: {w}",
            "smooth": {"type": "dynamic"},
        })

    stats = {
        "nodes_a":    len(nodes_a),
        "nodes_b":    len(nodes_b),
        "edges":      len(vis_edges),
        "density":    round(nx.density(G), 4),
        "avg_degree": round(sum(degrees.values()) / max(1, len(degrees)), 2),
        "top_a": sorted([(n, degrees[n]) for n in nodes_a], key=lambda x: -x[1])[:10],
        "top_b": sorted([(n, degrees[n]) for n in nodes_b], key=lambda x: -x[1])[:10],
        "col_a_lbl": col_a_lbl,
        "col_b_lbl": col_b_lbl,
    }
    return vis_nodes, vis_edges, stats


def render_vis_network(vis_nodes, vis_edges, stats, physics_mode: str, height: int = 680):
    """Render an interactive vis.js network via st.components.v1.html()."""
    import json
    import streamlit.components.v1 as components

    nodes_json = json.dumps(vis_nodes)
    edges_json = json.dumps(vis_edges)
    col_a_lbl  = stats.get("col_a_lbl", "Node A")
    col_b_lbl  = stats.get("col_b_lbl", "Node B")
    n_a = stats.get("nodes_a", 0)
    n_b = stats.get("nodes_b", 0)

    physics_opts = {
        "Force Atlas 2": json.dumps({
            "solver": "forceAtlas2Based",
            "forceAtlas2Based": {"gravitationalConstant": -60, "centralGravity": 0.01,
                                  "springLength": 120, "springConstant": 0.08, "damping": 0.4},
            "stabilization": {"iterations": 150},
        }),
        "Barnes-Hut": json.dumps({
            "solver": "barnesHut",
            "barnesHut": {"gravitationalConstant": -8000, "centralGravity": 0.3,
                           "springLength": 140, "springConstant": 0.04, "damping": 0.09},
            "stabilization": {"iterations": 150},
        }),
        "Repulsion": json.dumps({
            "solver": "repulsion",
            "repulsion": {"centralGravity": 0.2, "springLength": 200,
                           "springConstant": 0.05, "nodeDistance": 150, "damping": 0.09},
            "stabilization": {"iterations": 150},
        }),
        "None (static)": json.dumps({"enabled": False}),
    }
    physics_json = physics_opts.get(physics_mode, physics_opts["Force Atlas 2"])

    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<script src="https://cdnjs.cloudflare.com/ajax/libs/vis/4.21.0/vis.min.js"></script>
<link  href="https://cdnjs.cloudflare.com/ajax/libs/vis/4.21.0/vis.min.css" rel="stylesheet">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #080c10; font-family: 'DM Mono', monospace; color: #e2e8f0; overflow: hidden; }}

  #net-wrap {{ position: relative; width: 100%; height: {height}px; background: #0d1117;
               border: 1px solid #1e2a38; border-radius: 12px; overflow: hidden; }}
  #network  {{ width: 100%; height: 100%; }}

  /* Toolbar */
  #toolbar {{
    position: absolute; top: 12px; left: 12px; z-index: 10;
    display: flex; gap: 8px; flex-wrap: wrap; align-items: center;
  }}
  .tb-btn {{
    background: rgba(17,23,32,.92); border: 1px solid #1e2a38;
    color: #e2e8f0; border-radius: 7px; padding: 6px 12px;
    font-size: 11px; font-family: monospace; cursor: pointer;
    transition: all .15s; backdrop-filter: blur(4px);
  }}
  .tb-btn:hover {{ border-color: #00e5a0; color: #00e5a0; }}
  .tb-btn.active {{ background: rgba(0,229,160,.12); border-color: #00e5a0; color: #00e5a0; }}
  .tb-sep {{ width: 1px; height: 22px; background: #1e2a38; }}

  /* Search box */
  #search-wrap {{ position: absolute; top: 12px; right: 12px; z-index: 10; display: flex; gap: 6px; }}
  #node-search {{
    background: rgba(17,23,32,.92); border: 1px solid #1e2a38;
    color: #e2e8f0; border-radius: 7px; padding: 6px 12px;
    font-size: 11px; font-family: monospace; width: 180px; outline: none;
    backdrop-filter: blur(4px);
  }}
  #node-search:focus {{ border-color: #00e5a0; }}
  #node-search::placeholder {{ color: #64748b; }}

  /* Legend */
  #legend {{
    position: absolute; bottom: 14px; left: 14px; z-index: 10;
    background: rgba(13,17,23,.88); border: 1px solid #1e2a38;
    border-radius: 10px; padding: 10px 14px; backdrop-filter: blur(4px);
    font-size: 11px;
  }}
  .leg-row {{ display: flex; align-items: center; gap: 8px; margin-bottom: 5px; }}
  .leg-row:last-child {{ margin-bottom: 0; }}
  .leg-dot {{ width: 11px; height: 11px; border-radius: 50%; flex-shrink: 0; }}
  .leg-dia {{ width: 10px; height: 10px; transform: rotate(45deg); flex-shrink: 0; border-radius: 1px; }}

  /* Stats bar */
  #stats-bar {{
    position: absolute; bottom: 14px; right: 14px; z-index: 10;
    background: rgba(13,17,23,.88); border: 1px solid #1e2a38;
    border-radius: 10px; padding: 10px 14px; backdrop-filter: blur(4px);
    font-size: 11px; color: #64748b; line-height: 1.7;
  }}
  #stats-bar b {{ color: #e2e8f0; }}

  /* Tooltip override */
  .vis-tooltip {{
    background: #111720 !important; border: 1px solid #1e2a38 !important;
    color: #e2e8f0 !important; border-radius: 8px !important;
    font-family: 'DM Mono', monospace !important; font-size: 12px !important;
    padding: 8px 12px !important; box-shadow: 0 4px 20px rgba(0,0,0,.4) !important;
  }}

  /* Selected info panel */
  #info-panel {{
    display: none; position: absolute; top: 56px; right: 12px; z-index: 10;
    background: rgba(13,17,23,.95); border: 1px solid #1e2a38;
    border-radius: 10px; padding: 14px 16px; font-size: 12px;
    min-width: 200px; max-width: 260px; backdrop-filter: blur(4px);
  }}
  #info-panel .ip-name {{ font-size: 14px; font-weight: 700; color: #e2e8f0; margin-bottom: 6px; word-break: break-all; }}
  #info-panel .ip-row  {{ display: flex; justify-content: space-between; margin-bottom: 3px; }}
  #info-panel .ip-lbl  {{ color: #64748b; }}
  #info-panel .ip-val  {{ color: #e2e8f0; font-weight: 600; }}
  #info-panel .ip-close {{ float: right; cursor: pointer; color: #64748b; font-size: 14px; margin-left: 8px; }}
  #info-panel .ip-close:hover {{ color: #ef4444; }}
  #info-panel .ip-nbrs {{ margin-top: 8px; border-top: 1px solid #1e2a38; padding-top: 8px; }}
  #info-panel .ip-nbr  {{ color: #64748b; font-size: 11px; margin-bottom: 2px; }}

  #stabilizing {{ position: absolute; top: 50%; left: 50%; transform: translate(-50%,-50%);
    background: rgba(13,17,23,.9); border: 1px solid #1e2a38; border-radius: 10px;
    padding: 16px 24px; font-size: 13px; color: #00e5a0; z-index: 20;
    display: flex; align-items: center; gap: 10px; }}
  .spin {{ width: 16px; height: 16px; border: 2px solid rgba(0,229,160,.2);
    border-top-color: #00e5a0; border-radius: 50%; animation: spin .7s linear infinite; }}
  @keyframes spin {{ to {{ transform: rotate(360deg); }} }}
</style>
</head>
<body>

<div id="net-wrap">
  <div id="stabilizing"><div class="spin"></div> Laying out graph…</div>

  <div id="toolbar">
    <button class="tb-btn" onclick="zoomIn()">＋</button>
    <button class="tb-btn" onclick="zoomOut()">－</button>
    <button class="tb-btn" onclick="fitAll()">⊡ Fit</button>
    <div class="tb-sep"></div>
    <button class="tb-btn" id="btn-physics" onclick="togglePhysics()">⏸ Freeze</button>
    <button class="tb-btn" onclick="highlightHubs()">★ Hubs</button>
    <button class="tb-btn" onclick="resetHighlight()">↺ Reset</button>
    <div class="tb-sep"></div>
    <button class="tb-btn" id="btn-labels" onclick="toggleLabels()">🏷 Labels</button>
  </div>

  <div id="search-wrap">
    <input id="node-search" placeholder="🔍 Search node…" oninput="searchNode(this.value)">
  </div>

  <div id="network"></div>

  <div id="info-panel">
    <div><span class="ip-close" onclick="closeInfo()">✕</span><div class="ip-name" id="ip-name"></div></div>
    <div class="ip-row"><span class="ip-lbl">Type</span><span class="ip-val" id="ip-type"></span></div>
    <div class="ip-row"><span class="ip-lbl">Connections</span><span class="ip-val" id="ip-deg"></span></div>
    <div class="ip-nbrs"><div style="color:#64748b;font-size:10px;margin-bottom:4px;text-transform:uppercase;letter-spacing:.5px">Connected to</div>
      <div id="ip-nbr-list"></div>
    </div>
  </div>

  <div id="legend">
    <div class="leg-row"><div class="leg-dia" style="background:#00e5a0"></div><span>{col_a_lbl} ({n_a})</span></div>
    <div class="leg-row"><div class="leg-dot" style="background:#0ea5e9"></div><span>{col_b_lbl} ({n_b})</span></div>
  </div>

  <div id="stats-bar">
    <div><b>{stats["nodes_a"] + stats["nodes_b"]}</b> nodes</div>
    <div><b>{stats["edges"]}</b> edges</div>
    <div>avg degree <b>{stats["avg_degree"]}</b></div>
    <div>density <b>{stats["density"]}</b></div>
  </div>
</div>

<script>
const nodesData = {nodes_json};
const edgesData = {edges_json};

// Build lookup maps
const nodeMap = {{}};
nodesData.forEach(n => nodeMap[n.id] = n);

const adjMap = {{}};
edgesData.forEach(e => {{
  if (!adjMap[e.from]) adjMap[e.from] = [];
  if (!adjMap[e.to])   adjMap[e.to]   = [];
  adjMap[e.from].push({{id: e.to,   w: e.weight}});
  adjMap[e.to].push(  {{id: e.from, w: e.weight}});
}});

const nodes = new vis.DataSet(nodesData);
const edges = new vis.DataSet(edgesData);

const container = document.getElementById('network');
const physicsConfig = {physics_json};

const options = {{
  nodes: {{ borderWidth: 1.5, shadow: {{ enabled: true, color: 'rgba(0,0,0,.5)', x: 2, y: 2, size: 8 }} }},
  edges: {{ smooth: {{ type: 'dynamic' }}, shadow: false, selectionWidth: 3 }},
  physics: physicsConfig,
  interaction: {{
    hover: true, tooltipDelay: 150,
    navigationButtons: false,
    keyboard: true,
    multiselect: true,
    zoomView: true,
  }},
  layout: {{ improvedLayout: true }},
}};

const network = new vis.Network(container, {{ nodes, edges }}, options);

// Hide spinner once stabilised
network.on('stabilizationIterationsDone', () => {{
  document.getElementById('stabilizing').style.display = 'none';
  network.setOptions({{ physics: {{ enabled: false }} }});
  document.getElementById('btn-physics').textContent = '▶ Unfreeze';
  physicsRunning = false;
}});
network.on('stabilized', () => {{
  document.getElementById('stabilizing').style.display = 'none';
}});

// State
let physicsRunning = true;
let labelsVisible  = true;
let hubsHighlighted = false;

// ── Controls ──
function zoomIn()  {{ network.moveTo({{ scale: network.getScale() * 1.3, animation: true }}); }}
function zoomOut() {{ network.moveTo({{ scale: network.getScale() * 0.77, animation: true }}); }}
function fitAll()  {{ network.fit({{ animation: {{ duration: 600, easingFunction: 'easeInOutQuad' }} }}); }}

function togglePhysics() {{
  physicsRunning = !physicsRunning;
  network.setOptions({{ physics: {{ enabled: physicsRunning }} }});
  document.getElementById('btn-physics').textContent = physicsRunning ? '⏸ Freeze' : '▶ Unfreeze';
  document.getElementById('btn-physics').classList.toggle('active', physicsRunning);
}}

function toggleLabels() {{
  labelsVisible = !labelsVisible;
  const update = nodesData.map(n => ({{
    id: n.id,
    font: {{ ...n.font, color: labelsVisible ? '#e2e8f0' : 'rgba(0,0,0,0)' }}
  }}));
  nodes.update(update);
  document.getElementById('btn-labels').classList.toggle('active', labelsVisible);
}}

function highlightHubs() {{
  hubsHighlighted = !hubsHighlighted;
  if (hubsHighlighted) {{
    const maxDeg = Math.max(...nodesData.map(n => n.degree));
    const thresh = maxDeg * 0.5;
    const update = nodesData.map(n => ({{
      id: n.id,
      opacity: n.degree >= thresh ? 1.0 : 0.15,
    }}));
    nodes.update(update);
  }} else {{
    resetHighlight();
  }}
}}

function resetHighlight() {{
  hubsHighlighted = false;
  nodes.update(nodesData.map(n => ({{ id: n.id, opacity: 1.0 }})));
  document.getElementById('node-search').value = '';
}}

function searchNode(q) {{
  q = q.trim().toLowerCase();
  if (!q) {{ resetHighlight(); return; }}
  const update = nodesData.map(n => ({{
    id: n.id,
    opacity: n.id.toLowerCase().includes(q) ? 1.0 : 0.1,
  }}));
  nodes.update(update);
  // Focus first match
  const match = nodesData.find(n => n.id.toLowerCase().includes(q));
  if (match) {{
    network.focus(match.id, {{ scale: 1.4, animation: {{ duration: 600, easingFunction: 'easeInOutQuad' }} }});
  }}
}}

// ── Click → info panel ──
network.on('click', params => {{
  if (params.nodes.length === 1) {{
    const nid = params.nodes[0];
    const nd  = nodeMap[nid];
    if (!nd) return;
    document.getElementById('ip-name').textContent = nid;
    document.getElementById('ip-type').textContent = nd.group === 'A' ? '{col_a_lbl}' : '{col_b_lbl}';
    document.getElementById('ip-deg').textContent  = nd.degree;
    const nbrs = (adjMap[nid] || []).sort((a,b) => b.w - a.w).slice(0, 10);
    document.getElementById('ip-nbr-list').innerHTML =
      nbrs.map(n => `<div class="ip-nbr">• ${{n.id.length > 24 ? n.id.slice(0,24)+'…' : n.id}} <span style="color:#00e5a0">×${{n.w}}</span></div>`).join('');
    document.getElementById('info-panel').style.display = 'block';

    // Dim non-neighbours
    const connectedIds = new Set([nid, ...nbrs.map(n => n.id)]);
    nodes.update(nodesData.map(n => ({{ id: n.id, opacity: connectedIds.has(n.id) ? 1.0 : 0.1 }})));
    edges.update(edgesData.map(e => ({{
      id: e.id,
      color: {{
        color: (e.from === nid || e.to === nid) ? '#00e5a0' : 'rgba(100,116,139,0.06)',
        highlight: '#00e5a0', hover: '#f59e0b',
      }},
    }})));
  }} else {{
    closeInfo();
    resetHighlight();
  }}
}});

// ── Double-click → zoom to node ──
network.on('doubleClick', params => {{
  if (params.nodes.length === 1) {{
    network.focus(params.nodes[0], {{ scale: 2.0, animation: {{ duration: 500, easingFunction: 'easeInOutQuad' }} }});
  }}
}});

// ── Hover ──
network.on('hoverNode', params => {{
  container.style.cursor = 'pointer';
}});
network.on('blurNode',  () => {{ container.style.cursor = 'default'; }});

function closeInfo() {{
  document.getElementById('info-panel').style.display = 'none';
  resetHighlight();
}}
</script>
</body>
</html>"""

    components.html(html, height=height + 10, scrolling=False)


def fmt_number(n):
    if n >= 1e9:  return f"{n/1e9:.1f}B"
    if n >= 1e6:  return f"{n/1e6:.1f}M"
    if n >= 1e3:  return f"{n/1e3:.1f}K"
    return f"{n:,.0f}"


# ── Name normalisation engine ─────────────────────────────────────────────────

def _toks(name: str) -> set:
    """Lowercase alpha-numeric tokens from a name string."""
    return set(re.sub(r"[^a-z0-9 ]", "", name.lower()).split())

def _seq_ratio(a: str, b: str) -> float:
    sa = " ".join(sorted(_toks(a)))
    sb = " ".join(sorted(_toks(b)))
    return difflib.SequenceMatcher(None, sa, sb).ratio()

def _tok_fuzzy_subset(a: str, b: str, thresh: float = 0.76) -> bool:
    """
    True if every token in the SHORTER name has a fuzzy-close counterpart
    in the LONGER name (catches 'Aurbain'/'Urbain', 'Constatin'/'Constantin').
    """
    ta = list(_toks(a))
    tb = list(_toks(b))
    shorter, longer = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
    for tok in shorter:
        best = max((difflib.SequenceMatcher(None, tok, lt).ratio() for lt in longer), default=0)
        if best < thresh:
            return False
    return True

def _match_score(a: str, b: str):
    """
    Returns (score 0–1, reason str).
    reason ∈ {'subset', 'typo', 'none'}
    """
    ta, tb = _toks(a), _toks(b)
    if not ta or not tb:
        return 0.0, "none"

    shorter, longer = (ta, tb) if len(ta) <= len(tb) else (tb, ta)

    # Rule 1 — exact token subset ('ZACHEE' ⊂ 'Niyonsenga Zachee')
    if shorter <= longer:
        boost = min(0.12, len(shorter) * 0.04)
        return 0.88 + boost, "subset"

    # Rule 2 — fuzzy-token subset: every short token ≈ some long token
    if _tok_fuzzy_subset(a, b):
        return 0.85, "typo"

    # Rule 3 — high overall char-sequence similarity
    ratio = _seq_ratio(a, b)
    if ratio >= 0.88:
        return ratio, "typo"

    return 0.0, "none"


def detect_name_clusters(names: list, counts: dict) -> list[dict]:
    """
    Cluster similar names using Union-Find.
    Returns list of dicts:
      { "canonical": str, "variants": [str,...], "method": str, "confidence": float }
    Only clusters with ≥2 members are returned.
    """
    parent = {n: n for n in names}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        pa, pb = find(a), find(b)
        if pa != pb:
            if len(_toks(pa)) >= len(_toks(pb)):
                parent[pb] = pa
            else:
                parent[pa] = pb

    # ── Pass 1: merge multi-token names (typos + reordering) ─────────────────
    multi = [n for n in names if len(_toks(n)) >= 2]
    for i, a in enumerate(multi):
        for b in multi[i + 1:]:
            sc, why = _match_score(a, b)
            if sc > 0 and why != "none":
                union(a, b)

    # ── Pass 2: single-token names → merge only if token is unique to 1 cluster ──
    def get_clusters():
        c = _dd(list)
        for n in names:
            c[find(n)].append(n)
        return c

    cls1 = get_clusters()
    tok_to_roots: dict = _dd(set)
    for root, members in cls1.items():
        if len(members) > 1:
            for m in members:
                for t in _toks(m):
                    tok_to_roots[t].add(root)

    for name in names:
        if len(_toks(name)) != 1:
            continue
        tok = next(iter(_toks(name)))
        candidates = tok_to_roots.get(tok, set()) - {find(name)}
        if len(candidates) == 1:
            union(name, next(iter(candidates)))

    # ── Build final clusters ─────────────────────────────────────────────────
    final: dict = _dd(list)
    for n in names:
        final[find(n)].append(n)

    def best_canonical(members):
        def score(n):
            tc   = len(_toks(n))
            freq = counts.get(n, 0)
            # Title-case preferred; "Dr " prefix demoted
            titled  = n == n.title()
            no_pfx  = not re.match(r"^(Dr|DR)\s", n)
            return (tc, no_pfx, titled, freq, len(n))
        return max(members, key=score)

    results = []
    for root, members in final.items():
        if len(members) < 2:
            continue
        canon = best_canonical(members)
        variants = [m for m in members if m != canon]
        # Compute overall confidence
        scores = [_match_score(canon, v)[0] for v in variants]
        conf   = round(sum(scores) / len(scores), 3) if scores else 1.0
        # Flag suspicious: variant shares NO token with canonical
        ct = _toks(canon)
        suspicious = any(not (_toks(v) & ct) for v in variants)
        results.append({
            "canonical":  canon,
            "variants":   sorted(variants, key=lambda x: (-counts.get(x, 0), -len(x))),
            "confidence": conf,
            "suspicious": suspicious,
            "count":      len(members),
        })
    results.sort(key=lambda x: (-x["count"], -x["confidence"]))
    return results


def apply_name_normalisation(df: pd.DataFrame, col: str,
                              approved_clusters: list[dict]) -> pd.DataFrame:
    """Apply approved rename clusters to a column in a copy of df."""
    df = df.copy()
    mapping = {}
    for c in approved_clusters:
        for v in c["variants"]:
            mapping[v] = c["canonical"]
    df[col] = df[col].map(lambda x: mapping.get(x, x))
    return df



# ══════════════════════════════════════════════════════════════════════════════
# EMBEDDED DRUG REFERENCE  (RHIA Jan 2025 · UCG 2023 · BNF · FDA · WHO · GINA)
# 1,534 drugs · 84 ATC3 class defaults · base64-gzip encoded
# ══════════════════════════════════════════════════════════════════════════════
import gzip as _gzip, base64 as _b64, json as _json

_DRUG_REF_GZ = "H4sIAAKbumkC/9S9bXPbONYm/FdYXZV7e0qJTIAESXWe+4PeYnvGdvzYSTYztVX70BJtc5sStZSUtrK1+9sfHICkCIAgCZnunp0P00n8gkMKODgv17mu//XLMts/bX/57X/9cmXj8dx2bBvD39bhKvrlt1/O022URUm8jqzxItqFu8hC9jBYPVnxapOE690v738JdwtEv/WK/9GBP9Lf8f6XeL3dZfA7/nljpZn19e7zlfUf1jX9yiaLF/S3O9h3fTIk739ZhS//fb+Od9QQRP8Wr/97Fj3GSfLLb4H9/peHcBvTr/xyd3E5/s06X99dFEagM4T/+P2X//2eLumMJ7Zv26hi/eV6F2WPWbq2JtT2DxYaW469Wjyd2UOySjSmOxXTK8aiwKb/G9qCsa5orH+0lS8dwdoPsDYKLfdsla53z9za/F2j5nftDL1+XjUi2CW+ZL30qjGpfdXhU7qOt7vcFnSGg+XxEbD4CFfRfpOlSbwE032yejKzuWKvazsYDQNiZnBlfXTm/vG7tTgskgjMndhkPLaRaO7FfhWurXHysKe/9mjq5Ggq/bGKqRdfpbdqY2fkShsYi0ZWbLz4+pvFlwz5khb9MQtbD+lul0RbMPPvtjeZ2G6dmZer1Z5ujfMkfdgnVXP/fjSX/njF3Pvb+bRqrI0wkbYAEY1F7tFa+OnC3hjWTp/ypa3Lb5fnR2ullzpe72Lr7tmacYuPP0W3Oz17Hey++Hq02vNwgOlWatoHni2+Y7Dgw4zugE0WPa3D9eJgRT8ieoCoyd+on5iDyVUvd5smh+3ukP23vW1HAX3J233ymK7pCXxv3afLeL86Wv3taPU3wVdUrSbE80bNjq36qsHowoaILn8TUuOjTbxNlxE/a2gyhbNWNfpTsk+zdJ+FizixiL16OtOdNlSxcjyFwzarmIpHI9lUx9bb+vlm+vk363MWJvRw7dJd+hIvPtLP9qw8aqW1qH9rXdfHuFdryw2BOm6IRZgszHeE6zrUcPvUHTGt3RFEuqyj1SLehQ9wg9AHsi3qr2/TP5ZRZj3S13ifJvtdTC8k+Mvl+nG/pX/p9CFUHgP72O7+GPz1T4s3Tx3HR/oc5bunx546ZekavF+Hv0fWt3SdrthJppdi9W1rvUbFSJu+mWZH51RcBl+QXrAxdRJ0UWL9iMNkm7/m6UzxFuEioa/5JUqYS3syfYU11r3mFaLvE/mwwc5ZHV6oz6U75D3dCV817w9pdqszwu5IttHWBjvFbuVrIvv4AmcIjT/ZtisYtwu3W3qCrOfDMktfDvSyJu+sXbqJ6dGytvkuPVo8O1pMf13F4opfcAIfke4v9Uu+GD0aq3CXJunTgb5VaxMufj9Dbh7foO/f0UgKKLN4ne6iRQhPCU5MH0bqPn7HU8PI0z/+O+q9/iGHvZ/SbEWtpIGQRYMT8GEP5Gkw2S+jLQ3lIDby7Pxfj8bfHY2/E7zY5TX44Nt59W07I3vUPZi8nN6fXY0nYyteP4dJlMETOPaSe99ADcu27Mq1wlW8THdZ/DPlaQd939ZgFT0lEDpF0pc9TziJgk8OdLsc+17LR+FIocV5SH9N+pSFj7DRaYCRLqLlPov4XrfH32VvNv4ZJSH93Oi9t9RsaVuzpbFv+91jiA5b2sXYp2kM/jS7tlHVyBnNOn5E2cG6pxdOtMovCPisltYqWtLfK55I92g+/Z0V8+kuUVK81oeopnjXbDV4ih/052GjeBbLmrbyAwg+5f/KB/De4gH4aZ3ffPnTHmP0f+1j3ND05bvtiQd2Gu/iRZrwGOrd0dSbo6k3Qtxxe//P6QVYezNnebfiLLGNnKHT1VmyX3fGfxd1/II19DzzXFH5EMifcxgQab27ur/+ie18H8uhyyzMHqJNGu1iyJIfQ01Wrguzqfd15A0iFWk80aNXF3R54FKt00D8J5uYQnUGwr/APPrDI8fv7/oXPLr9p+yBLg/QfQ/w+MUTneF9mKwiHr+QIlAZWJ8gLlyENISJIKztHsCIZRDsO16LC+kYu0Bs+11ODq5pgPgSLyHJfWdZv/JUF62Sv5nFs9jzTaLEDpf/zHaptYKt04yG4qt4l65FV1c10HZrUyxMgtGoV/vubDSe2SPpbaZ0I/CP/NM+Y4EeEYtIwqeO9J86cTxiELGud1m4DrdQNkgzcMFb2I/xsvLxC/43+BMvQex6o/4uwfIEEqMTiMnpJ9ClQXIv2QPfM4FSaar4iZM2i+PhkYGL6LBZIJtUKh2fX8Ik3iThLi7TSds4n8QO8doKpCfeKLb/Z+5qxyX93Sv3NprPCH2pTuURvmThj5SmbNud5Ra792yVDL7EK+qSEmtAVvD34zPcH5/hXvgAPt9efKmYjr2RgXMpnGG6ed7R97eCimS5UcBwpTM0icFrcsvt94hZaUWHyFpm6WZrOafYjIhncAgFC4a5BXRd61f4tX+DByiDJnoix1dwL1Y9yjRNEnok6GPfxwndR2U1UnPt6HJ2jOyeL0b+xl3x4qlsFboaTTbfWeSUt2x7I9zdk9SvWv+OIXWR6ibX0S6DAk/4M00ii14RLPR4shZZFK6oL3zSvGtPE4PYHiYGnqVTEEJoECI1MqZQUN9AKYddN4LHFuwkOjsJ6TcWEcJq/OeE1Tb1f725v3J74K7bwzHfHXiE+41Qi3ga1cbT+J31Kz4xmqZbDg9HhPS5l89t50Jpfs53z/H6kETUhHAJ+5meY4dA4XJ6oAcb4gIInqpp4/nxAc6FqOn8nzdCVEI/zbZbBh2fgP30b9bn6S21m13vZxjlh9Cj71mqbF/vN3GWLmgkgt912AhSa3mEsdvrXriw8XgiO2W6gZ8PySaLlut4S3+MvkmWgp9h6iQv1/8jWoin7+Jo/YUAOZCsD0YKNAJrex28LV5jClhCoynBzaEe3Rw1s+ebrwz9nfrQH5O+Q38U+MTtJfRnGwQ5ou2z6DEJf4YQhZ+wC7yRQUDEt0FlQSuJ/+c+ZrZd22QyltOSf1GnuwT/u7DGi3hpuVKseX0081rYAmMzI0e2HOtXVoYuBN2jlvuhio2hbuecNWEcBXQyfU7pfoUfPU/XIfRb6LZibUXbuvzazYlVzSe2M0RN5hPBg/1mXUzP2Wp0MWrp2Ma87V0NLefbdBVtMn6t0ci+6lzHR7vGwodfjSlRgO3R0G3qF0soHmFJ+j4rLePyUGGjQ4XKfz/hUGG7n3x6xryuUEj8V7xeWHADUxPJOyuN17tVtN5ZI228gGsLRiigCUfvcfsY9oKQKWU0l6MxDhTMd2FGb+Edu25PSesgAjQI3qtL8xXrAvcxvXxnyJe2B2943j6n280ztVizf73mrizy/cCgnXl5fUZ//Lcc4GRtysU5qmDGCitVjzClv+QlXkV1rW8tqkDaqr7rtwSNlasW0B8PcUrPiLU9ZPvNx7L/cbZI99k2Es6ba1S/QqdXkJGPcU/njUa8F7Lj/RIuAMK42m/NglzYVHbfEAjubZEGZITfBGOEPC8waHy3FLTK9B61V4LMXQQJHMeg8nNctjand1T8wCx62YT0G6J1mmgcrqPZD8RDo2G/2Xte2HHfoLADrcV+t+/faRQ+kc2lsdciSX/EGYSxgK1W0wbBlxG9G6gxWB818HjxuLoaNFzJueSnaJ2jeCZZujosM97+yLerNbikAQiLx6j3hu+AOxobeLWT3/1dtN2UHrl42/RN07dNRF82p8l7uAkzGjLm1wyg6OP9GcD3a3M2ocvbcJG4AQCvR9L2dlvef8WegAWVltx9Qm/RfUKuF/TcfaJR9nQsT2vMi7KDcWDuuigwr2A/A9BsHQlejO6xyVguPU0YkHZE9/6K/2n3TGNnnkta9F+ch8WTFfIvUfdYfYAu27fd+tbtC9fElLhKNLmmIX7IjhayeVA3sHZ5v2BFr3O2KU6ML10c2D21DWZ0/86RjKSNkmgNh+5+nzwy7zAkuuKOLqhwkdt/uZ11ZoS6apqV7xm/s/7DKloytmBx5xdrB+4Qk57ebH4rY82tbGnfqe5idm3UL7Dvht50DKZcRVFchEm6ibJ4SU1EIk5ZwDKR+jPl0Gyi++ZkqCWarVfWXEabdJevnKcUbCpASimoY6VJBeS/p+YUjjfqPafIbwOnn9tAMDZwWnoqplcBc11K3Fg5UQxnvXhO0oyHCCe6K4eYuCvhSNvafHhm+1MFAPz55bCL6INDEE/zjPfWhfAEDgRu1oD/I7yVPKNb6UoSvuYgOk7vFdUckIB7BSR0MNR005SWOhpL4UCeZCru29R7G8+UeaLvhyTV7WFNBcqxA6f7pz0Ps7Mkhc97k0VKpGB3iBROOGG2hw168nXRSc0Rm7LbwZU8b7xePNPHgzR+cFvpYxytnh6tnmpvCdshBgj6MFqlWfZMP99i3uOjONdUZhFuYxbhFVmE97osAo8cb4hOTyI8IYkYIwwQu0Bu+lCDAbUAXcBvMSDu1tbMCgLb+npZX+lDDW0K7HvyvezZ+k1yTT9gmH7b7jebJILS7Udq9TGLgPEcuc1GT+GalbY1aYTmasOeiTf9Fj7Fa2oYzF19iNeP8BH+iEqHWr2MyXQiYwYg332OXmD3unLcIHgqzb7FNIHsHoJd7xdpcqDeCcx6eq44gzG4UYSUFlq8OCRFpwexT/7uef8QZg9W9AIX244BQ60BDUees/B/7uN1ymZ7D5rSrxCeV10GtpHBAYRByLMl3QT0xR8PIHvjFo3brMpJLJtC2KQpFO+NU0/kBy2BZnNPiMPI5T58dqAvNt2kMQd2S6Z1wpJTy1zFMtwIJp/ffmZrUZ/gvJz98bv14v7x+7Y4aWMZkHZz2O4Y+m9QTs8NrJsoXR0W1QHvLoevztYeDl85AIl0A5DGCHhDU1sH4Lzxd7mf8o+Ifnf8GK01TsHTvEMvQAZpY7nKsZhQichZjitwI4S7cJ1Xf0ml6GyV+MPX1hMQ8fuqJ1zZeDKRs94JfP9+lwccxO5E71DMxMDnWLXVHdlDx22e5K6ay/dBjQkst8TqFOF4tXmG+mkMoBZrYon2Csllw3VLEza7hYZCAYcIK09yTAi91KdT+kLFqzZ8WkdsFvYWcmZt608JCKz/kHp/yHHlmsLpYQHMCijQpitqIY8XoYxk/XH2h/UrJhZM8v9tYN1mcaL9cofRgpZneW0nIG9nBj23M9strWwOdcGj4/CXFne61q+wxN94VccZn8OnULX6W5hQF1KEFo5qulDb0dRLaW6Hhk7Q1VN8nZ7/ZvGFwT0xo4+ttY5u7gR3ZmODCpTWvZYZUVHnYYmR9Z/05S35/j8U07szZXp3SQP9BAgY7CF6Zza/i+yR6/cPioMir98X5qndxK6YJ3pMpncQpfuCC05faCiTJMWAg2VZg0US/tgnYQlHQnhItJ5ZE+20231KyU+ZBS1BkWY1PrMP/oZVP6QiAw9rJHjbVZrRTc4AJqQx6fFOtUs4UMpqyuvLrwtIghT8WJ6pM94Xelmb5jW2Z4JlMchrynq+o6vnW0n4EMPdSHfM78Cg8NSlvi8zQxHXG/YMN56NIVWvQt9us/Qp2nJUsXHi1W5i956fN1aixRm96mhuswTveZ1HtkiKbDvtWiMzq6vmYypsSEbdtTPb500TAc2QpA+syg8nkj4x4/yBLpRIWKSvLkOoO766mt+dC3RmtoucXu+Cc9ZmlbbxF7oX0m0BMidyGVK7L+RbodVY4VaIXopNAWgKYWPcsyqztDFKEIR5SGA7Xl8ZTgnEQ/0D8WzcxrXXBYhHf2+YFDwFUznOmoaPjxG9V6cxBzkqRWctYYH8aWOCh0EjWZIcAuRrR9YiX1xEs8zqSOxu0yX9HA5JwvL4tb5rqnMDdtBvr2bKQPry3FdMo8jIWsacl8gaMPzrZ/iLWWEcyPaavZZZYRwa/FMF76EWarVNfSPbjIFhuXHOv6Fx5SgJahklcQHMePa6UZIa6x3zURL3OEpyRVPXiZyWzGlYun9g5QViXoAbBUF/9Tc+6yDFzZcP4bo6cSDVirSjDjkV6Oe7LxdqyWgEjQ806hwJ0F92Br/pN6swh7nJ4+VU9ppwY6/JLXpN7qt6TSPfVahuDTpNrtBp4nAfrAwZVNoOZjifkee9RSdcRucfKcg81xQ6MfJIz43l8i06vb1FhHt+izQNxL4MCd4mLDfbxesiNu2WC8LUKZ+9N7FZCK+LtZfC2h8hADqjv1d4u/54iqQpKN5pXoTZA4+rsTYMUKJqIcQatU/LmpcIL+Qu4zXdMRuaSaws9NSJQZEmgdKpV+30G2kUq0ta/jGmYvbJLYXLVcy/l/DuMg0GkzBvKRHbXv1ZRhd2nFXWr9hOg6SJ0sG9BZrmdBFCxl0dOe0QXI2w7xnQe9KbKku3NL3n6QmpVH3K9q2QAz5zwq0t9wSwUQdTmhQmv6cMTTnNcUiGJY0Rst+mosFcLcJ6FpQCrob996RA254ha5luIzMPjPqG9twU5LuOmIPTwDHahCsL50PqXerbbTxsI9sE0AhFQ/qOk2gDdUNe5xT5XCZT0sRxPJBQAANWmaCRHqMSZtlF8K5D5CY8AVL6k68jP4aJjnMZlHlLc/Ml7JsPl+lS4cHTDnNUDA1Gnt27d2YNHL+ugUNO7d8EI9c1b98Qffcmr3mcj+WtcfscreMsZMSlRTFs8AUwjj+5s2FO5ZkPeXSoi5g8Q3eoczGOhnoYRwsCuln7rXi5U2Wo/j58oMdsa91HWbanrjqk3i7bL3ZVa4WyVz2dWeA7MoNIUy/6awYWFln6MjyAV7WXvpW/yL/b/uRcfpF348nl/P69dXkznn65/Db+Mp+9t/7rxeerufXt8u7rPTi7KiBM2MN+vdlem0tzvKpGQvgQR1vrR7hYsGPNb4B8rPq7PD0wZjy7cOFdlDzNg2PDekpjOJZSdRi2FiyWG6Yju9ELQwJJnVbNW+b5myzpEPF8ySJDljGdGSZtjabKB0uSdsiXDhd7VhKDDQwx0FnxmmHKayr3yO5BC2IXrnKUPr2Z4dut7SajD0y/F/62NRqHCTxkG9D2KAbAqh+YEdezy7wnZcFWWNGPPf8NDNBgj5XKxCca/n0Y05e9fqKvfbvPO3/7GD4HZNd/GALSwa5/Jpd+GM3QnEBqWIMpNOLLbYAPwvo/yLboPtoW5itEW1cptfsvsh/b0gOAMR9CbkvxHP8HH58AoPNK/YLGIBBm8MDvNv4R8vi1BlWmzbYqVjvE7dmHo/EnmVSd+m56yk+BnAaO6xjs9M6oN65LIBW1cu9pLkcQqGNejj7wry5jOWISpTAI5i4nss6jpEMKJUU/2Pf7xG7zQmsgE6nF4YruuyNHToeiatVGJZLUU/Ooa0HGVxx2BdzBJJpW8TKmJ+SDJR00H8BQ/Rx97LYRX8iu6zp+iZbHM4+C8szf2WQ2pv5ZrLTv6Ydx2OzAg0NwOjjfhxsacNKLep2j5GgGG642z1kkFYw7YKUDZBLM66HSt6zxLl3Xk2j985BY8B+eolbrQbdH4251dx2iHqCzA5gv6M4NM/o3aiDN4MG6XUaD8VWuDjRGaKowHkA9IGKqL4+cwcVlu4JmpaxDx/rLmh2hcVv2yKDkVowMlDmo1MemicGnsdzfmNIMD+pIOwYqtJD5EGVguwZMQ8p6lUGUkb20fqWn7GmdAqTCWsaM8vSJN+b/dmzGElWViWXM0Mt/Sg50Yw2sOf0odhlssWpU1NKdvf96dz7/fFN9OGTwcEWDtsYg+ln8EXH0To5FV8WELkEEbvApPZYCeMtnkdLcJdQrj2l2fKvpQhGGrn32yD+UYxGjMn2HxqyMUT//8d5a5w1ezIg+zujmgzKjMnlXftGo19f+KKfOwRLNHCwxHYM1e9vaKdgCqMyH46VRLNAmPMZsvKa+SoxpaczepnYIXhFHqwIrgZWUvKtQz+LhKbSoZm+1xgC2bP2sm6sWFOfZUwocJdYq2rJ6LlDmiu9YCJk0NQKVI6MJYf0tBBkbulqafbRwTfpaEhchI6Kw08n3/MDth7boHsisIQgRygXUGzzSWxXq1VGYbWmUvbTS3XOU0d+wjB/p3Qm8YeAp6Q0MUdTWeN/4dOP0RpjAZvWEMcPVPtnGjwe4mkqeM3pWuxCdVQpfvu86b1L8FHjZimqkcc2z3by6mud2v6W3O5A5se5/tdjJHRpWmnw5lqGUjTUPQdpNNeLHUGbkT0Oo+P0rM+AZTZxEtsgSdkQjd4g/We8kMaLc69/Qoj6Me6gPv4l1KoI7YrOVyV9s29/ZMZH6LF+eY8jJojUo/ORwzW5dDBNLq/lx/YLHMEBsjE4UPqQyMKTx4NMZeurQAa2Y6tn4jSZV8VQpREJ5OnrivRVbHDAQCks1g1UVGLdP7NHQkSpMQZtm7jR8iDJh7WoTHGma4MisCe67Su3/xCY4NKy+C0NG5WQj2G/jEyh4qHVeT667JI7BPRPH+C62+x0iKQchcG+DEL7jGuBIjeYgAPwklTq+RNlDvAZxQ/HC7gJ+arfU+F70eMFIoCjZb+NlmUgLqZMZf/0bmOvTEEO6f2bxY7KHTFni1uhCT/MWFn6fyhMPYvY5sDrh3/4si4upMiH0rU6VubqpMuIZzpS1227qDa5hKhnZ6mg6TWVharyYk9UAY/GJZta94pv78eUMjNztH6KqGMdMhbztnunr3cLUowVtxHdmLPW+kcx1x5aRUpKd0bCO7lMBjiURQHVoH/nYJ2/RPqLmTsdyY+4iXtILWKj2dbnlcZvQrlmP5oYGokrPfLLfxD9CPrwtEu4JFTLNS7R9GdcsxSEVUYrqSvnACLuXwnVIM5znaBuHwj2K+rtHbTto+axPvkj9sYJulCqkk/3uICJ7OrnTdptP6b8qsxCn8f2Y0U0b9F5zqBeug3q5p0K92q2tK3u4dv3kK9sD/gdk14ztl21QpG2DuqZtUG/k4uYWsrYN6lbaoAAnV/TyPr+kq2gV/ozXkdFgZo1NDSF+ZZUauhSHHneJw+XzI0B/kvRFk3s4GqMQadba7cTxxmM5mcS2Q7DZYRbU813U76n+O0jVy6f6fgNHbM1RHxY2C4Y8zzeA18tLfbS28foJlDGuj/zajBH+KV2z+ykKPwpaxiXVxaiF6sKQ6aL9ORp4Lo7kBtw8R8fdcs4YaeBHmNt/yLm4fcJoKq0xXS99gBugfJYODyG6L7MHaQEvl4xqgujuPnlk2zt94SUeup+sAfUl7N828F8roP9iHanXjGAEZltKDyOAFutEhYEuFs/QVgDYBLRbPzPFkjOL9V63+0UmYN070X95nuMPPdwU2chjEJXFakZi1YiGE+YDIDxcxtw3oncdhfqqhvYccpcCYbgvgTCPUJ/Xd7Kt7AJN0NXA3aKLvPq3F9k08MK1JHUNuDwdt0zv9tHLbqpQeMx31IU8ZjwtEDJBIXPRXCHuyDOAVE/DjB4BPp4gKIkqUxYcKR+9cKWdAUCHitGcDlMKVQN9t997eEwjP+VDroNRo/xqOF4hx69hu5uAVfVBbLkSfDKmuqQpRG9CU1hjan8UoZLm1nhFL+KXlHMUGjOEmlmqv674FKTXmB2OF9GmIBpB2H9Hb+xNunmOE8N80XPcNv7CE6Z4vysC21lIP3nGq+kHBRTOaJDMM+LdriVIgjCAITqJoJiSWZeXlxWphgrqqpAtNkJbedg3mK5pQVvdMyfrNO4FarPEfi7StxvNBplZ36ZLrV4Qs+ilUjsGLdnBFxr+hjwRcKp+rGPfyMO2b0CP36HnL1GcQ/312MXs1OJQJgk95PY72jRmCbDEes1Cb5o2MErFgoGLDfPdl2AM1sRNNLeFZkfbI4OzB5nEdhNuV+mymkkUxW6pTC8kFG9F/ubZfpu4xymV+u/IbmRxHljLmME2+RDfwLo+ZLuIlTKNCvj0KvB7KuDf2c50SiR/Ai08mo4yz1EFghuNEJkZWYsl5HeyRFbIhsYW8XYX8bnSorw2ALYt5kl+5gyiJ1zZJPDdPq5s0BsYa9lP3wujmlzaAj0ZKRGQgNg9CBFwtKsMmo5Xm3DDGcDc/Br0bG3ZRPcmbd+ACBnW5CtyB8EXLDaordKUFPNm/EpumEN3TjVPM+B2Pbs8O17MR9TrdCyH7UV4qxNH6HD7Gr7FxrsMKlCf5IIknKZwBaPywmkyqpAR34RRjVXIxGXZua8j15tOZHuLPII6puL1DiyhWn0SFwjxXbv3ZH1eN6mSHhLrNsoTNdKFD/YNrSwv3ZHm0sW6SxcHhlduu+0n37huw407q1y4Rnfsae9ac8fmDDxuHQPPUritJEfboYfSbqiefEdcVgnVncbaWK9phtlTtLAyovFEJra6iF5otpkzDuF31AGkK3p7Q1Z4D2PShq1gZq79RpCqKXI7Tbd2ABoTz3f6LVCVXhn9CV65d+vLiNLTRpRVn9wlYjSzsanKk6ea7utSzappvaPMc7wf7gnv9wYW5vPs6K3m2Xs3+bxurIDlBwA7DxOYAGPXMI0iaX5D7Ymj9QlIJUJc9w3qpqWgs/MGgs6EvMEOnmIpZrjcHs9YMaw4EFGX1g+YX9Kjb3T72+2ZxOjWRpNP8nTiONtFsMHpe4ZUcmBd0f3+SD851nZB8G/WLnxIop1mLlqzYVwij8540pWMpLB+PP1ieR/oWtbiOU6WFvmA3N+fhIPptB5MgIu/owFnnPK5C6tT561qt/MGDIpA8+y8Wh+IOCYiznp1oG/01Cko8S1Ht4GwTUo/ez4Nz7A7g1X0xHpIkfRV7Alu5NvxCb5pWl7E8QIDNOFXmuulwOK23Vnb3AnQx9jQMDFa7rNICNic2oCtS2g2v/kyvlJntInzRmEasB1NFaXyLfXPm6yAPgD/bjPbg5B6CH5PtVsvC89Ew6pLi7CBkqIWNVLU4oKiFr+KorbujRtQ1GKJotZT6R4E9D42YwBv3w+nzhfgf1cLy24sfpNurJm9JghOVZDvW7he5BW17sykzabajWAceUFgKSjPFWcnVvB9oE9f3mmYafVa77FtrPLc/mJPkOudjImjkgoekmjzzGKFAYs7NxWiQRoHwV+hUr3Sw7X62BsaDRusYnSqyP1S0HkAXoTHagO1N87GhwdX1WEU4QFwH4fxGTB/8hPATcEqw67AA0V/W9PlprMHtzlXoadSrmJtNxEbq2d32a/IhXhw+7fydvtP6nE/AJKjGtfnkx1IP9mBDAc76ux/df3yU01USQ/hS1iUME+m9mw3t+IvdKtqGD7LSS/c06QXPdhe70QoSruAPiUw9MQvLCltAtNqqiU4MOdCqS7KuFAKfUE2GSVSgx87gB0EA6t2+U6/sJdr25l8lz/g60htpQq1aKezdZ7dVGTaLmgcmEVJ+ALjukAxVdVm5NO60pjhfBUtaZTOK6TAkGJWv21/gUbTumzEgLySqbdqnu2dztJb234oJdFx75LoBJkIl3URfslrtk7vNVsTow1PEN+lRGbXXFIny7Dx4v3TZY+ioC/lp7KO7PZWR0bOqJc6cmka0ZqG/yLTQMprrnTJ0+yRX5wQUjjmPAEdzDOBqNFbxWnIOI33HLJ745/JK60iR8WeEeLdRmsaeRiBl4ntOz2zp+Q9b9zU89Yyweua3n2byfgvZRXEaboEFHM04sAs4a+LcLWNRWHlLqfGyO429K+rjlgfNhJH0hEA/O700TJiu2/RkpEYA3q/fPq2uuQnc2QvuooyCxyVZegIbLd3YXWkE1Y31FWXVLIIddFD8kph9QpdIQcYor4Bhu5oZA+x8yqEYTmk6TaQKdrEcPTHHdEb6U3a1k5fd7qZiY3RxmyC5UuJcVfmyYQzJG4TG+Wb2/dJmayn24/hzgZH1uO/xLgSEO++KSDeHdHkB+G+0DUuR4xWt+L1/Mvn28vr8b8ub+al0jNqwJW7mvdqG+LKoe9ZAyovwqW5PNFMr5hFAhxWRpGSmV2dR/lkKCZ1h8v8YzYViu/dQpaDy4EwTYw3kH+9SjCn3VQxFVdWzHkSlJJaWZBH+oI8Ma7Dt5trWocvQwsinnr4SZYBMXZkBmZPEhgD9u3EKNQwM7l1/mUsD2+dZ/E2Sh/3yQ99mU1jWdBzkXJWJws3BknhqBx6MatR15jYFPnc00A2ou/MYguyD7o6fVryN6Ee+ZtOe4saaCqZjuWplUpAZnHi/6q+y+ppUEhcZ+nmkAAV7wN9lMF1BA9tdI8aPUjjPcpPFFbIhMMjm7D5vdlqnhEmVenCSmxE9Njfprtwy4Z5bzgOyRCW6oIGFX4LhiK+kSUV0RvoV6wUmrcuW7jd0M57uMSBKwlyIuovhpttvKDvlu/UaszXyWJMejp0ZUaP/pSM3vVG9F336HRLRAn5M0SPXc/HrxY9FuRZUa086yvETutMtDtKtKrAAnbUlO7MDu4LxhBjtm1rTDv9rvik9OJfnUoZ2deU58GMhRQIzB/yHLS7vLGJZVXvLy9VQ1CV53pVktxrGtkWxe+BJWR+ZhGfka0tFfCcSEsoNcQQr+SGGpFpmX2+2jnF20KnzxVdJgjEL3PA3TsGdU2MMK6u52Ej/l76Hnbpz5RaeJwBrNIg46lydMWEflCZCiywV9bEaPbT9QjpYfaTG+srWHkBOxMW1PushDMoR1jFCVY+JfpXPARoMMxlspGvWRRagyuQ42qkoNWoL7geDnqezffH32XSBho8Af10+DMuGFtEFyUEfL7mHRoAVM4vJdidOorvVwT2ZF1iAGCcxPHYbuRJCMExYuIbYi8sWUDgQfT0N0i3GW2FxbVJmOWangBgUdjuN5skApGPKrwiNw3rTIPJuR+5ItfMsZgU2F9pru18GsvF5Vm64vJDWoiaxt8T0js1BAcCODI3Yfh0TD4sdzgy68e0G2oOBkC9VeeJCR2bPioqSUEcDSmIOwz4fH8qkoKYfeZu0PtnXrLMOvXiETeFqNg7M2dkZKoBeZSjUvNNeAcZ6o/8TwBMK+WXnIcFdQQh/0qWbtPErPll9CDN7Bq4LwyL6+Ke2DMAhSFrp8aPoYRP73LMW00yOOZoPCNSvPSFRvAcsTlYK/j5LuVlF70WI10y3OE3YLhzXZsM+6YDU+aHRRQ6m0qnd+QMeEPj9RLGHWFqm6wMS/dGtte+2zJrI/qA/hjSGyZtreaZICanAGlBWMfjUDZPRH6BDtBJ1wlG/UqcsA0rIycvIgbr03cUNPuzdRzNuKPgKnElzfmobTuYXNOBl3Qdzb4R7yUJKO6FBNTFJBh6/bKA5uqJbu/qiS7uVz0R6rAzZXQ5SR9oykvvY+CT2sRFBdYcBIJ7bsmC2CNjQgrqmZBuaFCXxD+jjClNm0UTRrZW1lyzNbf5msoVLuMuShyDmaPEdr+huhRZlCyf6BTKMBc5QU9IFSlpHCeLNE1gz4Z2EUEeIGAGFu13760o3L+3nvcxDYej7RYS9Cihf15G/Ecgz3vO/7xJ4x+Z6WPZ/cAXgdCkqZrQjdeiYpkdBL2PSH6SqQepH9jRROmEeofto7dQhClvTdTDrWl7ffsmmC2UjvttmLHyITgKYIHfgMjj0nSK0MzW2ihubPszhTT1imHUGFYRnVJ7s12v/9obT9hduc79EP2IOJlDTsfHODxPztdt3Hu+Pq4jfX6d3A4gtN9oeJ8mckhOQdIfrNr1wbpMl/xVv7NW6X73/Ee4fS65C4zkKt/sCUoBAyxrxJ+mU1BjqIFMgaBRwEqfTpskMjaWRG5/l8aBfjA+V8U0+S6wPhx3gaFn7dlMDrwg8qDoImH94IVZM/gk42qbwTc2nrC+S6Bx+ExTkwsPVwbKrVWYRKxuV83lhcHRPkwf0xv1KdoKwiWy4aNmwxkgjyEHYSA+XFv54wjz8dZ1mPw1j3Fr46miLnR7yOg1E+XPcMYU7jXt0M7mSZQ/I1u4GZ4jGmBDLAf7QyDT5yM2jggbXOdTNUkcWlBh3O2z6D2PUKEZ+hQmobWNadqV/2N+gTzADJHp/I3Zq26h8nXsZirfvA+K3g2glMULQ+47s4aokb3ahugdQ2xKQljfDwwEC8Esn+9F74xiMGdEcxxsWEpbRjTCp9t3F0K/iRlIj/+LoGiADBQN2NdPkDRwRrj3iSaugYreSAO1d4t5Ni7t4Ztw1zkZF4GRTn8A4zHIycmxGHTu94kQxgjRVj1gxwl8A0Gbq/AlrI/EK017huVwhFJBvE63vKheoxrdIW9wArtNMuaUxKEUM3L6VpB0/IC8DYed8/YcdtR4ZacGHUjsUMBY7DD54OQEdlAk+N4laTAqFji+E7wJfZI3HqtQPeqJ91sLErpd+lI990IbwDvRUq8SGhRrhcVa1EZIEnJnNFc0g8F9An1PInIIS1qyXf1Tq61V/3RcGhazfoXf9TdBdKVsryCdNgUmp7ZXHL+tQXkKL/a0Fg19X6KhB1aOgYYdO5iGqw0Nc4V2b4d0wvER6i6U2Q6JBp/gvUYsUzCt3xu01Apy5Ix7kOsDTS9nINS4pf72Qbn1O8gEOZ6PDBhUWnSCJjaZKDHKdbimv5dR97wrKVmZqCZ1JnsmBkN/wtLqGxFN+aDOdJkqVMI/F5aEiz19APpzlifU50lvKlyOR+y+sDQTJAXW1/MvF/+8uv969Wn8ZW7N5taXy/n155vLr9e8VNc2B6nbCm7vBbqSgNV9K2ZkM6s7zSD66h6+iDY0bqNPbE2s99btPosf42jJLrWnSHelnWqw41V5NqPNxPoRLhbsoreY5EhVq9npT6vZ8RwkjWycrtWMxdNUz5HVARDveLaBEhh8IM/xdsdpBBX9WIj7z1F9vfgq3GdwW+2Tx7TAHxczZ9OCn9k6Tw6LKEsTsxyB0Gdw/GZmg7o0gY31VSqHV8r07isorwX7+p/jVaifpjSyTqKXauBiFrMQsxFeebEjFK0Knc8bxqi2YdyVGsjESIOZCKza1oZHwz18wNoqD3WLs++ywj11UOy4sZtyaKi16rg+fr3Wal4YwX0WRgQbUd91kalCo11DwFCZ1OBSy9QDeWyw7DjOwf7dbIe2Po7ZDpWLlKwyCaF9tI4X0IOVPoqyiDmob3Z32c5Gn0jTeA8vCbh/QknAdV5REhB2DfpLdo0TBC20NwZwmBxS5NVDinBjvO3o7DsNS4Sly6ESwFrhlrVAby7uj1mCzGn9AXh+D8niUGDu7ZPSBceze8IwzxX+XH2VA59S5nBMdNeqZQ5cX+eASRA5VrgGBZok3XCohFka47ij/nEGeVXW7bcq6zj+8A2KsnmLGffZYm639VT8k6vBP8mC4l0Km0ZGGoCgfJV0XQfZhI3wdPrcvOMgexj0Ogbo8UyxJue5BaoYfkeYJTTYtw3QZq1dD8AXz2So1HgZboDyu4mJQMPe6GBQHepZs3DG7ixfzGhgmDJeWWOQDTFMZzA2Smf4SpEVsqXoFV6gsBRKIo7TkeWoooyua9HfQT3/BmKEMNkBjmaRrjY0U4q2ZnVDbBtMuLSUDbkImXq+aqLDDkjNdtNOogQeie/zM72dCkFDWZ69601q9A4bQ6oLGhwqLAmfF7ssSnd5Q0OyUagH6cQrHBUDTfSvEkqslTUFgBbiw0Kodq7ufSmAhW2LxrBn9AN4YnxUQicRHb9otiEQ6R13OlZm2BW0g00M4Q7I8fuEOzBwpJze3JtiIyv20ScekjdCRk6QUi2iUfXmkEF8JYyP/ZWWntsORxhhsbQJHwDXkcbyOROilqovPf/njfUflnDaWs0WOMCiF+s5zVawarz+H1LrEHOHKitUFOD3Fh49fOJrre7P6nIsuRI4s1wBeJFXkfnVEPGOCz3pAJB7fIyVr5xZ38JtxI6ZWRJr9II7UHZLQgunDYs6tknfuJ3wSx4cfNWRR2+jZHYDuAZ5eP0fUV63JzneHKZ76jicBMij9u6iu1+OUaV2BhLvrnJ953hzlVrVrq4Jb2vFqvkXauzvEh3ays3rd2nQM0/gc/uF7vxIIYc45Co95L98uE7XdD+l280zXMKzONf0I/S+/Wh9zfj3wSxF/j2ahv1IS6dV8zh6qDdn0yrsGxTrX34raMA+yWWwu/hHCKwmD+Haqq3QdWT9wiOXmNAwMkPlxXnAAxMVgo3/itcQfh8Y5d4iXVKfBg2bzzSW94guoKkPvqmZfQc0eQPc7W1AzczG5oqWOrkdQ0GXc4BdCDUtZF7QardUUAeqLI2UehaDNTwlIX2cVShIvyJdlWMgksUaRbZ4REPGnlFy4wmRAodZDBxca3i9kIWyOJ0zdIpcWAVg2nKwULjpUB5vfxAjRqwyQCc9wpHNXnZreF6GQr5MhPFyWDHk1CJKEgZItcg7KzpEFiCRt9KcbedN3mZ85yDjG4OkEBnWS/dIyikTOPF65eajrvu9dXn93rqfalRfdSNCNWY33+GlHdEqyp6i9eIg3eZM4wmh3jSecBA4Q+8VGk+Vbcuiohp1sU36yDBzFn+xfLapxFDlL/msRFOJfHSdgqWap2h+z6VVlbdL998nhYfuS8b20upkWCUO2voN1ekQebmKgx7ZS+vXZRw+rVMQr/lbdS4HOc1zOTVUIWwqpypaiQyHi3BAA1TSw1ROOffiaqRlzC7uwO7r4p7UEf2l+XxTkppUzkzs654uIXvyD6WlGgLHFxBk/mThfU00KZClVSvQl9dWuF5a1zQU/xHmMX9htu+5Qz/orm5SY8evl9eDVfm72aYW+SWs/6T/L6QtgSZtIbqshXTjoxUerM9MpQyYcEPAJGI7ziHlbEIO6gx3PIMmRrdC8XwutzEmMXw7jVe2O8t+zwlIKre5/R6Iir/SIEqYPOlyEMD+zge1agZ8N0MSVrwjRCaMqe9wLMkF+rDVqIyAPf9tygglgNvRQn9m8esIVdptN+57AmubI2vDbHfpsYao73s6J77g7pXD3HFUrfv/zBxCqzGmswVXRVlLrF8AQIAXJHLtJ73+VNXS8fTUT5cLTtUvnJcnyszD67M1YGZml9RjPJev5iorNA+djk6KmN7RRvuxrdA6aQHLGUty9mheAV5xteAVq35IpwOSBXveGzBcupOZrdAJJuB6RJIUwe24kn5bo5FNUc3XDJxioRZbx44FrWAJYgrSJafy9mDP7lljecYnnUQ4xePO2uTdCmvwX+ldk3UhmgAIyO28etuQUe/0XbMxcWXqKUbJyIfImQYIvSVZhnPYHLJcIA+4/w8J+zMAOXcx3KCDeyNUXs3jNHGnfXmO6CHji+rI05CqnHcyUgATx3sbTJOjC16JMabJzMjuTNZ0F39SR7WewmSZhbAl7oEH8xl69LpgT5PWEgXXMrIbJzYg7/hoOTUfd1kj85SxwmNx7JSSr0tGzSArI35WRnZcdVifWQmvct9j0wvfdR0DAFPrjV+GdY6OqpPFMD1kgy4i/QZ/E5t8V6aKc/waaH9mDzy2D4YuiFr9EEcM2xRgdMOF2LWdZk4SRy48Xkw/O5Z3LIfRqPO7Ikx8uQIB2vgB+NfNo1XVqHZ9VHHF/NJ9kXcvatq9xpvXdnqMVpnICpJUz6/SrEBfnSKz0m5i9UWqiwkOADRgpIh/Fm2T4w+xemFNDNi7oeKyshjM32m2fC27/GJm/kv6klYHYzpMMlLbTCYZi5V2bKXKLOPfbcLxP9VuzbcwoR6IhtE/4qxRJal6gq8rxjmBNxyJPh4HWlekLpfHH1VCSUWZuXqzo3dSJ0/x+10iEqMxhg4RyQ19tzM5YLoImVRuvMznH5C2X0C6WokbOqb3/5xe/GZVF+XZSSnUrHj3182C4FNnQa5nl2dKdb3samBNV+OU+KPVQl1PA0kdjWMoJ8tRgkwFi/JZg5zjDmh8BwKL1X8cjJMnFtqbBXkODUqlrOrkKE8hmpuCDukzzBVyHnlrYk05SrlQMNlqivHV/SomVjUGN56sNWcXE7Nn0EA/l2FL43VIFwQtBakHpFVAF+70GsPaL/V8TZbx8DbWVJlZv99Tp5DF9C2uO/SlJCBMjVVeC2Cnsp7lnf3xe3F/T+vGHee7OIkeq+yiRnOj2EFOc1/YUaZGea0z71dyCgi++xB9d/IOfKZvdgHyOY8w7Q0lWdDLAU/JRUugLqbZg1qLDWpO33KhHqWMU+bBufqCK0Jm6W3E8vlFVYjBaM4HO/ZoSEivAyh4MpU5IOZbelFuMn51rkq3RB+gUPdCWjklYd9Kp9zGw8DvfpoEM1iPbRNGdLNQO6QpY7hIL+Sy/WeaKv0MN8Uu7nCJssuwYjBW2MOaCiT5VSovmwukj+VOyOf1MlxvGUbHSBEdY29k0pluUUTPtdKwrJVGQHN6nKSR9SPKwoHFEHKfYUBl8E1x89oi2qlmd1IVGM2+y56LRorbMAMnhzyWM49XCc1Gy09jcASkpbvnmCutIZYCaFzcSOggC8/jDpHXeTOPp/P4bHw3qa+sck4GonIyQCAQJ0exDrN8H7tmpAzyaiICpjSS/BsbCfHWRIbpdI63Tq6rYerb/J5CLhpyz5R5IbCRpkE/VVSlxYCt0LiO1XqKsJexrp5SY73+xr78xgqYzweaMOxgyo1RoVTnifIRV69pxBVr2Z6Dzm/49Q5EGRmX/cUP0aGYOgnkkqHjdr/y5OUlf1YDKvn18vpvwtw+qc/V6HOuFk/WPb1PD2YJGzJBMqgrbmFF61dqGaspbGlydH55M/7N+n+wtQjXoCuWbc8OGasI/hElyQd6+uitmCQRPR8CucJIT64At9VOpIZ5TYu09ZmN2EwZqxH+U1mNMKIxuO8S8jpeo0khzSWRyjAdbf89vQssGngDcgYe44GG7eE6Svdbc3A/UstVbZLepR0ECDSsbXV9juU4Vzg5LmgetLZoBpHFkEVa5+k6XKY7SOLXQMtmx/tu42HNpmstpz9JDZ+es7VKEfJPSoXgioZbOR+IZxvX+ZHBQIe8kkX9OxGGObBmmANGD04e5rBH/jBAo1eMc8Dy/HPGKvfKXbxLl1k5r0S0U3+a+9wOWtQPKhfica3yAq7O196w9yjVJb48x6D2DjAqBHNI3yqA7JMwwrbfUvWTMMIVA3DV1pwK39dAbln1dTBNl4xPxbFNYbU2zQJwd5+UL1R1STAOw4ircf1FN5hk6eqZsYsNzvchtM5plLc2g9nWmHm61FGNntiJLXH60EPyFoIXpDlUE/Zjl2Ctb0Opp1P51T/tgeAlV/vFHbuhshtSDXW1rrJYMCwXtH6F3/c3KY4ngPVF5CMjMUWuUKFA9RUKclKBosZ6bFShAA8ORrIMgj6BMNSFOw114RMq7K3bQzfUhWtJikqXUH25lbMPZQAZVHPNuY5l+Q5DOD48iN2L0NtUSVorPk1648B8Vnq5yHKh0EEf95E/LaQORk+BRiPS61O4mhQkt7P8WPicWuVJsFbO9W0tL9i8hQyjJ7FUNPJ6RtyWdRhcX4c5jcHTyM661eooPEtTvX8rU0UjS6pB/PZUgzWmY7eFatD5AHcLEA5ityAcLIsXqL54QV5du0Aj4g5d95TSBTGqXJSFXNSxkEv6rOMaP2ZjHXeMmNiZODWaLODqeTqioRB3fNaxAsm4/xj7Ex9ps35CuV3TYkFYs7UwHrqd2xXX9FVCs4zNlUQgwfaROrcjZxGUUR0Z2f2YhD9DIP2RQn8tgbUYbhmayHM+ZdWibYmlvv+cXiyVnK+enqJD/xKN7JGM7zu1gcklXAXCx5A6Emu5t275CGCW7q0wi3fxY7yIo8T61YNR0r9xtatc9IqpMWSRRfMw/kMrJvn6K3VoluGF2c/8YHlhBm9wYQajni9MCL4VauS7eMsRLzyC6hp9Ayr8Zi7C7FEQmMDs8zBcNiCfp5jKWcJR2c98mAIF3qjHWY+p3GP9N7FtRj/gsYTlXMD3rSDuLCUmibbnQE600DyN5aU+VF/qa8TS6d4iwTKUrlOxj+PoiLUIN1sBHk+0IDpjeDwKTFTLDLSMYJRTqlJBFZ11BNKXCkOpxWSG2T9Tm1aS7kKn92vCBMneVsWAM6a9UAbGAhMkH/SS6r43aVaZ83JMTxRNR3FvHFq8quprmy4DVRPkTzbyGl6hLMBCjfw9YtO01DHxUABXryCBFUlXT0V+4MmU2vpjxQSMeAQgESvOaaZA03EhH65w3VvOOwtmptc/0+JfLM8MlYpgdr1nRs1PCiRszFskEWtWcV1Ls0/dd0h/7G7nSr4wzdLVU3JYHCmyIEDCf5mRgKuayXQb8+3mCGgaVMIl+rdpAkHgc8ZPVBf0weU1xCL3X+/O559vqg+ByNCVOuX6of/L6zP4Fb9ZF8PNAQpMAPOiCYEFqW+4OdDN8Xu8s178gsp0UsNzv13R4xZtt4LhWhJTCOut/7CuBZsdY2qv6rKS5gSZKPgJerZ2WbRT0WDaOGBWtY9mBN1zxHvgm6f7wWILpgznIdiX4wOdvvGByAt614rhMnd2/zJ37caaxvk8RA2UMmqtuEgXf9BqYWd/AEUOhVhAKXL0V9XwvGBoALVsrmrcsg6rpHQ3Th6idU4u5r6TiK86CEvXmdiAs68ux3aALBU0nynTMzQUA3GDQmC8q6RC1UYj9s98Pa2YAgiuKVYeqeoHR9Jla9DKm+JrLDbpSHRMWcZyUX/883hVHUU1oDpj6ALIyOvuAvSLHr0BIyGpENYw7np5X+RueW1EW4+IT8ylLRunfzyVg+k2S3/ET3EiJ1tCUcLTF9jASNOLVFqzepxw43EyPU3EQ8MRMT9OGmkSBL0FYV9+oU4TkPliv2xg3QJVNdOclb6yOITrdJE+hEnYYWBAftded5RNMTNQIkTKTEHasfSonE/kwHG8in9nKSH7hMR0pmOnu8ZaZLcMjVSWrUJEnDEj3BdZoleQ+j7x8Eq/cx19OU2hW0CE+Abw+4oJ9fQEtwWWjYjz6hWSzGM0g2z9fabb3a7dI0smm66XW8KXoGmzrpAAX+TzGLclPzCU4cz5KE0ewyDmIZxHWggbADocrdIse6aB8ZY6H0DxMc60rSbI0RTnCE3WcGci+YuwXLRCXHVW5atCtkpmc7lNd8V1EVlYmzkIl4awiRUt51dkDjSrmXyX4URziBO2Scn8qml8CMA8TUhGbKRM6Ohxb7DwlmlHg1urVj4moD0n1xK/ZOEa6kdH+gLhbXaxzx35EhGm3jp5OYA1CnSY4+8K/WERfpmdGLcNhGCSJUyVcdZpuF5GZZ5gngq4IzzEntNTLsATGel+kk3sMZHp1/qyb+W/QT7r+l6/+WwJqCFieXOXs3Rm1OWvGYarCqLhzWWZABXDP5r1Cl2aoPWAgZzSGJzFuUL3Ld3uf4TbHdBfSG61ui8kAs+qcdg2adsr63200vWC3pJhnBwYBM5l/1j1tvz+cuWkEiDm1jLmU0TWIB+MO1TVG7pcXy6CaUib9HmD5QXnTvUPbWm5enu59ELodejmmslfKcPQvK3AMXpEHjkWCvY6Muo6U/VN++v9dkHjvSxKwpdwvROa92XWK3u5O5pBQAb6UgTi1nsO5TAPxt3WaQe/undfjpkvfaEfhOjW/4AYtyz9r/vR4oE7x5yWhXzUVMg3T+Ncu8+qmHLC7sPF/iGmoXxiYVdAEmGvBF1tO9wj0/Hd7FLo0zv+aNg6EyPEY822/Jov8bci4/hPem18hJYqdSRsoEl+UFz/oO5I+OUE/YUPWmdL5UFFD5k/llP/WCNfRILZzl/4XLXGaB7sto6z4fZAQ8pdlMhY8y7FTjBc9vVeQ7WT2hHRw0KjBMZ0UZ1v9cbKoT5qMX2Ou8yJ8daRmG07vkHhsFU08o4R6Uq05yL5ryu5+Q6ziMgh3Zkl7qLtpqI0ILv4vHwU1JSPGCa8ZDqA6wjm2wabsorEvuFZ+A7z8pHjBkPH6bt+BOTQM3nesBjp+GAVArPvrOt0v3v+I9w+HznZpKJHB/Jo5DjkTcijS8oEfDJlghAdqHbqM8cKZ4Lc0r8BWKQM0PoET7Q+JIyNTdrU2pmok62D4a1yRc5rzKa24GUFOqGEJ/7Ztup14X7eX6GGoL4+h9cFql4B5BvShzCjbjrMa9BQwsAdzHW6bkq9sQxhLZoAA+6A32RpATnyHLtyqT5arumGTTidmGkU1XpyOkdR5dAe0epkAIJ9s432yzQqQbKIg5DlARjrOjQa6UMOcocjvw+tjBua787k/u01L86GKy51ZciGhhw7kLo2zVSRx+XoNoAN4OWUDLXUtmuufmYPc9taEce6d0iTFw/3xcDAy0jNzWbTOgwejQzqmO3o/0IMA2Q+LmSwxrc4WYK+xmZXTAkZqHwIRgfDkUG0qK4LQwofrcnNp9/o3z9w50SfpMzCxLARHkbdwNGOZgKrEnI5sJRlzJ/LHroIdcf6iyYMdM+J1U/Js4UHw7UPFhDNg538uZ36fMySVzyfU/t8uY7kX/rJcRvMH23G7i4sE4vSu24dspRBymY61YVw4A8d5DQzeAoXWGVB4GijBgdsrOkYuZ7briqTPdvvQmAIKXWyjTnwqa1k6NmkNxb8cqZt1CcNEjUTmUiVtdAgXdH4WYWaAFp/u0nLGRkNHaKr3b2qjbgVXqCsalV48r7L3nJO45s1jVl8+90Hi755bGNv+M4MBYe9oG9VgUAdICksHVym5kC93k0sg0CdWBpnmC6pG7ApdUOdyYbUDaWgG2kSdGPdFw6ClBoY/CvWUe/N8AEIDVT7kXT7NFGEBGrstQa3Uszt2PIMNIQUZm0Z7PYycz5loD6pJDljswsMYxBoY8Tq672Z317cfYZqUl5Igz/KDhhkIkfSpec0s6NJwW2dHz6nMbkCsb6mCRnT0QJeUNtmP/JA1DikAwcMtZsm9N6oMXmwJWofaf18+WNLQRr3uN/EBWWVc3351byRgPFoGNi4M65HXE8KGzhvhZ2TWgiw1frZ3KBn1CpGZOgHbl+4Va4DLx3ULyJLmFnSjm2vJ2kKYCtWOEG/UC/4EgOPg+SftWzFdNPBiQPiYcFMmbi0nbW4XLvUyb6xfbXUfBOBQOcTc3FVziL7fWN27uvep2KofsjG9pUibjm6j/6E0X0UBDTPf93svhUu98nuGO7KlDbxugx3Twp2EY3Q+hR88mjcQ5QdwI/PYBLCUEWoxyh7mhfpe31jlHlEHvTKMoF8t/mwG7FMgJFTmRaT/tgefmpFD5NPCg3hmiJSJ3sVOWa/yd5i4ctvhdJn7vEnE7lxc8vmlFebjIGE6ZEKRZ5KUz+PHLe/6YRCN1p+sVzMGdsnNejpVSQN/PjtwtF84hROeUELJdjrLwWQhv8akIZgau8zP7OZrEpGn5G6T861OSSGI7SqhURLkMgXCl+KlartA+A4vWrQoyggDVzSBf4fsrNNFjLMw35lPWTpqqwk0G8waz6qj3Fy87HMhPy/pGSP7MBAwKohFTq3nQslRGAKEVwutQOXpnyB0YjAMfAN8OMg+1kuKRPDq1HWjMao0QYaCjmYvrXPRHSvUbVV38LNaR/K1fP+AhGCAdxnMGDTYMDpr/BFr9ipgj7mcMRtBNwFtRJftbr3nWS+qP0tfkPitCxtAVNIjkgoz9u5nCveRI8pfBASu5pxYaTVzE690THMVSCF+4P1oZnw8BAT68mCHrr1FHURKxWuXNsLzIT/jnqlOpbyHB1Z3Q0MDJmagyFV65qK4uwd/AhZpr0rWCKO3EVwmJSk9UsEc9dCqKo9RoJhbo/x9A3DxUvH/Bp8OUunHD2Pe0dpGGqvPQw67kVeni2XBwx/bkHRjFVSq39B+MAa3v7wJMpKRMN0c6qc47LiHLU3mchtznEpFIbsd2Zpie16Q8/tN5JSJ0rY5Q18ThEot56cnNguap5GrAaouiXrKPvKPYpq9yjqYY8qphvtUXTco5CsjGUfJCcr1iBeL8MNl+g6qSxlO6OhrUC1XzVDwXMsp2OOdZLNwbC3sYlS4x01abzjE1QzkY2DV2u844K3nO5j+rVcOQqr2nVHwj2aKvB5OQs9nSEt7E1HtWdjX55k9bQpzfWTtS2Wo0ZGi4SexW0cfqwKkt7XtWl5ZJJmu7hQOThd0wDgYX2PKelmgIIT9qtNhiOM+zpid/A6ZY6NSoI1AJYTeltLHBsdBFSZpZgYyv9usmgj8udxaL4gu5UdjgPu2lHUTpcDTUyau4uV26FpWYW06opRbCgf/SZaxLvwoQCYGDuBGnsRbvUDx2UjkAK1ysN/zdI+WVg9idbLjKk+MKIlsXAlDJmQmgr73RchjLFNwhj6O87gF/xmFUaEu5J6dq5oFma7aLtnTKFefttal9fWe3Aedfdul9q1am/T+2Wl64oZdOH/tgfFK4SrCQviyrDCvGyaq0Lepuku3DJfG9N/W0ZGuOJRG/HKibDisvhG9JRcFv5gKAw76pl1hyYvn1gC4wqt2VVJF6lEjNocsGKkZ1B0F/I/LcifUQZiaUb9c1aO01cITYxoA41M7cwaeMNGz6QGLMMS7AQukA4I45HnmAOM2TIluLiY5VAsmsTbcJEuD2Z6P8BiTPDrxH7KsBRrwlLy6sr/iKBh0GMcnUMYnHoIg9sXhGHk+tTsnvEL5ZC8Wz8kz50RKYluTpvpH7lu563aPDJfzvEi3RwvMR7jHdEshRhg8+TlPsLFmEVbkATdgV7YcahXmlQrjHd0xiP7z7UeSUPIMtMD+a6OeZVXaqljEQzdmqpr3LJVdBXXkeMOiQFY/h9TNi3hVUutUzuYKjiXT1EXr1GPlVONanrH0kq5BksJ266+43PbmTH34YkETE/RNm8l4AbGW60emJm999GL9Zxmq7QExpZVTGc2kWPYm3QVgXkge0z+MtNKFgWvlf3bPHlRbXwli8IVGyWTcWKHBaNHSBl9yXX6GO3ipDF5cbUFrZpT0wqKFdcnBfEWh+3KiKVwAZnqar8VBVI6YXYNP3JmnLRgQbRLpHy12q8cWJcPe06524Fpt2Ke7TdjLIRUmtHsKk3IctxFD1IzDVOMsjtxqSaf7tDkQ1Iuu6SOhr69Cc2xV5uEnjhVXVlw3fUBaTAyqO3AkmePfOcdX2alZ5oTfDkagi/SD7+XSvP/KnovZjbypC2aRU/hQ5hwccdOZotnOwhGw5GMXW3qnB1XPDun/wWJPSBtcGoBrPn94/Z6/4DJBmDbhgvInSq983HyuN+mW/Ui79iTNjSuwywGZ4L3mpjg3TomeIZdPQ1pFQReZ+I/kRL+8ptubnvKSuxERio8pS98wNCmt+kZqt6lHZRMgsDt7hbyxYq1VgUvnSAU6DYJBSJjocAgQEPUu8CC3HFLoAGUo55PgoMFPhn6Xle+FmW9Eg4mErcwtLUrmO38W5qN3Eo4BZmUQo0xpZt7uzvkcKTxmukZ0fDB7ZJSyUajIe7sIMSF3ad8VhNZ23DxHO0s9MHjr7tmWNP2xjMknbhrCH2eYFzmyaj+Enh+G6KxXWsZaMtlaqRcDTrP9walcvQC1KIHx4xwUVCaFnpT02fD5K/mEbxGuM1dvH6iPizhtohZYIkVQvVYoSaoELHNsULGxh+xQu8Uy78r0aT0MVjUSY7eyTn3n2n+TThNuBGC8ddQU5JBRGVkblXr2kJwjrtu6g7umUXp4Jh3+4eoWlscN9YWLQZttCq9eugJGYfugecMsdGwUQv1OQTD0oUCeh2HxGKyHZx2852mG6OJ1j00tLu/2PlilwJH6jamiS59MWDpLotCJmp0zMelQcrb+Mcq4vl4uDoRBh2ANl7X+6NmwdaLryw5o24l5xP2g0sDeLe3/cBHDVRwNIR3sWEU6RITuD5fQULrc9CgQqgCo1hPyZ7e5PQmg2rgENgNz1zAQ+dB8CHZiN9Gv2OhozHS7WO3O65b7CjJ4MdpUVESrgtAZ7JrXWYE7BL6ut2xo+VCR8A2Q+aqUj5iniH2DY3bhoHbb9sQ4PFT2fl/gtQO0kmAfZXjkNNws90nUQf8u+QQTExmGHgJ+AAjeTOl4V1i9gzn8YLeuKrLZBfXJ7un4K8DR56Wf02qO2Yad1Jp8GsGnW0ovD6nR00p266j2RMOdD0ON8D0wu9c6riINnQnPsRJHGaH+hIH1A2/Y0mq+ByIlrf8xOWUKNR8gEXE0BhxbT0wQ7iqvt6dQ1FJ8PbKA3h2Sz2xihym18i4MUo5oeUZYBqQ9HcBwRu9UDLcKIkWNHl/kO95/csTj3UXG4XhfnHB+nknkeYHXYwdW6b54aR2EyKLBmg59Kr1Q+TK0w2jRm4fXlfNi1yjClw8L8MhbRlOHm3VFuIq1tl+M52HrgbHrsaqzAWbwpVlAg+8i6hD1fgam1BngurzSwlRoaJA/KVAjYVrqbGg7Lp4MmLDUs08nQqrvLhd9eIGDJbRfd32+oxZRYBMUKCRyecnINO3Hqnfyf+iMRP/SWYC+YlfmxBfHDb55AzXVzIjankzS/G/vaXIHn9SlaDKApyRMlHfxlGXdq4QMdETveLMBKRbT0K6YWyTufTGzvMnhXqFC6aK/P1mqZiBY+zUcy5nkkf9zSSrRp40kuzD37YVW13sIOTbwfziHyIucjy7mN9ffptbX8a386O57tFc+qMn7sjArnKcLfMt+YP+NFx+nsWuv21pHrHJHNh6vTrz5lfj+y+XU2syvpmNz/86S0upNKddKg2ZblG/rcPbrpSGiKKURhQONmeqwI7mu+cYCJeoSeEyBuSJeKV3a0v6Izy0R0bc3TovwFl4sVw1omaykEOc52xiVeg4seT7ruHA0tGYIxSrvKCkWG7OM0zDG8n3ey8ezKayWNN1Cnzn+9/D7U5KfbT1gmrq43u2PCXT9HnLq+XMhtqcYmr7NfN/OwDVpDmR8MBi+wDyfd6qbKgo+7qnIJ7B1QUc5x8eknTxe5TpM7ip3Oy7jYGNY8EKp50SuCb7GgigOIRmu99s6K7bpZC3I/tMEExUBX5FESxkLILlEzL0UF8iWEwVRxaXOurfGJGpgWW9cKn54wmRQuQJpwKDvfhpvwoZVzyUs606vixveMrOdD2DVmn7ziwfA7c8BptP/Pd9jhJFh+tRdMgYRee7RGF68ExwdOiIo5sqllGPdTz7HcXqq7Y5ndEo9UL1V4wdQXakQD6RpVumi9iJi6w60uW71PVjUw6yfE02MFJMGNXORD5Fa4hfuAaT847RpZgpF/iOa6AeaDRilN/x6PV3fKuNxupLePxdFpWcwmR4FsHQ1rdwGyWCD+3QJT7NytousRDr+33H+h0M7R7sf4NpIuQrjNjPh4Q6w3QdMy2rI6vIu2LfWNsCNSDu2W/HB/hmn/oAlaP+mRHgwf+Fm2jPdWSqYmFjhKZTpZMJ4JFdRoN8Yz0T33GGDiI9FGQhyb+TL/gqvJw3DSxrsEjCH/skXBcDK5403t8xs/aRPyQy/T7uUgKQuh6ss42RQPeG6ujeZNJtE0M747mOFG9stfoufIXsc8amZrHCXv+cHrZhEq0bIj8hvKpGfggNjVK+IvQTnFgZngQzhSRpFifxDiiVFIyDdrREuNiRPXRlktcmC+l61eWs+7umrkdxgeHWC8xYese332ZEtkxPXG16cmYaodj2KyMUoUbl9VKjan1/BpdBiW5qgLEINCQnQZu8kWOQiLbhmpDKvQ6BKP3FPABrVDmr/5y9QDYQd5TApr6pykGjQNiuPkypXXvw/sB2J7ZuhAtKX4L2Ajx0HBQYhCprrr4mTj1PYOpZzt/v4h8hPSLhQ8jdpEIwK8AA9R7e83zjcR15bQF/iWvxl6iZq42YIxhVwzviL8EUeYDwXGl4jsWx48Fl9SwRY7dv9pqn04mOHxeN53U35iGhz0bjkkQvDKA5Qophr4v4L2w0UW6gzy+HHVChsY1gfb2sK41eHG29EGxt/tCdxg+9WJgue/m12qa9YBXdUc1FWWG/sfXi5oK1uOObbdBYq1k511orZx/dVgVp89nH1k/ftA9V8owizdyePaTB9NNgvKVP+1AAiOwh3TCgXlChGKWXJfAPGjEj1jwObgCVVOw6CwuexI+WWzvsRdTr6y7elgwZuDs53XFErdGbtbPUyeszO4HhQQIUTUFlLt1kSrFXsNPr+EobHZe0UmGRPIQ43y6q32m9qVHSYtwsIfMnfWf+7V61e6wHUlH/UG5Wehk9JfFjkv6MhQBg20kgygrXS+uanuUfYe7XSsNNBlFVK+TKOQ36HOmzH2dPB4YOfWdEg+W5bvfKeQMHlszmfBSooPf6KZIYnouHrhf4nfNkZclCRbthWmv6yZZAjZ+S/aKCQyAntfg9xx0iuQ4RaC2vW1Pyl7wiEVSyfEC6KpM5t9l+vd9a40dqB3W9kTWll1/0YjWBPfTAV8+xeyRNLbnSsI4r7Ug34BnfuI7dua0urMnuXLaegIizBAwIhwNgHRwAla155y3t1qwpRQ21eICSzAPVk3kQQy4PD/tD9yQyDyJO1lssJhepvr2x0g2mJyR9gS5idMJVi90h8UcGXOTyatW2Ota01UnHrroyPu9hbDJJLa0JvfUz+vua+uucgUTyFEtevSqgseiEAXUP20PfJz1NqBNZ2jBcbfcJ77fYQ5fmiq/XEPSw3SsLdB6CIV0IhsxDMDQauu5pMRgqQ0My/iSrvfwLrhhgMXiOX7imVVcRXon5GSzsPKtR8D6Liy+jTbrjlrpqpfoq+pEu002YF2YHU9BzZv/QwPzgCkRgJU+FbLpnwJ3Of81vVmHPWWmIEIOhmhhMVJHuEoUh7/VR2ESptnx5jit00V1jmapZaBg4dq8NlfLFOTUvbjAvNDIMXx965esrE2ysSbAx10gTE2xeUBDy6+swiQBaYZZg24Ey4N9Thl1W4VB9FY70UISz8XDkGQCDtHU4ug+/K2qn19EO2F5zLsoTNrFtt0w/+4JOr7xYHlcd2dU/Wtf3oCP9wS9RjN8Vrie6HcIlTG2W5S4TpnWhOKfaTxrZ3Mq1yxXzsZy8XzCRi97iQzPJmrrCV5fOgcm7FjoHFQajY9h6x0Z6iFxCjtfp5vnAmwiYDZxryokd1IBqLNZPqYhLAwlMpfbJRqPkfXB7gELSSxmw41WXjrxoXteWWzGGVX7sOo6asqng1jUVip7CwLpM11vriHMQ2BFO6i0oD9Ott3AWgyFScwEELaVk8iotxOrtfgUtyYgMR77t96cbMFPGEy+zh4J23XkD892+BheBn+ucvnlkiyxt9CQX16LkozsQ8ZKRqzIpkKbYHNh4+aol8V45YYtgtr8G0LNIH8KEHRGfQd8HFVoxn7V2Px2nb9lNSO/2gzGChoxo1GSTXkYaXTVtVxIj4zSo1sDXZEJTBtR05L75egdKpLm3dBtQHTq4Jgl8ibeiqXx/eQ2AfXVRXgA8XtWe/D4hM2EXJqnr+HYcLlBtbRq8LYClwtJF8nYhc4Z/TsL1z7CGdrVr0mb2IvOsTV6VlfexS8+lNx+PxXtuHa41lfz6iFc16BWVfPqpKsNk42S/4tg95kYZPocllIOjeIj0FWtwzyg1Yg1+R/ckeIhGBgPMNOwBkFv9oH2ZETsNGfEkWm9BBosxGHVMikWDbe/0NLiyeonxkBoR9MXTTRNuN3CfgXsuBM9KnAL1uNaM5nm/w2GzaJKa7uFm74YFsf7DEpwEjeDwa9EgNHu4U6A/lZYCRInW9eXXgSVHy6cgFYnvD52TehsDcfkOxDzjNkSTaTThu0O7P32XvK6KNHVV99S6qqGZ0ppd6qqci1wKo6kLKanB3W5U5FWbaYBsUFeXF8tZ84D0JiipsrX2T2w8UbBFcxqKbA9HNuNmAJTGJXp+65RFtR8jLwlhkIDQIhqEFq5t0HaEaBGPDJFrhGpVYFpFe2BKM6oxcWRIa7qFYHOarh7iNesIbzvIKcpW2i1MKAaKitQlXY+RVCQEx/7I2488MDYrr6gG6jNUdamP1uXsfmyRpfUcPz1/WEICun1Os12ertLNStig81HG3K2XMTfSCCRe71rLHE1EWpnUjVubrbaeAiZS/NbVgSUU40W0OyTbklgDUmkNwW6HuqZq+imKqqx5iRTwTQLMJCZSBMIlRvxhMPJMIDjieo1ocfCr0im73y/o2XzUAl51b5B4Q2/UU3DJdVYkh3+bcZmVhkgAa1+iEZC5k5pKCVv3XjNVK/lQYg9bkSRCfV1Yr54VCYTKiiiLyVPLfnWcbODVJvS32KeJrILdNjZN25R1qxUnVF9xYqTqfVacXGfouKQ3ssaaNuEs3BxBnRqHqkmADI0TVsoN5Ceda+tiWVs32dAXtzo1VTe0rpTWzVcVlVFyNjy3ng3vWWBYOIkInriYOtJ+qeA/NVHqyAIAXTMuaqeNuiKylPVKH1CHyspp/dSB+zIVPqdpb7jSO3+i2ajUb/mkTyI/FgHK/Z/PZUx2GqNNnZ0Nyay0XAHTQ67MaAOFXMVY1tM+uZ9C3Dfop3AFxUChTgPFAEk1r8NAGHGCNxkIg7E7xRl8g7FS+pSJhV1TQSfimGiqSys1D9zNCrK8RhVdTcjvd7T21SE/l6xwmyQrcJ1khWucY7XZrhWqoOufsdddKoTL2MZrBTUyha5scUYrzDdWjhoyn4dQ7W8Q2qiuLknXMfki+Z0fadG1IxsdZhuNtod2SRFRwmEBbi0sgLwSFdD6SttBAceWA57M5TJLFfuiRQR0SQFN3usxB1TEoRhsSAbbyXT+QKRvBBwysq4WOAREyGNZPfccHHJRNzbkQm41qTMXMpciV7UbmIMKwYf+RYa5GEEXCc1vvkph3ucvXz7fWBf/nN19vr24vJrDhjvXNJZOtLJ6e86yaEsD4aezY2Op7Lwi3lPSXemNmk460/CbyOHm1ZyRwMRdKeMIbflTKzk46KGSU9gqYAdqbUV/raHQ0J6rswpPNBKgxq0iXq7eWm5DbqxJPjExUEv7Ro/pMqZrpVm9WtrUJuOZwv4RrxfP6SJkCt63WbRc04gpEc67tvrcZGnDSboIo1WaZc/weo5xaJWnpKT7cl9P99VqWje6r2umKCmPGkQP+5cUmiqS4qAQROg+XHfoEZO5CHGtsmHniJfxVbwF4Bn063A35vKqTdiAlEtaSSN8KEpGbTaQaTA2eKYyP2BaiJW/X4fr8ClcR9vo+K9GsojE9puJG32BSfqZbkboicGlIxzqnGcQ1fEMkjqawcbil44Gjdh4GJgwxbczoY0ZwEaC9F/Pv3y+vbwe/+vyZs6lcQEFefN3Tbbp6iJJMLezx9zFEQ0falxmSdaG6snapBpoJ7o2GiMMfUcGoCvQ6EbGNrauojSO6pXGcV9K4/SjGnq4b6nxKxt9/y5vgxzFsgizB47hblK/RRqetzp7vVaeN83S3FWosef0EK7TEoRn6AFsGS/UJGsgLCTgoAXFOFSnGIdsM8k4dzRqKYO0KcYx1gSA0XhCoP6YhD9DqHfAWOFTB5IEsSrnjoKhI7e328kOlXWrJGmoE0ka6oMkzR11b3u2caTdQeNYxildpVm4C5dqA1EIQDSfeeAPscHom7SWMs6tju7cbqP9Mo2KcRgaxx0TcqMU1w1oTGKTVyW5DN6g+J2YHn4GcGgi7NIjG9zAlpQ1XoFsmLG2oUQk8Y9ol0K0wfM1M1En1brXahrUyH+8rqLZt4mc7F7us0ePNOcNX1KhEojM6peqpY2U9+WCNAusiKT9nSmUSBHG5Y4ei+IzbuoFYoONKHeGVOSusCwOCi6uiaKg8v/ueQeI7lVGYI7ra34dZmvAzMbpCYwkfHyxNg2JveNrLFFVbjvvlTmqyvW9ITIJfvXcV8DQMbNVas7HNMsbQBCy38c7IMvY7OK1hIVtIOoQ7bWl2bUmP9S6fs52jtXWhmcXT6UIP1V/CaqF3XV8GDIMsMGcirpuC1d7GaogTajinRKp+I5CMmwWqHj5jmFO1quZF9tyFn+s3yE6J+vTFFCklcR2k9rSMtsDRadtse9W1MFHXB28Gund0SDVAIsrvzu7NzJBTn8oU/aJFYEeQSJ9mn7PBHOlyY5jo5kN85xBKphoat711UTXG3UvNM3D7Iwm8Wr4NEaI13SEoct9QoMY3mI2amG7nt/CHlwzW6SqfpWBk1dTFCmFHLVSmN6Jtp0QNikq5VAvKXsYMAUWMTZ7e2gaQPVtbQkDdVpJ5YxhoO3GniZuhBRq43QXMrbhYwXeON6rsdVuiviKJaHs/w1oM8SWr5x6zqKX32lsn/dgxQmMjpjAGhONJCaZIqY8qHMTP8LYUSEkS0wlQMzeW5sGCNORlU28D5OHfVMW6fRwVI6LKNuvbOd647kc202jXZzFP3NaAlH9q0tSbuQY1cUUW4+YKRUvzXrOFXKNV52VztgIadFKL1/AzuJ+FGlcDwrZLVhEPW5WBcx+ZKMpPJJkNRnsy8y224SVSnYKK1WnDWAPR75B/6VcbllSU+lGnr9DcCEAJI6zw4iYzju7CjC+Cdd3XOrsnP4XZkbjdT2+j2cXUpnhW5wsxRzltOyC0DC9O95LXlQsenFDVeLqKImLH5IILN/ARnm99vxnLJeNpwd6Cz1Hm3wPDawS8zkuxh81j0AapxrVBzGfaQTU93cZPylyv1iDpCBN6AECfurn0Ab/RtO5TCp8m+UN7km0/nlINuARWYxl5nwVM/XOt1yxslaO81JAqtdptnmG73WMJsqbd3Lb7Lu4aDFEVw6UO/UD5SdNJtTsTmw0UE6OOkpzW+p8/OMY2qGTZLBdYtfogDTU9OQV68dq+cQHh3sDPmhMpJigUHPWTPx1gAi5bvA2EKE6fr1sx9KmmQXAlh2nr0Kcce3xMYqVt98BfUODhWEweiuqaBqFSSBrwI+EP6MNjTi8Tvu40dQOwzXSgke2EdTANgJ3BtCNnIerVUgz8XWaRAKF2MAqiUeM2UZc11WkWV6nn+7o9NORfQJLZJ19J7FEsowLS9cvw5Ns+d7l+LbB4jld0zss3sUG/uPr3Tk08QUHXGO4UYpYcjm4tVwOTNCUBgxHfkNIZK0vcuHdnOfQdEs0Mx3e2N60Rk4ogQ5ILiZk0cf4lOxZYPrC+Q878V52sLvD6FNpy6BiQ1HPnMh0L3PqZkJWzmQCFDfxLovWeUBkDK5yXfqofY25wXizAk75EsGzHRYNChReV9Neryyr+Icvz1FJiGZWUFDNa3K/1XU4P1ZDi0VgqcGvYKnp8A5NaGpgeh3J/uuwoM41MUOcOK9FnAB1xURu6H+K1ulj/MCCgoaWvoa7wnXIECMD6KO0mMAK3TRVjV4/Ve0qIPGTZ6r5ral4R4aL2ryWYdk1mrDS3pxlY9qpaUzzW5P1pesg110a04qV3RvTuZuettLhmfaWHGcY+H155jJN8V+XpjTb11DqVRIUfdW3LEzjPgvTZua2VaZpPDv/LuXzz+l280wd1fKQWDSEOzr6QT7xyY+QNSg6WhNsQQxOA6gOMbO4NxAgonoImTnlp4QrW34IE/okHwB7tHkG5w40dPsTIntqJ7Z7MBNKAFcK5GCfbGKuaWp3G1RXrbMNW/bymmy7OhOFi3QSLpK21o4mrMAjulE7x+zX++0iiawsSuglud7lb+4Ywt/baK5U1GFghRMssCcRD1OHQSoXB0PsBd19k7peWY0EhIyqW5YjVPKE6DyJV1Hx3jtVVKu2+kN8AjiGLS2v3BCyzYrWBWodoe4C3lOsfq3ecj4oguoHRVzTQREXO0M0OmlQxM0PDjTuWQVKiIfi5EeUWWwMeQnVWyhJs64RTNksjXMI3C9NEnTzZzWf8gPd4NsUcBP0d/ChEZZCnj2Z9fN7NrdUs6vGTRdVyKaRcp1qXmflOh4ocV8pRe30dqTfmCye4wWgnCKRMa+Ty7RbHLmJxyybzqPaprN1Q3dyEv+km5Q1FWjgCaAceh8A9dfWLIVUDG9AZlVMWDMTtoUJ3AIG2WI2MMiWMCVfBqhN7DmmASoaUQ/QX4dEBRfPAHbCCwedeT2q9nlDgjtidpWlJDMZT4pf8rgxEnqiCkCndEeVUySWxHQp8Qd3yUsUUQ3ckfhfUa84aigRBfER5cUjpxt/Rj4yVDWTDEeB13krSEu2KVf7/JYQAximUBfRPf4cZgtg32fApMa2sM7nUutlLd5XASR9VWiRBkD07LIpvSYWaV971IgJzlBarIkA1B7NJircVOyvXlX7q2bFRUSTrt6Ki6zio5RPksd9Vn/H6gJC1ajXY+ZYdxf1391VbTVq74qj6w5Nq6Um9BW9T2iE+sShxZ0wIJJUqUJbW2NzU8RdsaB+QBwowBTh4NcrabmoTyWtEunv1OYxAekD6G/TzNAlJ+Qy9cu34vxhZk5uPnze0VuODctBIxEYtOt9rSbgsYMhQSZ1y3i7Cbcr6o8WOkbInFoTaag1zdWAXNsdEvs0Yk3S6G7hCCqkDDR7WOcNWXzKAWw2vGmDHFeuP3l0Q09boMuDFvJazZVrO72mOXS7TSbye53sd4fk+ZBuGQfUA9+trEGmYS3qtH/tZulu3eZVhs1Lh9FY+Hi1x7CH5OTah/lcEFLJKE4fElBsf8WMwNR2VN7YT3uQ5jnuClwvwST4jK7bQi4tVyGb8T5j/AOgvlQFivMKklwLi8shPWj0smaT4Rxo24s0LoEAnp2aGYj5cZX8xnlnMZGHn6WijuX9tVbzIE0G31ALw91zXZhGX/N717r+elnnKbpEbib2V+yoxG546ILsdCk8PJLvjTCryOmay8rVmOg28Imqq31sEKYmqkjbcSYnaLCVdLTV0QfC0kIWYVb6gviT0Ah7rfhTzZs0lX4ShINRT8LBNW+tu2gwD2MK5rqxHNx+PyTpipf1uY6TbagYanRC+FD/MqLOEBrS4XqXT6cBHr8A488UCoIsZaHKTz22w3u9bdVVjiyVVXY9JeSTpiBEtdUuDRiTy7GNYo/fjJ6OYs8aHMkA/yIjXewg5NvB/OIfEhP07GJ+f/ltbn0Z3841wBjnxE+4Oy7moi78ZAi0kmlPO1Ost647vENZivuSyXe5sjONd7DfStBEV4BbXbLhjLyhLY/JN2Yb8uJVRBF+LaLIGTnDAJNXYYpKa5x/B2vOi10lZggMjQNUeFK9RovHMbOqExynxIXiOlwoqtH4aqgtuRpbafgz6k/jiw1AydKmbEjkKdkvANBs8S5IQXRMg0Hxy6hh8t7RPoPdZ7HjwkYTpQT6+eWwS1mOS/NLGq/WpDCCx0EaFjXVWqdR27RY1yKXX6sqsiVBO64dZHeN41TVsE7U7HW176rkRQk9cIXTLvT7LSaWlmfFotaA7EE7QBGcEQJ0cPeMXLJmIGXo9VWba3tEvZg04rXj7BHkv3y4TkEAmEOVImsGDFNQzSMcyv01499ID0H+TZpO7Kj7EzqNPGfcskGx8CLcbPdJJEzacZkyX+KcSLf7UjoMd8N6Vmsjzsge+rYBzYuyoDQOSFMexSd+zYDAC3qLz0ddVo5/yR90a6Sn4AT+EPUqpzBhRQdXATRs4mX6lEUJk5XtwKEjvFrVSqdl8ktesZp8wJSt5PPm9Gd3J/GeGdomLNTYpXQ4LkRIP8JN+CPKuIxrTm5pUHYUnHPgNjtniSy9svRxxNoRgNO4N+C0EzhD73TcdHEUxPNUskIHOqZl5r+uPlgJF/7ibEZGw15OYA9dn7zJtNeYQW9kWpJS0/W2dMH34eJZ27LW2O2PhiMX9aWp5XBVDSLphtKLIgkXu5q5qQ4lUTARud1DzZr1iuP2nzKF70Rh8mN8vZ/2qxB2VVUDu8TaTiuwW4QH/4rXCyP6TscPhthIa3G8DqNVHOaRM2p6867+zRPzF++DNPapb75uzh3Nx/I4/s6awIzoJoleogpzhKmYjeO7EilqD2I2JahBAAp8qMAaOHDLGNgg3HG+Mwx6A+cD9ZO0pWcxRxCHC00EpnMNsl0d6ub54B+ydvuHcvZe0DjAjRoHxFDioNXIbhIHTPaDKJfuUfbD5YwAUxqDwcd+9Lkch19MCZN8AHBgdnn4aIic7pEuN6Ka4fOhDHnQIdxk6UtErSKnTI6DVd0nq8TFtEPjH+GKPmruqsJGX7JwDTVD6nbDmhJ6B8VdxwuGbnMQUfEFtesxT5tbCXhnFY3NQab5lLgR3NnxaNhoUNCRlqoWqJ0+CtR19pxaoi4J+52eCPsdhRDIiLC/Aa2W16qFVGAXJ9Fj1gz10Lghjwz97gjWaZgt43Ah3o78slFGXpNVnN8tbj4kXUXVWebXjecOgxHpbVK3FKZza4TpnIbKhu5N0kC6lzmfvEwn1INLxiDZsg6tOUPLDIiK7lgg6sod5So6BneSfm80tuFQ30XbjUzUWDIgo1oGZEse9ulSfvO6Cw5WqY95mUNQrxipkXuVOnTwXEPdg81pOR3iD/2gt5OSs90iDdvt6XRUYCdpCSF0XLekA9UT73U7skrA05pmrTQFLSU9uQpCgZBi5RfmsKhbftjvDlmJ6uhw8qRohD4hQgbEuEyeY0M9+oZx4HsVbg9Mr0+pn3UdMzb17S6Fu8OGV/BAdD0RTdxBvObR32ooev7Pm9+syqJHZn5f5W3fxWzY1bQ+pdjTtCOEdTTQLPp3jEa4EsKVdxbW3FlBP3cWIcORibBuK7cvZx4mde6jK8GXvD/d3gh8bwsRhDppE+bT/ue+ZpKlwxiIQ5AJ04RuTb49SrlXGU9UQZph++kMPRlpvjrEfhPNVw59knqy8+xwhCK9gpG21ejK1dawpkaatNT6dFq0Pj1Tqc9Wu9ukPj1F6XOivOQrXlZVSq6vEeFTDT9F205gRyGvYEdpf48G5Cg0TpjIUyDn9K0lPI7xjLuFLr07O/smaSXr+q5KRyLmLNd1QNMv0fpUllXH9bo3aMV1PlqTm0+/SbN2H3ibFv6bl0knTKBUqPiul0yMGcgdG1y/Js528dAbGUjUy6vp3y2TaEPSu72JHuIfXKPtJIZK3VwYe45+ddroYp/a7LfECbG/3uiyJuX0W5NycUt00FaSOhJFE1kEZJGkP9Wpqy4R4v/f3Jc2N25kW/4VfJkIO2ShkAkkltbMBy4qldySrJHK1W7HRExAJFRCNEXqcbFLjuj//nIBQOQG5AVR8fpLt11F816Audzl3HN6vZJB/W1Dass6mc4Cre682b2yo93KSJRY3EJ+Eo81RdmodBCLSgdKB8h0GHyEqXRwq6LZr+tMPNLD4DXf2uYcbC39MAaAhBw6+qwVraXZHNREk7lSb+a5AJnC0E/ImEgmUX/ShzmZi7lBkcuRcAboaBctzqWCoZl6zac/bravh1UOq5CN5Nk0QDqRxz/y7TNbeJ5pBNoK2eh0rutMacy1DxNBNaGOaUpAYiIwEJstY3iwoCEcWCgM3tqnsGQPaCRJT+fDvlGP4GEQkWtlB0YB/+o8UNG7ArumKmprRddERdWXqXy+1JTuOVkKe0guA1ore3teDWVkxfVXheLFJTsyPIr9STjRTcuBDxFn+bJmRp903Y5VsaDv4mkgZ7LDIpY6NLK5bsLkdgOMdx3V0nm764j4NJme2LnkR72vvP0IxwypmXurM7p7LhOL1KDjqVgvW0JzskCFVIlw9k8pl2VSwnS0J10CuOWjMi84WTX/zameIbtnLStiidI3J8mOCJ7QaAyeUNM7c+QJ7Ri2fOCzQpFhroRPlWw53xwsrgStPdWUZfAls3PBSmKRDuMk/f45z5PcmHDUt8WWbvL95u0gmolwCeKQEcqlQAVik9mG8VJz8e0l5xU9WP0w7NFdlOiRaLhbVQ1jznfzoY2sSWdanUhC1rhN27dcw5mPITVi1VoPuclEVeiRy9zjScGFOPUjNGLP6aN6P06YSPVmV5zScMIpZCBBMih6TaH+srPqZZOpxg08y1cL1mlabp7oNcWrhzRe3LA237GSqAOhrfK77SdJ/ABwqvaI79IQ6EE7rI6jBCKeOvN+2TZkTvBxW5PLLoMMZ0er9oHbK96MDtXTdrff1Oy9WETfsIEhjLrro7ZpIV6xRW0wL5us0nJS+lBLA7zUlW09xAEAd1xxratW28NB2DocBEuXsSr1fmq2XM8WoL7ZAkSM0rcudQiQ0w51iKaEE40otKp7OUhmVeBMFJjETfEHU1QTcXx1OO07tQwgtEgwzyVfzHM+jfxYaJIfWw6RH9N9BGqPtQsCjLdJQ2tNVs+H3WbXgHDBtE0hyhRi1FNom5pENDQmohjMih0iesjHeEgaWtEOhhLpsXHK8rV44yIuZc/RadnoiPgJBtydZpNVAUXj4J8VzznrPzNID/yO1IjuVNEjuXgiWbLk7mwnSr99JXbeeB/9p3ifSjkommqKR+0qg+DomezosfYkT5ad3e+Kw3JDc4ViyXMKXh8CcTPoT9LLzaAA0UQ5MFFJ8YrlWvDBdozEWxJ8hAHAZsXShRcF56wvytcBrwCytXEe8gynWTE0nq3AMmFwQf+K9STyZxpAiaQRT68C7c7Ybzei5CUfZ1LCaHsi5IMmSkTC2JiUmktopOYS9SkAuKS2l4wtRT6+qWaN5St0erPlZQAZvlBNdY7xZfqRyymsGD3Qa7H7YVXSUI2dgX4QBN71r17++r5if5LWf8AyDPbVRcb+OKZ/aInDMuuzuedo85LRodBsxqsctM14JZPurq5DE7ftI5OUB2Ttkil783yC0M9T+eX/0cxvJQyC9pN3c07T0nXJO1k35/taA45rUf3k1R+/ZFm8Q4OmDtUk0ZIwQH6CxkAzV/MpyDbCKFEOecppXvEoHLXsXmG15CDww+803MiLeyokYyBNTBgE7klmZ1lvFqRz7aD5Umw5dKOqy4OqUzhL/ARAqKjYumA3jPL2RCdB1HCvAjzTGD4YreZmseUo3+4gGY/is8FeFUd9U0fE8Xgj4jiLASe4aUK8BgVP9S4HI+8otl/ZyAuPf7zAx5614+j0TjUebis1hm6eNR3xEb12xbkIlVLNJTVJ8yhxQELLNDiLxinTPARkxicnZbndJxqmblYKgk0KJsgwv2Ra/sVm9b4vFywI/vrSCiUboo7QQtSBxibqwFnoJ0NYOtwYOirBTaKP3px59+/0e2hWsPbOFu/5etNocIKxEPpDDLrEZnRNTEhoAAlZZWId6rewN9xfvRWN/MjYyHcdeel7fW59fEx6mBHRxwmjKpBigr9OInLUXe3gRvxLZ3H8x6dfPlzPHyd/88g58/r3c47E/kFEJfRP0E9VarnE5+RH75cH8V+KpwuXF+LBsJi4jGTywgZUjjp+Buz4XHb8gWbJYzPD9F+YIsbxbMP6Zct0ZehZmIIPtx7n3A63x1p8Gw3jB8SnO3eZbz+sNgs9hxccckoF976pkKYBlI8SM86dOHSXLFZsCTDwpaZkonAYsFT0bEKTNtHOg6KDmZfuAMuazldlL4knE6xqqrNp63250tvLju/OXa1MMyXGZ6caR83ktdwsc94hEipLVW+uiwbGdgylrJcLIKhv2cZHDlQynaiZ6mT1thX6YSyUG6T3jtPUD2OoWLZuuH3SYdtJF4BPujTxA4Dcu2atLmo2lCJ4TDIXnEL0R7q4XCzUSdqOZtChir8IBX39bqvXkR8SMrLawJUGNNeJRV6gqAKbFIHpETpH1Gc2RAGf8FEcf8hZHWVZnxE2AnQHQiKs0VPZEZktqy0ql+svEpuLKNEey+E3QTT5TRtlZdjG/bb4xqmkOuqvkT3KSwM/VGoksV2Ci/cTbz//5v1ZFP9avXvR+Z//4iqhq/fm8GIk42qMUek7hMBWjcG7rP/QUgzWclUqEuaxWNjQDA6YIoNrHdpUbGuxC2nz14ZlFxhJoKJQ1yb7TOOfipvf7fqWdw3kApdN0ZV4e74t6JeWi32xtNcqm+AIjxUcpZAL3hAaNTTdURdNN4KxdOtencbSXWnrRWZtPaGZaD6GHKCWurP28cKWVekYumC+NmdRAxZnyoU8DQ7VA4ihpI4dOe/sOG/Z1T62PEBC44C+eXmImGHT0EDdDY2feNui4P+Ur5d1V+NHUOsCJ+R7tC4yHXPYMHehAEpOg5PQzzJ3lINsqjOOiSdzhFR9rcX+sGIpKIQ+mvkYJ6eSWSfTGVF16Pb7/O2w+lrSC5Oed7esDsiCNY8L6hWFF34FDX2ZHD1h6qsqjIUGTppPUk2vXSdT/qZVNYOXzBh51hh61VdBNNXURj7nS+oU/Rbvl/kjfQW3V2D0C07UA64jSbifX5Lywks/4D//VeV/l1OkDCKKEREWonI5RiiFEo5pzodidxZ43d6RloA3GsKgm4v2zFM7aHzcMpEJKqxNCOUHjxMA4EmEfy2HziY7VvddS2lMaE5j8OA0RmPTOn327mdll12vucz9irVXJ9uv5Zp6/JN35NH8iQ2/+Yw58ydvGv/kPRargtOY/lnuHXaZ1F2FPU/HiB5mNNeRAW58JbJjVsVAwf9q0VOiwMpPiR3fvT2guFxRK1tW5it4GNFOZ1qEFjMepRFVOvbr5puIz4ijEhU62eHKaCv1qvyba5kik57afOVKCFyYKG6WsjejATF96lWxZFEIiyAcHJaidRBbHGsFyt48PhzJomcB0dNEtiVeitfNdvuyKZc7735Lc6PtkKaF5mkHx8mnvDHZ4jj50FoJc52o/XHzvPdYH/75mT5ZzXy5syhnZcO8HKCcNdHLhm+NUNaRYNpDwMaF5qpdoflokVphazaRhKawUWgqAgtN6S5F2EFoihuqTtsjqYqOf52XvDbE2sN7MTel1C0cgB2GH3gkWMcDv8PUt7ncCjjTOl9VfJYnZGj6Zrd20qnh4mjTK17pWyvWi/dGEYgE5JLFi0ZFoMubyePn65k3ndzNJ1cwcaD+LeTOC3PHicuUZXDPoDJ/CdUO4sZargYuMUA5nsctms02cQbuIM5Q+qcOxBk4DocTZ2DStAISXQBnXu4Oq2c2dKhubunNJY5+dYWq9FuqqMMsS1wNiSYyn+vXopm4kf1zAXLE9DSAUKNYFXqYd9r44N1mW+xfyto/AvaOgR6j050TrVjlh707rwL7xTt10ECU5tKTJZlPcHR6V1YUHZTZCnl0EFp2IKkfjEjg2iRzyJrMYXADj6SA7aEbO2ZyTXGUGIujCMOroyDfNFMiQfuoMifcFvsNvcHftjlYyhqTkdMyGiRdaUzCf6c3tsBqtLIYYJClMUoi+2F8tNdWpm2IypGJqBzL+D9HugbdLRI7EJVzY31kDaJ7ROJanE+rgNJz+u19Rc/CzYEuFD6CAVTow4T4KWCIxmyxKixrg7PTXJD/K/o2LhVkEvkhSsarIE9rbkDp/b287zevNDpblpsqJEQ0JjQFhS7hAgn9mDiq29QyEH+vA1GaftRpykcNBCZmaNU8hX7KU2CWThtJczPp0MttLLcMJ1rsikaPXXUvhwev4h5UEEBf6KlQlcRj+DWI6fJ0vwQVW3ILrtFW7cBOgaDozLtgLIYJve9SufZ0YPsfg+klMAkAE1BiWKht8tgJ1hHTNCCse8E2+hqXpnAU+5kzZZDUFG5R2EhKLqJWFpkof5GPh3D+4oiMxvkrYAnqhAfX8makS900IhHAxf4pZ81qKxXQ9/ARb46hWUAUQrBAthyATySp0Q5NAJ4Oez4n1AU/szFLYhAzomqsV91NqYnO6dKtNdZSMXepjmfCAsoo8MNxwUoVsCI0AytwAEdWaD46Iit0vUS2AK60+zrf/kEvJPabxHx3aw0Th+k03cm4W9VDNtlButQwytjFPsZjlIG96z6FqSrIDMcJMnvfMCTCZHPuc9W5TzlNy0TdkTV6xVwMMdcdXZBUusfWCLOCUXEPCmFc4g1kzd25erd/XOV/bL4Z9BOtnd2Wc2EKOLf6OA2aSxNbZDag6zCk+yIbsW4x1cqhNhDyAAxyCLk5VQByu4CBJxOSqfpRTMJLlDqkwz6SJzXjATX9MPpONf07E3eMSungiKlu+0uzDNwzXNpFIlGDqe/oNaC1GwW4Y7cvv/IWSQyu8xq8G17pbUSJkFmUCIFFiXCI/RRH30OVaFqDAOWGWMEwMUwfk63WIUAJ6jHOgEgJ3azI03WtAXXgpVEqgCXoIXaWc7BaPBIZtnFJFbd/OBK3Pw6RH4KqwSZ+f51+NzOojzDeI8GEgaCcGxinCtlOr+6IbK3aVdvibUXPq1e6bvk8Fj1Cz8OajUPn4WgKoljX8GtPPmD4nAWmCaozBZtmyzu2B/Sa2IyzOG6r0p535n0armYlexyMdvcaXuhsxVJ/3mTtkEG1vU5QKs1wGO80pNoX6x1HK5qcnNO9pp0P1zuG86eH3kYX/W5DImzbDUNkfh5ZA4q1IyuLSxX4neh4ddG6WhULiZvT5arSXBt+UYmcWjn7p0KuiuU3KADyfGCMfYhMhmLLDp5v8JGhER+JgwHMYBgHvkprBUBH/mySb+Lo0+aOgGvnYZSpbXr7lWQwptzxfI4YIym5x+bkHg1I7hFNQ8iQ7B5p2X2TcGJTwokJNOFEfd0/SMLJnJuqI+1a8YbG8+zs/otzmtd6hkCvaeaUjOb2rMbytzuYlzT6z4fJtDL33E9F2VBHsYxJ72p68pPVSjAcb1aqiKeD+i5GkU9CwEGkm+NVvEhvF/yyXubrHS+AW6p1NpfCHu5Rmal9X9Lwpx4TZp3KdpU70anSWkcplJ2p3zXLmd3JzVShFfHYaEXN2ZHQiqgXrTiAaEFz1gmvWCmmVZDFh4DMtZN7tlmKYhdzccGugQ3DOQkUM2bY9t2mZGDb1sfom8lC+jdXdJ1Xf7wq859giBnI67fjZQQNXaYXGRaMxbT8awgZKWwZ68Y6jyiGZ1DgKUxLsxbKA/N+Mme7YaPtF6ma6lSdiHDduNNVfconofBY9zDOPpe8j8b+4CY/bN9XbBUJfjTogCfCCkPTwPnOO84MHmqq3c8C8U7c+EuVcgSi+XMSAEQxRUWiMdusAmLVZ5QqfHVtrwCuCupmGLiDMe4eJ9dzuTZeAaliE5DKAUGlvzfnAUmBGm2QVNdfxE7Hhr7d6u0l9yYVEyKb4CkPlr1uSSSRPobdcQ28bDf5XmUrmSD0SePVWq7Oc+bc+X6z2NDYmu4Ijz8P3RL0Iby2n65DYigYY35FsKrFMn7USA7IkQ/eVU6d2Zw/FPR72WQtSOBJ97kzehJTTZVh+uwVDEjrdF/SZ9sJVVBsVCVwAf+gvrHntgirapC9VUkjEps1IjFYIxL2ylRTXXPjIY2j1KuSHjL7bfm1j9EZQjje/2Kl+/Poga0KTqZaW2nGiqiMaoD+IPebfb5jwuKwRgLIy5a9urdxo2mzH1YVL9ywqyXIfKRisLPem0Wx2sKbxiPhTYNsXLxpUw0MjdXAwEe9/XfLjg5SH8cjlwUnvCyI5ePyyLbHGNWgNfgg9jNAk0g1JogCMP/nupBtAw40Asv6Ye+1pJXnN+feSsgukyCAthANzzNeEzEUUtZq5bjiGQk6Gp4dZ1bbeeKTDHIOVKYtDIv1MC62cAqFZOAwbkDG3YXszV6ptHlfaMyy3UiaHrD+ouZl1NWb/XV29TdP2GSrELXRY40yDv5+yjg4iHxMvpMyTkPcga3EHWEanHntaD+EE3cEyCfpmLwdNNxCmjbegX4fm3DmraVv74zv4Db/ui7YtevNmFwTKxjAqJsCxmwGWdJsSssqPfvRAC5tXXmt8guqKKeA+0/z1s40JW0/TUWelQPnmi7z5edf7q9vJ79f311yNdXHw9vbBlYZ1F3sGLg4VgY5jxQfw2eKOiho9595y0muWzJK1G0hkDLsdT5uVocO3q7E1dcT6DsaSB82Q/oEDG4QpE/30xHTh6pw3Ho108N3ommtyYho0oHoQs5rwH4YfMlp2E1/TnoQlOfl+rlmuK0qm8cXzIqbD2qPavLKNKDalcazxSr/47DK1/VIJZIbLk5Hsr4wuss/T+Vmr11yrYYVq7ZMCeltusxL0STKFx4BV+H6l0lntYWVMecqJvV+u+ECwg1siQ2dHM8xWE1Td9A+6WI0fP3lw/WtCD3D1lnGACoMohRbKLIjYHCMsowRD8IAKm1rCtiXS4Ao1GQPm92hEeYgThIg7XYlc9E9QVeNqWDkbKJh5G9KGolvAKjJtm8pgN5eNqR6VkluRjbJTVdeSdk7usoBP65u0FJd9354y+nJmzP6PBbnv7BcvVz8WEUHc/v2P2O0hYdt4U0PNCt6ed/sZCyFQ3CANJr+k5lyOMdMZuCYOWukRuYgRhmURX7mTpd0S98Bu+k5xyQHUMmN7KqPEZ7Wx5ArICgLfeIssK6aY32uejyzFW01GTCxlTuVBBiBE2CT32MlwEIiMlSrX8UT27bLXCTA3tljyUM5w/S/0/LFfpKMKhzZnLvEdu6iYMDBi30cDTt5daQIZ8o2qIswvHrFkEKAs+IGB7PuYoJiTBSUxY0QanhKCdEHqikzx9yvrE+zz79r19SsU9E4CsAXVeBnwVjg/ruazL59Wd0ynfAatqcF/9Z9La843cu4hyRKNlv3Ce6N08xb+lP+14GG3wUberR3SrH1HcqrLbIrz0u2FIkCUWLPtCrbU/4XYzKACj2iNIOV1lt2Gm9ihc+yqr+/sulroDcpU8aEMrDLBqskU+tDsOn6bb6GoByvb7WOCUpjPwOA8mWbneLYDOGjibnvj2BJKDJTd7VbpU82deFN7z7+jf3zeUVL+Gd57FbFrROHSavMGVBTkVYphdYJuNSn+510iaq0zLCTQzl1eHaZ1Imb9oLnxTd+VnNIZX28A0Eour9xpyKu0WJH9aFJ3NKxErcUcsV0J20mAjN2fjH9Xnn/W7tC0ikOcs29LTRVy9OtWYkU2hVCSeITAK5EsXXh0AOKHSNgTMARcBL7GH+XALihYUYmGmbwWANKIJMDkp3q8LL111mN7FLj4C/WfDyevlX2j++rN4ay5jUz2PGVAHRxW0aP5jzEc6LjoaUpa7EMYr0s6H91m68K3pCBHlpAHu6jRVSfs9arjDtNIp3Yon3mhT5WRyRxZFG4PPHJeo8QhpbhTonnaqGpK0lUbJRE7aqqWK7kJPLTlAzQRO1BjNyYYtbP9K1+K58LbRzGKjx89c87FvHIfVfmtHuMKDhFGtNNqCjmoZTrVwaMnzYX33YYjzYPdcshWOoFVzIht4UWcDuUn/t9kw5fyVAdkp2rSNCq1ivzUGAzDwWs3JAgZ82W7k7aLIh0dMDl9itd3HSTvxYCsbXzUMevHTn62Cm/mu+of9TWZmuHBWiTvPd8TX5dHRYcNpuCyzYJTf+SbFSdoQYcSPrAgWru4AAGRHHmE8gwRQ0HPGp0JXosMKUHPTtsNS09hyE45hGKorEGticmytWe5nnzh6KG98KLeKBGOopTWPXZ3kdnG+o3rBJk5s/7kjr5zGDdnSmjZS/FNLIddy9NkUKReXv5+dM/bx5/vfk4+XzpzS+9z9eXt7/cXf96K1gTdh3d9NDq9aisOM14UmQZT8Lg8SQUxxoXoOt0Ukv4Dk9+I+r8IQ2ockZAEiaiwCyhVDCUZcLkaDfFpzDf7jk0o5vIKm2BhkhboJgMl7Zo0mhsSaNDaBoYR34cDe1/hmoq3fhH/jP9ExQ8anO+fH3L3wRXcAQuPjIHIQMazNrRVlOwFUTGSpJ0875mCl3FWp2Oc+AvY45hQMXbymDGSmNT9s4SuTSWs4ukXA3AwTHXgk4NSLk6Jlmip3SqUFZooaKiTu15Z3RVPBXrxSrXpdacViH2I4j6i+KAbr6zVIZnmjrM8QKncdF+R0NTb3HUpIZ1ZuMAQHbR3Zidmaa1WiIfGKqzozvXrUkrlD2EobY8RjVBHBnzNHAuFgeDBog1kAPdELea6Cub6GtwpgM2E4n9LItdQaW6tQuv1sROlt5scv/h18/XdDHSfxVkJT/w1/qjRJMaGmlSwWThzHeUDWFJTQ2T+EgAjOXeA0NYvdZJBAHPO5o87BDKONrLG3tmqFhF3twefWw0bcNuyWdYK5ZEfhoP1bTFRL9BWWoZq22pmrfJSwgQw21wcCSKqLs6AlZ0wF6Kb++8ird8V84Cyd3I1d0e8njNXFu0DkFE6xBMtI65GoyFeW3gC6ENvpCA4QskgvVtFVvteBNZ4k0MvelJ6BM8LNrEXcXNxtPoP8VT9fCcqT3w22JVF+4I+NhkpMXEXdlRttUx0t60mLG5xYzhLWYS+CkgwVCNeY8PHf4KEk1lp//y7f3psGfqcyD+TESCEfkzKxHK2CBCedaoTjpMDbf8i5gsyRiSd9VeSSx7BUH3SpQObBYjI4wLq7MfTLVhX66GobiiBHAMaqba8wiheR5hKMEwimI/HTSN0MMv3AyDEeswGAaTTRq8pZGq8zRYU6Cib3KmUccz0O5msS3f9gYosrR1LUkXiOlcsiY1RNAYDZEo8lE6Kn44CD9qlATzzaugFdZad/8jPtL86aMBdJZvy/0Lfd1cCwZOfIY0lviO3MtkreqH2ab62HycXhbK14zLSAzthPZXa1uKyCfu0luaMbmDdzwdvR+i8/TPf1WiNCJyrPHwqgDIfrt5k0mKakV04CoJRmrt0StNV99+4LyIK+AdOAprRjOogc2DGgicakeQ2EGxZBpyUQ7Ix/K1AT9jN6z1UOdUU53RVyUPh+3ycAiqXgf0tqOueqlJ1z0yGqNlKWahBhSDNNe65BI1WxdedB633yWrCdXz/0dZFaUZ9lgsjqIqMPIlg7/DyJcaRU/UoejpBT+RU0Q9DT+8i6inmJmQZD0blofYWLVOCRjOF2Z+QEL3OopirEJf2IG9+MbQCV+zNu3qwPvIFe+vaN9522JXbN8qfg0g2y4KUz+NRgZQihhZxfXlT8U+b+DCg2Jk6mwcA+hnJIsXjBPaqsJBz1neeGlvto/FevNcPgnZgDiBHrQhRINas6Urm0zVYcLJvhBQcQLOw8PYR4n7z942ZBvTILqwagWAGCYHYPCxlym8bavdm0jGYTdlLg1nN1V/T0315XFTLdcYrlOEQhAdddvShXd9e75lzc9ysS+WtpizDpF1on0O9lw28qo2qXmXeDkMfdchwZbZ6y9NiZ8drUQeYuYrIB2L4kd3EUbwY03Ub00yatfrpRjYXpRDKCR1Z+2jEbophT3CPCYh3FZa1q2B+5rD678Oatl6kP9dFJgtq281cRi36cCF2bS0sL2lxXgE3hesNg7kbOxfMz2EjVUJJ7aWcEJ4CUfzCUfuFZzwCDGKdCmYacmrEYU6HGPt+SjDwqAtJrjSjiYlvaLQKPIXgAWLmEvjCfyh3zRo6/Ury7HKJ3WTO6k4Il2kBPdCrmWL1YDGNxWJh9yQeARcdApV8MHJdbFkolHC3pXPhy0jWxd6cXYnE6uTwXhMNtncPu0i9D3GArTj1EfxaAof4aWB/bqkTlYAG6PbJIB2z3Hsp4Am27w8bBv28xaluEEK6Gpb7orN82H1hyHVlwjFkdU1RAAgx7ZBm64Gk5CODJJFNDIzFXId36D7ISWMvbStmQhMlMT5GNTCy2J4WPRsmu8nE7PekywGxsoPeLC+n+5vr76f4kMlI81ii980qfOSZfNNCwbMS4Qw8bOwDw3ZGVc0hU80EkMNSLpHZajpnMBr1mI4UiaHQz9N0/GVKm5q/WhpabKokdHvv71UqMQBdzwOAXIllVKzyW4DgxUjF6qedLle5xXhfUfKbrkvMdOgG2niQkiaI3R587vSnZ78+vvlIB1zg3/DdcynvE+otA2uGYbo7OOGz9LMrufemSeNJ1qGamzuMsJcQHEuL17LvHql9sKc+ouPoY9rcBUokKu35hSV3JnWB5GCJ+D2B71ZyVDH5v/ERICMek9Vxd4ulue8ArrCZNVUOz0n5vQ8sVfpnS4hzT97wcZgtelJNg1CJWie1gUnCQL58r7fvNIvWpYschGsn4gx6BmewoEunj1F4EigV3URvb8Lq+IZmtKYKgD1kTGgV0p4gff69SfvCC5/4QVLHuHHhP9dzQO24KSrsMIZSv0UsAE/4fOn1Wbxr8I2iVXDgsKxYEEo8aN0HFzQg0lW7X5XHJab4u2lWMooDanZhKzOOY/g3+U7eicoQhFiuDJRwflvDNvFBqvU2pejRCpiclpBBKyF6HYF6Z+efn4ptlxiYNXF8mQh/EPETwJ1wLKz7iBb40c+r9dG6ZHuHGlXwP87YJRhXqpDuGG9YCIrdKl10n7evxTr9xVdEmzS9JijVllrpcXG4MKwgTcU0a323XSVPxrUSTny+hur10N7KLqvDvxIqsX2hGNomXBE8AlH6lsUZMNGHJu54VsT8Pw0ZTjmWHCaLNzPAbqaqrHhFX2nOYcj8czQ9/hkhUxX7YI7QKEfJc5k0Eer3By7h1p01QwthLvRQmgAVgjhsbBCtZS3lDq/C0RKASvmIUTf23jM1AIuphL75ut9g+CKwHAx6iIGTWBq9izR6YX3vNl6l5cfri4f5p4Aj8lyWpyWDY1Dy2Z8jCHEbM0EDDJPwGDwBAwK/DQg32cEplqq2K46j6Cq8wgkZdOdUH+qcTtEmYjbv7xvxRQh/Uv+Hz0RYJ4SZH5M4gwkrSeZra0ewfKJASzPb2xOT1+RuXpnbzV8XvzlS/tvz2bvORBWD1LE6YIU6mmLLOZ9NoqUt+7vKYX+dK5BOOblqtxT74pXL3YLEaW2WZAonNWdVHflqmXKhNe5qMoBzVLG1qX8SO+TAxtxCYYt6BhCti2vZtJezU2wGnUHq16lBshxv3RZt2LXQgpeYcGq9hxjxar3nMkvVAlRnoq1gCfC6E41N7Gd7rRlxMNiblYhPZ0FaP6byifxqXjju5VFjyldltBp6QBS3J7l22WZL9S8NZxrKu+fX4rN28s7F5KFISRBskFtM1WBr665mVFx5zUaEVGfFbLO691mx2qa9MRaltT2VhKIcXqZ4akvs3GM/Ic51ojaoVFF7VCAAVQCiqhd023E1m4jvF8b4OHdRhLodCYmFmX9VbmwLgSBjyAqbJ28C4KcJlQh+KtlsebsNKhDySVy9g9ATPNQ05cQiRGbsT8f/pXv9kqUaT1KTGzFBs865zMVq5rOxFy7pY9FoBRacsogZFiynV72SaXUaWLdPWN0ZBwht9tYz+rYzfPhFJN8IISoMz6b7fCJkCzxI0DKqRlrmJ/PeYGhg0T1VmuO3LEDofxmIY1wqIZkMV2wrvw2R2ttyohIp4xIvR9+/Xz9oxRqRtZQMyFDQsyM+CGB8OTbjIrZG50s5JJaZC0SpWzoMHWThYDg1zp00wy6I+Oge+JD0aDUrWjImHviHzuM1t7dfYCmGnHn/z2wwfFCLSI6TNtkWOHvj5WeEkYKuFKyVR1XyfL/xOFRUQxrQcV+W74q4baDlBh1DqHxpcQaEhDUQQKCgSQgWeAn6WAOEHzkruC1GWJnrgK1s6hXCGFA70MxJoDqkb4YqxYHb9uqha/Jomg6qHAseJr5fXSY7V9dM9YBXWdXKFeyUeuH+3ypl70c7ss09RFgbFa21N7p8hUkQUpiDVIyCFOSpj2tQ3dEyV0QTzUI631ZpeRMWx5Id5AmgIBJMSROIAlSia2QSgW+5gCpTGM/GRVQyaNitf/Orkz6TGUFaOJdvwGRsuQ2ChLYnX10QLLNOOU2FroQ3lpQ9Egud5vXYnBrISV097vD3VRbHaCXpteF7L2uqpUUDel36Z67dbsitdt1Y8KStDjP8Qmc57DXq3Ke46pt+MjHlJF26u/ZDyEQfMpOezx6+Wh7f6Bupm6s+pmls5OMAsfTPBt+dPJhaaLXrjff8vXxSc5kDVZ1IsAhrEtD9XfuGKKWrGnUdFIRC/UUsbxeXJ7FX+RHITm10oYnU4Oa7cv7arl5y0HzximNm0afNg51YMupVH4gP41EftUQZaQQd9C8eUsT52NrgHnebnW9tQUuwGotmt+4Sya6KvDLc364rRIt7fp2BrxuN9Zc9rrq2AhhUmgKk2ABUg+nSWgKj66/tBEKDcAyNgMsW+OPYIx/kil5OHj6kei1LzH9WNNgZno1qZmTcx08anf8ktSPRupHisKr0i755+alZFy8ch7pUHRN0hH53ARPs9rfLZ7zVfFtELeR5l0PS3PLjrmAJZZmSDeKVoBhyTHd9jakr3n/Jgm9SJyTx9vDbkEv2G2xyukFvK/qQurkiYKVOGG0KIl9CJJMnS2yxrZ181aWqqsatZM1h/AedlXLVkGgVn86k8rIoMPc8FTjIQ0raj1kodZD0FwzCQcy6x21IWN9TVyfINmaQKAz17pca9NniSx9FuKUPN5fzlm+IB2SEFIjxaL3A/3CD9e3P6rVdaIHaZMFuxP+KLedqormfmOc+ZFcWE86QgnZkEf4kquw0c0NGZlvSHgpS/etIyVULSmsBnW0wwpyzNsvnGlFeo1/0gB8ywF0pT5f8OXo7hcLiEt3V9vFR3f/oRvjMdmhrrGG+iAcAzpvnnKaM+zzVafKtflwZyozAPJJXq422KzqMt/UxikyN04JtG8ax34yuGvKNPw+qsfLTLCzcxE9eO9J96fr1lZN2e/tJi+ILXkBAqu0GFw9LTFo1iG2rkMCXoaak50ldcMyJDqTQUOmjWxk2uCrTvfTWQq8vux4xRwrIurzYrc6lrIJsGaue9U5Pa/aUivlKMJxkH28fZDfnD3vc90jwxO/hvoUj0V9Gscwea0+EgoR16rTvw3uEwzvjCFMtrKdbtWTyW/y6VKPngkJleOQGkeq/l6uFx5X36yn3+Yh5zP69RqEy2FSNOlIqJxmeBlbhpcT6PByHI2GtJ4H6URLvBosk3dfbF/z9Vcaou+72SjMMJgYAlF7fN/tC0YzsGRG9yytfb/QJAGjPklAx+m1lpPYz0KUgQUB63E1FjReqsH33wUdPCuZg2spsYpW66qlKIY6KyisgqtxO+yZ+uP7ghNGM1pGqDZPjE6FIQoOZk1A8/0UBuY48JMIO4c8urWqLGoPfCqiq1AjumJbpYJ0q1vGJbKgCZd6DRF3tqudMEzq8S6pKhqeUBU1+XU6IUF7ywgqAg53OmPkBCACApAWUz/9AJdkUpfkL41CEg4GrEdnQnDFzoWVBlwA2w1D/EfUOenoG5mvPpJ1U8lI+Hb6gxYr+rec3Ai1R87ExU0MF/eZ90d9N4PuZJLSLT3SnTzl3CeJ1r5cb+gjvOZbFYnmcI1Q9wB9GD79rBussj8NAD153Xx7H5z9ERo7Js58yqzV2zbF5pl3B0NbkCtWSYMj8f9E7VF/urFKj4JDVIncHnJqRBAlhAEU20AgvFKqJUv11/uhfPboOtrzqZJy/fVH6baXNUHZRf++8rzP20FDByD/u4chcF8fGQXDGslj+MgoXJTT4TM76CpyFTCJE8invl5TRV4f2snroaqghIxBXX9rAlWeNnkPcsw4dd+g0LENhY6gKHQS+Uk4EIUuujg24HkTVeLOqBJOoaprvkEZVPlxwy/Td4nlOzSxfCOw3BYJIc1YyU6n0AOeaOQGl9uvjD14W1byIiDFHhI6c/0f7VRk/1o8jk+IxzVHhkfjTYsTjaKLSrAfhnioMKqd0bvpcMbm8VRpEJVGmcpUTDO2CgsukB/H0ffpbDY1jbivpsGnVa7oGbLdnD/QL6SHyXoPi05pPhyCSxzc7uWsnhWc6dAqekQ85a/slOjC/Jr7dhFkgL1lSkL6IivSV2k9OCB9mR5cPCLQ96pmWDRRgoDGIHTPxpiC+Nn0Enkknr/yjbPQkV0O6UYU+4l76cVoz95xEpl5qFJBshLYcLVk6nGgkuZ2KXbp5o5yyazk9uvn6w+Ll2K3/3B1bU/fWb9nqhHrsu4Q40l57VKdN/dVItDct2KoDeZGNjA3hiY8UeSDWGI0a/rtfsGQieslDdw1dhhWV7pk5f5QqysJWnzP+1hseTLKZY7a2Y9DkSmKRi0y3QdYh4PeMQzAruZAhRZvIjWMGlS8YRpUl5rkNHthbKaUkZSzXyflYaF3RlNJ/sfUhVdVEMRl74V+mAL2Hh8ka/nwoVIGse2yJlLFo+jRRJgebmPJ0bT0a1OdNHTyuqIB/1vZTbCdjuGmbOnC+zif8LEvVNHddTOJ6o1EeQB0GLFJhOjCGDpzqhKb6KzWn+nJly/pT6PQ7UFxzAY3u8f+KrOkVVqrcz+NUW+Wb1nM81fx1mjWgPK+iIZ9SIVToL7Uz2BVrAYe46nqCyYxpc2Oz9NMDzSWfdpuXg9bmBgpk1QjI+t8zi41Ik6GyPj2/soHBN5Y/CHKngNCHs3f0A5a6TSrRj7nKBDBz3kUXHj052lOjGzSw9sEytrCbFD/nnSPW34yCfAZ2Y/woMn0MKGbb+ARgdtj6Q3mClkwV/A1ofvmiLnqiYCbIhHqLBLBDwvTywQqJStVogb6EI4FfdB9HI59ELKV6rVQNkNXBDxzZfjJhwlXCliGyr+02b2vFiJdYleV+KVBgIyQjAXIkApZ6IRClubS8EKWyA0VMMvdZtukahE4Mwxp/qJME3QEp5otc1h6pOx4qNmOZAXNFnmLWqN0mPwNI3eK6Idi96YyRP/MyxaqrDhrQL4vKhw8+DWGPlIAVR2vUbNVZ9cJy67N7/SiilCkFiqytFAHHOb6Azj2UDFx6qE28snIIp8MblOAdMBUSxIHSfEHDVLWX03RfyOziDpkFkEVAyYQ5t7mV03RF3weaQUDxsgdy/5is7/RCP6Gbv5GZqoK4W39dhP97U5YOY/euEtv9pJvF/Ty8DxR7NS6pQ5cw7r7JwmHGY4zmbwNPN8ansDdJtCltm7bJ1PuwnKfiqhG7ny4RKOBMgSZdXee24aOJBY/m0T3OE7v+bBlJMq6vqbLAUYzwcBVYNNkzSir6f1wWC82r28reiuzFVnfaw3mHlkx9yE0PMWQdoEBcc+BJouXcrWsB2ia0hAylYY6Vqp5K+HUD0B94ZahmuGSnIv5FNuabai8QxuVN7hEiyHjppqpzpHEWF8E1ydMM2ieOo7JHUcZyGyqvrzpdvP6UvCMMLUGWsTNH9Q1Z7rYrN735YJF+V9fWuFW0wZA5jYA/DUlw9oA9WuaIHSl3zptbahOGirz+YNjPzmdrbqR08N2OT0EbkTq6nn9BPBGo03wrGGH9tu8jmjhwSdWSe46Cgm6qR4k84RjcVU20/wt/0NAwiPopApW+eROLN3R3adzEc82y8JOQ0xsjnVuV6nLwL/ea/ZqcnSFnh/Kj7t9E5/9n3CJ3bRT0037bgoUXK5Z5IdKqzl1uWhb9rxYqhE37STc1U6KDO2kFLxTkB9k4bBeEnXgQ9rZSmrwoMiMB0UEfPciwMmjWupTohOxggpiP6GZiwM/zgCcu7+ordyP88mHu0+P1OfgPKpbCZXnvJcr08/wJm+VB811+c8NTT54JQyU8DAFtRGlVWq6Rmyka0wJlK7R4F7mRNaYiiXBXy8rhWBCjm8Y4yj1Qs5AbyMXblqQyNyCJMAOpOFBHFuQBNiBFANYyvly/XSoxqIi8NC+wfUuChHFlGUESxSo4qAOY+aqx59YHMbnQxXMnUsE0++xIYZRLR7nNkLD3MZNvtg3cFDQ5AaG0Ft0D27Q8PKTRbrmm37duciopGOoqFzZiccKCQaqgNIcEKeoLz0Lpfq4xWRLSpdMzJgvWZAaxvgPeYnmABo34n+JzvN0v2UZisiWIyhiAyUqLW+3jJJsqrpkrZzszQWARuLrpe6G5BS63uMNEARHtcLIcvi3WXzZaw/HYj4xPEcXB7bGffL4YCXR0++pj4ftZlfwrlgEVoeg6TJAX1sxVWdXglK+56ZqmvZonKF7prmJB0/d92iqtMvmgiBdEeS6L7cVbTkGLw7N8a61IRviMaI0aTjTjl7OA6BB/FwEOYgfjXZVVY7h/zDHmqkIZJuKwNA7igycisCN4utMG9K5K5+LKiLEYMnXyI8B2DHFFD97jGlV3ZzhrEePD7UGnM7VW6nADYhHqOcoBWQA1pikcQwZHCMD/IpG8UnHOT4UezazBXUnGKWcqHcD52Wt7AitI2pORaQ3DJKt8Vw31gs603KXLzbLd5u4qbkQjEI/dL+Bb/JvuRgUoOdrVZtuNiYyb0wwLrXfJcu+rDR7qv3IOyM8yAn5n2dV1eDCm00e5te/3HjPq/yraPKxw07X5Vyuyx0N8tZdXT7s+AhdyaFq6sLLF4d98Tfv8ObtN1w+kB80PyAhEPdjfeSwThBrBr8wYuYFnyxoAoxjdyVT8/WSHq1v76tBNIT9TyZn7NzW/mhMnJlHZgItHzryEogYT+1Xu+jv6T7aIe5ti1KYg5dS4QAZCwcYHvNieu0Mqhxgx8oBV7vX5p5WJX3ORUWFTqAJB/YDBNA3UI3VgaRPVyipV2gVVvKwQ9Tzqp5rk8AjYwIPv5fQGBcB4+sypO+bhWiN2WuMNsE5ug7G7YHEk7n21rYbDhMz5OwOJGxMrhdhiL6wYq3S++oSK7tUG2Gzo8IFWGnF5HBnZ1O2Zddaabo62NDVObtjGJu9Km0Ka/No0siDGsV0C1xOFNhvxaQy9WYMhVF8A+2djIaa+PTNI3C0+u5p6RcoPU8XVSo/jMh48gUNUSXu0rYFDX5kPtakqUaaDBbyq0S729dHYVQls3CAhWZ+oNTBgLjQpmqNLFVrqPp6Bq8aquZEQWiicS0aJKthqgrqAd510nyaff5dlQhFevwzJM9Jx9H4pt5gJQa+lBh4YFIdo5HAhH0kMINZYNLTxXGxCXtpmM6HHx2Jn0bfiSyd6E43KSa0rJ/A96dkq6krqyH2eiliCcG2B6VH9LPIWUdTM6XdDgJSzso6zYD5pZYUaGgFZBp+xWBWOHdSOBmrQM1/wJ1YhYZiD1ko9ppJ6EGj0MSPR5yEFkm6OnFQfDNp2Dpk6H3OtUcN2kbErAVPwy/o/7JO/25BT0kO3P/fERs1z9febr958/Kn7eFtz9CZogjxQQJBTzSwwqn3EaHhEPA+4uGuDUfa9Huwud8DJi3VHIzdmj1tIEudM1YVEoxEzyo5lkFamfCDETapKiFHwCRDe4ysD22uWOvIiFhRRF/lJW/ZLqqR21LTN3ehbByDsXFqQsO3uBNM8wUunAl+FgL0TvrPChreTmbqW3zMV0+HSokKGBZHfgAYNZYNVS+k6rmKil7EsPznkXC6SjbFv4RBsxSmJsHhceiMQRwVCpXxnWk6xuhWCI55Qj/D34elESH9vL09rPZlxTMKCrhDGppF6RgVJQMiZLI60A8zMA2/D7691yQf3ZM7thcag3C27No3V5ea5BhZkmM04NcOvldmLHY/Nu/+CLj7MWBiRzbT7H3red9s8rBvNynv12GXY59g8DavQOr/Ft/+/5fFc063yY55x0QR6P/ZiwNhwId+cfenUCo+Frp8bBK42JwEqk39NxIVXLePRT2uRfxTxO3LYuVjen9NoFeVj2mAU/6pzMkmCnpsVh9Dbt+mvtzY/LHQ5a0h4vCpqfarGx9gqv3squChiFncvkt1TImQkKhJuHzXTPPL8rHQ7WORy/ufBcRhp8y0dWaxmbp9LHP6mNt6nPftdb425gF2+lTo9KnI6VPE6VOx06cSp0+lTp/KXD6lvXzzp1ze/VXg9ilt+etb6cpt8V9pv5DxY5/0jamb/BQ4HWSf3Dz7ue9dJPxD2OVDkZtB0n3aie+KHU6Bn/tWYRgTLjHu8nPfuJ12N9prjXT3b9zOulvNsdh0qd46HRW3mmOWL3P6lW77DovqU1n/b3nnsqjv9GfEJvfvep+SO3anPWVm/hhx+7bY7WOJy8futfeBDbvpXt/laqmDf8rlfnhw2gAPTsvsoc8if8YHp7vmoW+X8+967PO++pSL949O7+tL36cSIUnZE+oi/qm09yD797//G0E1yQe9DgQA"

@st.cache_data(show_spinner=False)
def _load_drug_ref():
    raw = _gzip.decompress(_b64.b64decode(_DRUG_REF_GZ))
    return _json.loads(raw)

def _get_drug_info(code: str) -> dict | None:
    ref = _load_drug_ref()
    code = str(code).strip()
    d = ref["drugs"].get(code)
    if d:
        return d
    # Try ATC3 prefix default
    atc3 = code[:3]
    return ref["atc3_defaults"].get(atc3)

# ══════════════════════════════════════════════════════════════════════════════
# COLUMN NORMALISATION PROFILER  (Data Prep tab)
# ══════════════════════════════════════════════════════════════════════════════

# Known system fields with their fingerprints
_SYSTEM_FIELDS = {
    "patient_id": {
        "label": "Patient / Member ID",
        "group": "Patient",
        "patterns": [r"rama", r"member.?id", r"patient.?id", r"benefi", r"insur.?no",
                     r"affil", r"no.*pat", r"num.*adh"],
        "dtype_hints": ["object", "string", "int"],
        "sample_hints": ["contains digits", "6-12 chars"],
        "required": True,
    },
    "patient_name": {
        "label": "Patient Full Name",
        "group": "Patient",
        "patterns": [r"patient.?name", r"nom.?patient", r"beneficiary.?name",
                     r"client.?name", r"full.?name", r"nom.*prenom"],
        "dtype_hints": ["object"],
        "required": False,
    },
    "visit_date": {
        "label": "Visit / Dispensing Date",
        "group": "Visit",
        "patterns": [r"dispensing.?date", r"visit.?date", r"date.?dispens",
                     r"service.?date", r"date.?service", r"date.?visite",
                     r"date.?soins", r"^date$"],
        "dtype_hints": ["datetime64", "object"],
        "required": True,
    },
    "voucher_id": {
        "label": "Voucher / Claim ID",
        "group": "Visit",
        "patterns": [r"paper.?code", r"voucher", r"claim.?id", r"claim.?no",
                     r"bon.?no", r"fiche.?no", r"reference", r"invoice.?no",
                     r"receipt.?no"],
        "dtype_hints": ["object", "string"],
        "required": False,
    },
    "doctor_name": {
        "label": "Prescriber / Doctor Name",
        "group": "Provider",
        "patterns": [r"practitioner.?name", r"doctor.?name", r"prescriber",
                     r"medecin", r"nom.?medecin", r"physician.?name",
                     r"provider.?name"],
        "dtype_hints": ["object"],
        "required": False,
    },
    "doctor_type": {
        "label": "Prescriber Type / Speciality",
        "group": "Provider",
        "patterns": [r"practitioner.?type", r"doctor.?type", r"specialit",
                     r"type.?medecin", r"provider.?type"],
        "dtype_hints": ["object"],
        "required": False,
    },
    "facility": {
        "label": "Health Facility",
        "group": "Provider",
        "patterns": [r"facility", r"hospital", r"clinic", r"pharmacy", r"pharmacie",
                     r"health.?center", r"centre.?sante", r"etablissement"],
        "dtype_hints": ["object"],
        "required": False,
    },
    "drug_code": {
        "label": "Drug / Medicine Code",
        "group": "Drug",
        "patterns": [r"drug.?code", r"medicine.?code", r"med.?code", r"item.?code",
                     r"product.?code", r"rhia.?code", r"atc.?code", r"ndc"],
        "dtype_hints": ["object", "string"],
        "required": False,
    },
    "drug_name": {
        "label": "Drug / Medicine Name",
        "group": "Drug",
        "patterns": [r"drug.?name", r"medicine.?name", r"medication", r"item.?name",
                     r"product.?name", r"denomination", r"generic.?name",
                     r"nom.?medicament"],
        "dtype_hints": ["object"],
        "required": False,
    },
    "quantity": {
        "label": "Quantity Dispensed",
        "group": "Drug",
        "patterns": [r"quantity", r"qty", r"quantite", r"qte", r"units?$",
                     r"nombre.?unite", r"dose.?qty"],
        "dtype_hints": ["int", "float", "int64", "float64"],
        "required": False,
    },
    "diagnosis": {
        "label": "Diagnosis / ICD Code",
        "group": "Clinical",
        "patterns": [r"diagnosis", r"diagnos", r"icd", r"condition",
                     r"disease.?code", r"code.?patho", r"pathology"],
        "dtype_hints": ["object"],
        "required": False,
    },
    "amount": {
        "label": "Total Amount (RWF)",
        "group": "Financial",
        "patterns": [r"total.?cost", r"total.?amount", r"montant.?total",
                     r"cout.?total", r"amount$", r"total$"],
        "dtype_hints": ["float64", "int64"],
        "required": False,
    },
    "insurance_copay": {
        "label": "Insurance Co-payment",
        "group": "Financial",
        "patterns": [r"insurance.?co.?pay", r"couverture", r"part.?assur",
                     r"rssb.?amount", r"rama.?amount", r"insurance.?share",
                     r"rama.?part"],
        "dtype_hints": ["float64"],
        "required": False,
    },
    "patient_copay": {
        "label": "Patient Co-payment",
        "group": "Financial",
        "patterns": [r"patient.?co.?pay", r"ticket.?moderateur",
                     r"patient.?share", r"part.?patient", r"copay"],
        "dtype_hints": ["float64"],
        "required": False,
    },
    "medicine_cost": {
        "label": "Medicine Cost",
        "group": "Financial",
        "patterns": [r"medicine.?cost", r"drug.?cost", r"med.?cost",
                     r"cout.?medicament"],
        "dtype_hints": ["float64"],
        "required": False,
    },
    "gender": {
        "label": "Gender",
        "group": "Patient",
        "patterns": [r"gender", r"sex$", r"sexe$"],
        "dtype_hints": ["object"],
        "required": False,
    },
    "patient_type": {
        "label": "Patient Type",
        "group": "Patient",
        "patterns": [r"patient.?type", r"type.?patient", r"beneficiary.?type",
                     r"categorie"],
        "dtype_hints": ["object"],
        "required": False,
    },
}

_GROUPS_ORDER = ["Patient", "Visit", "Provider", "Drug", "Clinical", "Financial"]
_GROUP_COLORS = {
    "Patient":   "#0ea5e9",
    "Visit":     "#00e5a0",
    "Provider":  "#a78bfa",
    "Drug":      "#f59e0b",
    "Clinical":  "#f97316",
    "Financial": "#22c55e",
}


def profile_column(df: pd.DataFrame, col: str) -> dict:
    """Compute rich stats for a column to help with mapping."""
    s = df[col]
    total = len(s)
    nulls = int(s.isna().sum())
    non_null = s.dropna()
    dtype_str = str(s.dtype)

    info = {
        "col":       col,
        "dtype":     dtype_str,
        "total":     total,
        "nulls":     nulls,
        "null_pct":  round(100 * nulls / max(total, 1), 1),
        "unique":    int(non_null.nunique()),
        "samples":   [],
        "is_numeric": pd.api.types.is_numeric_dtype(s),
        "is_date":   False,
        "looks_like_id": False,
        "looks_like_date": False,
        "looks_like_amount": False,
        "looks_like_drug_code": False,
    }

    # Sample values
    info["samples"] = [str(v) for v in non_null.head(5).tolist()]

    # Detect date
    if "datetime" in dtype_str:
        info["is_date"] = True
        info["looks_like_date"] = True
    elif dtype_str == "object" and len(non_null) > 0:
        try:
            test = pd.to_datetime(non_null.head(20), errors="coerce")
            if test.notna().mean() > 0.7:
                info["looks_like_date"] = True
        except Exception:
            pass

    # Looks like amount
    if info["is_numeric"] and len(non_null) > 0:
        med = float(non_null.median())
        if 100 < med < 10_000_000:
            info["looks_like_amount"] = True

    # Looks like ID
    if not info["is_numeric"] and len(non_null) > 0:
        sample_str = non_null.head(20).astype(str)
        alpha_ratio = sample_str.str.match(r"^[A-Za-z0-9\-/]+$").mean()
        avg_len = sample_str.str.len().mean()
        if alpha_ratio > 0.8 and 4 <= avg_len <= 20:
            info["looks_like_id"] = True

    # Looks like drug code (ATC pattern)
    if dtype_str == "object" and len(non_null) > 0:
        sample_str = non_null.head(20).astype(str)
        atc_match = sample_str.str.match(r"^[A-Z][0-9]{2}[A-Z]{2}").mean()
        if atc_match > 0.5:
            info["looks_like_drug_code"] = True

    return info


def score_column_vs_field(col_key: str, profile: dict,
                           field_name: str, field_def: dict) -> float:
    """Return 0-1 confidence that col maps to field."""
    score = 0.0

    # Pattern matching (strongest signal)
    for pat in field_def["patterns"]:
        if re.search(pat, col_key, re.IGNORECASE):
            score += 0.7
            break

    # Dtype hint
    for dh in field_def.get("dtype_hints", []):
        if dh in profile["dtype"] or profile["dtype"].startswith(dh):
            score += 0.15
            break

    # Semantic hints
    if field_name == "visit_date" and profile["looks_like_date"]:
        score += 0.2
    if field_name in ("amount", "insurance_copay", "patient_copay", "medicine_cost"):
        if profile["looks_like_amount"]:
            score += 0.15
    if field_name == "patient_id" and profile["looks_like_id"]:
        score += 0.1
    if field_name == "drug_code" and profile["looks_like_drug_code"]:
        score += 0.2

    return min(score, 1.0)


def auto_map_columns(df: pd.DataFrame) -> dict:
    """
    Auto-detect which df column maps to which system field.
    Returns {system_field: original_col_name} for confident matches,
    and a full scores dict {original_col: {field: score}}.
    """
    profiles = {}
    for col in df.columns:
        key = re.sub(r"[^a-z0-9]", "_", col.lower().strip())
        key = re.sub(r"_+", "_", key).strip("_")
        profiles[col] = profile_column(df, col)
        profiles[col]["col_key"] = key

    # Score all columns against all fields
    all_scores = {}   # col -> {field -> score}
    for col, prof in profiles.items():
        all_scores[col] = {}
        for field, fdef in _SYSTEM_FIELDS.items():
            all_scores[col][field] = score_column_vs_field(
                prof["col_key"], prof, field, fdef
            )

    # Greedy assignment: highest-score match wins, no field used twice
    mapping = {}           # system_field -> original_col
    reverse = {}           # original_col -> system_field
    assigned_fields = set()

    # Flatten to (score, col, field) sorted desc
    pairs = sorted(
        [(s, col, field)
         for col, scores in all_scores.items()
         for field, s in scores.items()
         if s >= 0.35],
        key=lambda x: -x[0]
    )

    for score, col, field in pairs:
        if field not in assigned_fields and col not in reverse:
            mapping[field] = col
            reverse[col] = field
            assigned_fields.add(field)

    return mapping, all_scores, profiles


def apply_column_mapping(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    """
    Rename df columns according to mapping {system_field: original_col}.
    Unknown columns get a raw_ prefix.
    Returns new df with system field names.
    """
    df = df.copy()
    rename = {}
    for field, orig_col in mapping.items():
        if orig_col in df.columns:
            rename[orig_col] = field

    # Rename known columns
    df = df.rename(columns=rename)

    # Prefix unknown remaining columns with raw_
    final_rename = {}
    for col in df.columns:
        if col not in _SYSTEM_FIELDS and not col.startswith("raw_"):
            final_rename[col] = f"raw_{col}"
    if final_rename:
        df = df.rename(columns=final_rename)

    return df


# ══════════════════════════════════════════════════════════════════════════════
# RULES ENGINE  (in-memory, no database required)
# ══════════════════════════════════════════════════════════════════════════════

# Diagnosis-Drug Blacklist (ICD prefix -> set of ATC prefixes to flag)
_DX_DRUG_BLACKLIST = {
    "B50": {"J01": (20,"Malaria+antibiotics: UCG first-line is ACT not J01 antibiotics"),
             "L01": (45,"Antineoplastic for malaria: no clinical basis"),
             "N05": (35,"Antipsychotic for malaria: no indication")},
    "B51": {"J01": (20,"Malaria (P.vivax) + antibiotics: ACT is first-line"),
             "L01": (45,"Antineoplastic for malaria: impossible")},
    "B54": {"J01": (20,"Malaria + antibiotics: ACT protocol not antibiotics"),
             "L01": (45,"Antineoplastic for unspecified malaria")},
    "I10": {"P01": (40,"Antihypertensive + antiparasitic: no clinical link"),
             "L01": (50,"Antineoplastic for hypertension: diagnosis fraud"),
             "N05": (30,"Antipsychotic for hypertension: no indication")},
    "E11": {"P01": (40,"T2DM + antiparasitic: no clinical indication"),
             "L01": (50,"Antineoplastic for diabetes: diagnosis fraud")},
    "E10": {"L01": (50,"Antineoplastic for T1DM: diagnosis fraud")},
    "J18": {"L01": (50,"Antineoplastic for pneumonia: no indication"),
             "N05": (35,"Antipsychotic for pneumonia: no indication")},
    "G40": {"P01": (40,"Epilepsy + antiparasitic: UCG uses CBZ/VPA/PHB"),
             "L01": (45,"Antineoplastic for epilepsy: no indication")},
    "A15": {"L01": (45,"Antineoplastic for TB: unless concurrent cancer"),
             "N05": (35,"Antipsychotic for TB: not in RHZE protocol")},
    "Z00": {"L01": (60,"CRITICAL: Antineoplastic on routine checkup"),
             "N05": (40,"Antipsychotic on routine checkup: billing fraud"),
             "H02": (30,"High-dose steroid on routine checkup")},
    "J06": {"L01": (55,"Antineoplastic for URTI: strong fraud signal"),
             "N05": (35,"Antipsychotic for URTI: no indication"),
             "S01": (25,"Ophthalmic prep for URTI: no indication")},
    "J00": {"L01": (55,"Antineoplastic for common cold: fraud"),
             "N05": (35,"Antipsychotic for common cold")},
    "O80": {"L01": (60,"CRITICAL: Antineoplastic during normal delivery"),
             "N05": (35,"Antipsychotic for normal delivery")},
    "Z23": {"L01": (60,"CRITICAL: Antineoplastic alongside vaccination"),
             "N05": (40,"Antipsychotic at vaccination visit")},
    "F20": {"P01": (40,"Antipsychotic Rx for schizophrenia needs N05, not P01")},
    "F32": {"P01": (40,"Depression + antiparasitic: no indication")},
}

# Prescriber code to speciality mapping
_PRESCRIBER_ALLOWED = {
    "D":       {"Dermatology","Dermatologist"},
    "OPHT":    {"Ophthalmology","Ophthalmologist"},
    "IM":      {"Internal Medicine","Internist","Physician"},
    "AC":      {"Oncology","Oncologist"},
    "UROL":    {"Urology","Urologist"},
    "GYN":     {"Gynaecology","Gynaecologist","Obstetrics","OB-GYN"},
    "GYNEC":   {"Gynaecology","Gynaecologist"},
    "PSYCH":   {"Psychiatry","Psychiatrist"},
    "CARDIOL": {"Cardiology","Cardiologist"},
    "DT":      {"Dentistry","Dental Surgeon"},
    "NEUROL":  {"Neurology","Neurologist"},
    "PED":     {"Paediatrics","Paediatrician","Pediatrics"},
    "NEPHR":   {"Nephrology","Nephrologist"},
    "SPEC":    {"*specialist*"},   # any registered specialist
    "HU":      {"*hospital*"},     # hospital/inpatient only
}


def _atc_prefix(code: str, n: int) -> str:
    return str(code).strip()[:n].upper()


def run_rules_engine(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Run all in-memory fraud rules against the dataframe.
    Returns (results_df, summary_stats).

    Results df has one row per original claim with:
      _score, _decision, _rules_fired, _risk_level
    """
    ref = _load_drug_ref()
    drugs_ref = ref["drugs"]
    atc3_ref  = ref["atc3_defaults"]

    # Identify available columns
    def _col(*names):
        for n in names:
            if n in df.columns:
                return n
        return None

    id_col    = _col("patient_id", "patient_name")
    date_col  = _col("visit_date")
    drug_col  = _col("drug_code")
    drug_nm   = _col("drug_name")
    qty_col   = _col("quantity")
    dx_col    = _col("diagnosis")
    doc_col   = _col("doctor_name")
    doc_type  = _col("doctor_type")
    fac_col   = _col("facility")
    amt_col   = _col("insurance_copay", "amount")
    vou_col   = _col("voucher_id")

    rows_out = []
    summary = {
        "total": len(df),
        "rule_counts": {f"R{i:02d}": 0 for i in range(1, 11)},
        "decisions": {"APPROVE": 0, "FLAG": 0, "HOLD": 0, "BLOCK": 0},
        "total_flagged_amount": 0.0,
        "rules_available": [],
    }

    # Determine which rules can run
    if drug_col:  summary["rules_available"] += ["R01","R02","R03","R04","R07","R08","R09"]
    if dx_col:    summary["rules_available"] += ["R02","R05","R06","R09"]
    if qty_col:   summary["rules_available"] += ["R03"]
    if date_col and id_col: summary["rules_available"] += ["R10"]
    summary["rules_available"] = sorted(set(summary["rules_available"]))

    # Pre-build refill index: {(patient_id, drug_code): [sorted dates]}
    refill_index = {}
    if id_col and drug_col and date_col:
        sub = df[[id_col, drug_col, date_col]].dropna()
        for _, row in sub.iterrows():
            key = (str(row[id_col]).strip(), str(row[drug_col]).strip())
            dt  = pd.to_datetime(row[date_col], errors="coerce")
            if pd.notna(dt):
                refill_index.setdefault(key, []).append(dt)
        for key in refill_index:
            refill_index[key].sort()

    for idx, row in df.iterrows():
        score    = 0
        fired    = []

        def fire(rule_id, s, reason, evidence=""):
            nonlocal score
            score += s
            fired.append({"id": rule_id, "score": s,
                          "reason": reason[:120],
                          "evidence": str(evidence)[:80]})
            summary["rule_counts"][rule_id] = summary["rule_counts"].get(rule_id, 0) + 1

        # Get row values
        d_code  = str(row[drug_col]).strip() if drug_col and pd.notna(row.get(drug_col)) else ""
        d_name  = str(row[drug_nm]).strip()  if drug_nm  and pd.notna(row.get(drug_nm))  else ""
        qty     = float(row[qty_col])        if qty_col  and pd.notna(row.get(qty_col))  else None
        dx      = str(row[dx_col]).strip()[:3].upper() if dx_col and pd.notna(row.get(dx_col)) else ""
        doc     = str(row[doc_type]).strip() if doc_type and pd.notna(row.get(doc_type)) else ""
        fac     = str(row[fac_col]).strip()  if fac_col  and pd.notna(row.get(fac_col))  else ""
        amt     = float(row[amt_col])        if amt_col  and pd.notna(row.get(amt_col))  else 0.0
        pid     = str(row[id_col]).strip()   if id_col   and pd.notna(row.get(id_col))   else ""
        vou     = str(row[vou_col]).strip()  if vou_col  and pd.notna(row.get(vou_col))  else ""

        # Look up drug info
        drug_info = drugs_ref.get(d_code)
        if not drug_info and d_code:
            drug_info = atc3_ref.get(d_code[:3])

        atc1 = (drug_info["atc1"] if drug_info else d_code[:1]).upper()
        atc3 = (drug_info["atc3"] if drug_info else d_code[:3]).upper()
        instr = (drug_info["instr"] if drug_info else "").strip()
        price = drug_info["price"] if drug_info else 0.0
        max_u = drug_info.get("max_units") if drug_info else None
        min_r = drug_info.get("min_refill") if drug_info else None

        # ── R01: Drug-Prescriber Mismatch ────────────────────────────────────
        if d_code and instr and doc:
            doc_up = doc.upper()
            instr_parts = {p.strip().upper() for p in re.split(r"[\s,]+", instr)
                           if p.strip()}
            # Hospital-use drugs at non-hospital providers
            if "HU" in instr_parts:
                if not any(x in doc_up for x in
                           ("HOSPITAL","INTERN","SPECIALIST","SPEC","SENIOR")):
                    fire("R01", 35, f"HU-restricted drug by non-hospital provider",
                         f"{d_code}|{instr}|{doc[:30]}")
            # PSYCH-restricted drugs by non-psychiatrist
            elif "PSYCH" in instr_parts:
                if not any(x in doc_up for x in ("PSYCH","NEUROL","SPECIALIST","SPEC")):
                    fire("R01", 25, f"PSYCH-restricted drug by {doc[:25]}",
                         f"{d_code}|{instr}")
            # AC (oncology) drugs
            elif "AC" in instr_parts and "AC" not in {"DAC"}:
                if not any(x in doc_up for x in ("ONCOL","CANCER","HAEMATOL","SPECIALIST","SPEC")):
                    fire("R01", 30, f"Oncology-only drug by non-oncologist {doc[:20]}",
                         f"{d_code}|{instr}")
            # OPHT drugs
            elif "OPHT" in instr_parts:
                if "OPHTH" not in doc_up and "EYE" not in doc_up and "SPEC" not in doc_up:
                    fire("R01", 20, f"OPHT drug by non-ophthalmologist",
                         f"{d_code}|{instr}|{doc[:20]}")

        # ── R02: Diagnosis-Drug Mismatch ─────────────────────────────────────
        if dx and d_code:
            if dx in _DX_DRUG_BLACKLIST:
                for atc_pref, (s, reason) in _DX_DRUG_BLACKLIST[dx].items():
                    if atc1 == atc_pref[:1] and (len(atc_pref) == 1 or
                                                   atc3.startswith(atc_pref)):
                        fire("R02", s, reason, f"ICD:{dx} + {d_code}({atc_pref})")
                        break

        # ── R03: Drug Quantity Excess ─────────────────────────────────────────
        if qty and max_u:
            if float(qty) > float(max_u):
                excess_pct = (float(qty) - float(max_u)) / float(max_u) * 100
                s = min(25 + int(excess_pct / 20) * 5, 60)
                fire("R03", s,
                     f"Quantity {qty:.0f} > max {max_u} ({excess_pct:.0f}% excess)",
                     f"{d_code}|qty:{qty}|max:{max_u}")

        # ── R04: Restricted Drug – High Value + No Indication ─────────────────
        if price > 50000 and dx:
            atc1_dx_ok = {
                "L": {"C", "D", "N", "G", "M"},   # antineoplastics need cancer/relevant dx
                "B": {"D", "N", "K"},              # erythropoietin needs haem/renal
            }
            if atc1 in atc1_dx_ok:
                expected_dx_firsts = atc1_dx_ok[atc1]
                if dx[:1] not in expected_dx_firsts:
                    fire("R04", 30,
                         f"High-value drug ({price:,.0f} RWF) with unrelated diagnosis {dx}",
                         f"{d_code}|price:{price:,.0f}|dx:{dx}")

        # ── R05: Antineoplastic Without Cancer Diagnosis ──────────────────────
        if atc1 == "L" and atc3.startswith("L01"):
            is_cancer_dx = (
                dx.startswith("C") or
                (dx.startswith("D") and dx[1:3].isdigit() and int(dx[1:3]) <= 49)
            )
            if dx and not is_cancer_dx:
                fire("R05", 25,
                     f"Cytotoxic drug without cancer diagnosis (ICD:{dx})",
                     f"{d_code}|dx:{dx}")

        # ── R06: Psychiatric Drug Without Mental Health Diagnosis ─────────────
        if "PSYCH" in instr.upper() and dx:
            is_mental = (dx.startswith("F") or
                         (dx.startswith("G4") and len(dx) >= 3 and "0" <= dx[2] <= "7"))
            if not is_mental:
                fire("R06", 20,
                     f"PSYCH drug without psychiatric/neuro diagnosis (ICD:{dx})",
                     f"{d_code}|instr:{instr}|dx:{dx}")

        # ── R07: Duplicate / Early Refill ─────────────────────────────────────
        if pid and d_code and min_r and date_col:
            key = (pid, d_code)
            dates = refill_index.get(key, [])
            if len(dates) >= 2:
                cur_dt = pd.to_datetime(row.get(date_col), errors="coerce")
                if pd.notna(cur_dt):
                    # Find most recent prior dispense
                    prior = [d for d in dates if d < cur_dt]
                    if prior:
                        gap = (cur_dt - max(prior)).days
                        if gap < min_r:
                            fire("R07", 40,
                                 f"Refill {gap}d after last dispense (min:{min_r}d)",
                                 f"{d_code}|gap:{gap}d|min:{min_r}d")

        # ── R08: Zero-Tariff Procedure or Unlisted Drug ───────────────────────
        if d_code and not drug_info and d_code.upper().startswith("RHIC"):
            fire("R08", 15,
                 f"Procedure {d_code} not found in RAMA tariff",
                 f"{d_code}")

        # ── R09: Malaria + Multiple Antibiotics ───────────────────────────────
        # (checked at patient-visit level below in post-processing)

        # ── R10: Immunosuppressant Without Valid Indication ───────────────────
        if atc3.startswith("L04") and dx:
            valid = any(dx.startswith(p) for p in
                        ("T86","M0","M1","M2","M3","K50","K51","K52","N04","L40","L41","G35"))
            if not valid:
                fire("R10", 20,
                     f"Immunosuppressant without transplant/autoimmune dx ({dx})",
                     f"{d_code}|dx:{dx}")

        # ── Scoring → decision ─────────────────────────────────────────────────
        if score >= 75:   decision = "BLOCK"
        elif score >= 50: decision = "HOLD"
        elif score >= 30: decision = "FLAG"
        else:             decision = "APPROVE"

        if score >= 75:   risk = "CRITICAL"
        elif score >= 50: risk = "HIGH"
        elif score >= 30: risk = "MEDIUM"
        else:             risk = "LOW"

        summary["decisions"][decision] += 1
        if decision in ("HOLD", "BLOCK"):
            summary["total_flagged_amount"] += amt

        rows_out.append({
            "_score":     score,
            "_risk":      risk,
            "_decision":  decision,
            "_rules_fired": "; ".join(
                f"{r['id']}(+{r['score']})" for r in fired
            ) if fired else "—",
            "_reasons":    " | ".join(r["reason"] for r in fired) if fired else "—",
            "_n_rules":    len(fired),
        })

    results = pd.DataFrame(rows_out, index=df.index)
    out_df = pd.concat([df, results], axis=1)

    summary["flagged_count"] = (
        summary["decisions"]["FLAG"] +
        summary["decisions"]["HOLD"] +
        summary["decisions"]["BLOCK"]
    )
    summary["rules_with_most_fires"] = sorted(
        [(k, v) for k, v in summary["rule_counts"].items() if v > 0],
        key=lambda x: -x[1]
    )[:5]

    return out_df, summary




def generate_counter_verification_xlsx(
    df: pd.DataFrame,
    deductions: list[dict],
    meta: dict,
    prepared_by: str,
    verified_by: str,
    approved_by: str,
    pc_col:  str | None = None,
    ins_col: str | None = None,
    tot_col: str | None = None,
    obs_col: str | None = None,
    dif_col: str | None = None,
) -> bytes:
    """
    Generate counter-verification report.
    All dimensions matched exactly from reference file.

    Sheet 1 — "After counter verification"
    Sheet 2 — "Counter verification report"
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    # ── Palette (exact hex from reference file) ───────────────────────────────
    C_BLUE      = "003366"   # Primary Blue
    C_GOLD      = "FFCC00"   # Accent Gold
    C_GREY      = "F4F4F4"   # Zebra / metadata bg
    C_WHITE     = "FFFFFF"
    C_TEXT      = "333333"   # Body text
    C_RED       = "C0392B"   # Deduction amounts / NO amounts
    C_GREEN     = "1E7E34"   # Verified YES text
    C_AMBER     = "B8860B"   # Verified NO text
    C_FILL_GREEN= "D4EDDA"   # YES background
    C_FILL_AMBER= "FFF3CD"   # NO background
    C_TITLE_BG  = "E8F0F7"   # Title banner background

    # ── Style helpers ─────────────────────────────────────────────────────────
    def fill(hex_col):
        return PatternFill("solid", fgColor=hex_col)

    def _font(name="Calibri", bold=False, size=11.0, color=C_TEXT):
        return Font(name=name, bold=bold, size=size, color=color)

    def side(style, color="000000"):
        return Side(border_style=style, color=color)

    THIN_GREY = side("thin",   "AAAAAA")
    MED_BLUE  = side("medium", C_BLUE)
    MED_GOLD  = side("medium", C_GOLD)
    THIN_ANY  = side("thin")          # thin with default colour (matches file's thin/?)
    NONE_S    = Side(border_style=None)

    def border_hdr_mid():
        """Header cell middle — medium blue all, gold bottom."""
        return Border(left=MED_BLUE, right=MED_BLUE, top=MED_BLUE, bottom=MED_GOLD)

    def border_hdr_first():
        """Header cell column A/first — thin left, medium rest."""
        return Border(left=THIN_ANY, right=MED_BLUE, top=MED_BLUE, bottom=MED_GOLD)

    def border_hdr_last():
        """Header cell last column — medium left, thin right."""
        return Border(left=MED_BLUE, right=THIN_ANY, top=MED_BLUE, bottom=MED_GOLD)

    def border_data():
        """Regular data cell borders."""
        return Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

    def border_data_first():
        """Data cell col A — thin/? left."""
        return Border(left=THIN_ANY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

    def border_data_last():
        """Data cell last col E — thin/? right."""
        return Border(left=THIN_GREY, right=THIN_ANY, top=THIN_GREY, bottom=THIN_GREY)

    # ── Alignments ────────────────────────────────────────────────────────────
    A_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    A_CENTER_NW = Alignment(horizontal="center", vertical="center", wrap_text=False)
    A_LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    A_LEFT_NW   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    A_LEFT_TOP  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    A_RIGHT     = Alignment(horizontal="right",  vertical="center", wrap_text=False)

    # ── Numeric helpers ───────────────────────────────────────────────────────
    def _safe_float(v):
        if v is None: return 0.0
        try:
            if pd.isna(v): return 0.0
        except Exception: pass
        try: return float(str(v).replace(",", "").replace(" ", ""))
        except ValueError: return 0.0

    def _safe_date(v):
        if v is None: return None
        try:
            if pd.isna(v): return None
        except Exception: pass
        return v

    def _get_col(row_, *keys, default=""):
        for k in keys:
            if k and k in row_.index:
                v = row_[k]
                try:
                    if pd.notna(v): return v
                except Exception:
                    if v is not None: return v
        return default

    # ── Deduction lookup ──────────────────────────────────────────────────────
    ded_map = {str(d["paper_code"]).strip(): d for d in deductions}

    wb = Workbook()

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 1 — "After counter verification"
    # ═════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "After counter verification"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A2"

    # Column widths — exact from reference file
    col_widths_s1 = [
        13.54,  # A  Paper Code
        17.91,  # B  Dispensing Date
        22.54,  # C  Patient Name
        24.82,  # D  RAMA Number
        28.36,  # E  Practitioner Name
        18.18,  # F  Health Facility
        19.73,  # G  Date of Treatment
        13.63,  # H  Verified
        23.45,  # I  Total Before Counter-V
        35.82,  # J  85% After Counter-V
        23.36,  # K  After Counter-V 100%
        20.82,  # L  After Counter-V 85%
        19.73,  # M  Amount Deducted
        28.00,  # N  Explanation
    ]
    for ci, w in enumerate(col_widths_s1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Row 1: Header ─────────────────────────────────────────────────────────
    # font: Calibri 11 bold white; fill: #003366;
    # border: medium/003366 all sides, bottom medium/FFCC00
    headers_s1 = [
        "Paper Code", "Dispensing Date", "Patient Name", "RAMA Number",
        "Practitioner Name", "Health Facility", "Date of Treatment", "Verified",
        "Total Before Counter-V (RWF)", "85% After Counter-V (RWF)",
        "After Counter-V 100%", "After Counter-V 85%",
        "Amount Deducted (RWF)", "Explanation",
    ]
    for ci, h in enumerate(headers_s1, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font      = _font("Calibri", bold=True, size=11.0, color=C_WHITE)
        c.fill      = fill(C_BLUE)
        c.border    = border_hdr_mid()
        c.alignment = A_CENTER
    ws1.row_dimensions[1].height = 33.75

    # ── Data validation: YES / NO dropdown col H ──────────────────────────────
    dv_verified = DataValidation(
        type="list", formula1='"YES,NO,PENDING"', allow_blank=True,
        showDropDown=False, showErrorMessage=True,
        errorTitle="Invalid value", error="Please choose YES, NO or PENDING",
    )
    ws1.add_data_validation(dv_verified)

    # ── Data rows ─────────────────────────────────────────────────────────────
    # font: Aptos Narrow 11 normal #333333; border: thin/AAAAAA all sides
    # row height: 46.25 (exact from reference)
    # zebra: even data_count (0-based) → white; odd → grey
    FONT_DATA  = _font("Aptos Narrow", bold=False, size=11.0, color=C_TEXT)
    FONT_YES   = _font("Aptos Narrow", bold=True,  size=11.0, color=C_GREEN)
    FONT_NO    = _font("Aptos Narrow", bold=True,  size=11.0, color=C_AMBER)
    FONT_RED   = _font("Aptos Narrow", bold=False, size=11.0, color=C_RED)

    data_count = 0
    for _, row in df.iterrows():
        ri = data_count + 2

        if pc_col and pc_col in row.index:
            pc = str(row[pc_col]).strip()
        else:
            pc = str(_get_col(row, "voucher_id", "paper_code", "Paper Code", default=""))

        ins_co     = _safe_float(_get_col(row, ins_col, "Insurance Co-payment",
                                          "insurance_copay", "insurance_co_payment"))
        ded        = ded_map.get(pc)
        is_ded     = ded is not None
        ded_amount = _safe_float(ded["amount"])      if is_ded else 0.0
        expla      = str(ded["explanation"]).strip()  if is_ded else ""
        verified   = "NO" if is_ded else "YES"

        # ded_amount is negative (e.g. -11342); use abs for arithmetic
        total_85  = round(ins_co * 0.85, 2)
        after_100 = round(ins_co - abs(ded_amount), 2)
        after_85  = round(after_100 * 0.85, 2)

        row_vals = [
            pc,
            _safe_date(_get_col(row, "Dispensing Date", "dispensing_date", "visit_date")),
            str(_get_col(row, "Patient Name",       "patient_name",      default="")),
            str(_get_col(row, "RAMA Number",        "rama_number", "patient_id", default="")),
            str(_get_col(row, "Practitioner Name",  "practitioner_name", "doctor_name", default="")),
            str(_get_col(row, "Health Facility",    "Health facility",   "facility",
                         default="PHARMACIE VINCA GISENYI LTD")),
            _safe_date(_get_col(row, "Dispensing Date", "dispensing_date", "visit_date")),
            verified,
            ins_co,
            total_85,
            after_100,
            after_85,
            ded_amount,
            expla,
        ]

        row_fill = fill(C_WHITE) if data_count % 2 == 0 else fill(C_GREY)
        ws1.row_dimensions[ri].height = 46.25

        for ci, val in enumerate(row_vals, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.border = border_data()

            if ci == 8:  # H — Verified
                c.font      = FONT_YES if verified == "YES" else FONT_NO
                c.fill      = fill(C_FILL_GREEN) if verified == "YES" else fill(C_FILL_AMBER)
                c.alignment = A_CENTER
                dv_verified.add(c)
            elif ci in (9, 10, 11, 12):  # Numeric amount cols
                c.font          = FONT_DATA
                c.fill          = row_fill
                c.number_format = "#,##0.00"
                c.alignment     = A_RIGHT
            elif ci == 13:  # M — Amount Deducted
                c.font          = FONT_RED if ded_amount != 0 else FONT_DATA
                c.fill          = row_fill
                c.number_format = "#,##0.00"
                c.alignment     = A_RIGHT
            elif ci in (2, 7):  # Date cols
                c.font          = FONT_DATA
                c.fill          = row_fill
                c.number_format = "dd/mm/yyyy"
                c.alignment     = A_CENTER
            elif ci == 14:  # N — Explanation
                c.font      = FONT_DATA
                c.fill      = row_fill
                c.alignment = A_LEFT
            else:
                c.font      = FONT_DATA
                c.fill      = row_fill
                c.alignment = A_LEFT

        data_count += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    # font: Calibri 11 bold white; fill: #003366
    # border: medium/003366 all sides (top = gold from reference, bottom = blue)
    if data_count > 0:
        tot_ri = data_count + 2
        ws1.row_dimensions[tot_ri].height = 21.75
        FONT_TOT = _font("Calibri", bold=True, size=11.0, color=C_WHITE)
        tot_border = Border(left=MED_BLUE, right=MED_BLUE, top=MED_GOLD, bottom=MED_BLUE)
        for ci in range(1, 15):
            c = ws1.cell(row=tot_ri, column=ci)
            c.font   = FONT_TOT
            c.fill   = fill(C_BLUE)
            c.border = tot_border
        ws1.cell(row=tot_ri, column=8, value="TOTAL").alignment = A_CENTER
        for ci, col_letter in [(9, "I"), (10, "J"), (11, "K"), (12, "L"), (13, "M")]:
            c = ws1.cell(row=tot_ri, column=ci,
                         value=f"=SUM({col_letter}2:{col_letter}{tot_ri - 1})")
            c.number_format = "#,##0.00"
            c.alignment     = A_RIGHT

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 2 — "Counter verification report"
    # ═════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Counter verification report")
    ws2.sheet_view.showGridLines = False

    # Column widths — exact from reference file
    ws2.column_dimensions["A"].width = 12.24
    ws2.column_dimensions["B"].width = 25.04
    ws2.column_dimensions["C"].width = 31.86
    ws2.column_dimensions["D"].width = 13.49
    ws2.column_dimensions["E"].width = 49.24

    # ── Rows 1-3: Title banner — A1:E3 merged ────────────────────────────────
    # font: Calibri 36 bold #003366; fill: #E8F0F7; h=center v=center wrap=True
    ws2.merge_cells("A1:E3")
    t = ws2["A1"]
    t.value     = "RSSB - COUNTER VERIFICATION REPORT"
    t.font      = _font("Calibri", bold=True, size=36.0, color=C_BLUE)
    t.fill      = fill(C_TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    t.border    = Border(left=THIN_ANY, right=THIN_ANY, top=THIN_ANY, bottom=NONE_S)
    for rn in [1, 2, 3]:
        ws2.row_dimensions[rn].height = 21.75

    # ── Row 4: Gold separator — height 3.75 ───────────────────────────────────
    ws2.row_dimensions[4].height = 3.75
    for ci in range(1, 6):
        ws2.cell(row=4, column=ci).fill = fill(C_GOLD)

    # ── Rows 5-9: Metadata ────────────────────────────────────────────────────
    # Label: col A — Calibri 13 bold #003366, h=left v=center, bL=thin
    # Value: col C — Calibri 13 bold #333333, h=left v=center
    # Col E last: bR=thin (right-side border of the block)
    meta_rows = [
        (5, "PROVINCE:",               meta.get("province", "")),
        (6, "ADMINISTRATIVE DISTRICT:", meta.get("district", "")),
        (7, "PHARMACY:",               meta.get("pharmacy", "")),
        (8, "PERIOD:",                 meta.get("period",   "")),
        (9, "CODE:",                   meta.get("code",     "")),
    ]
    for rn, label, value in meta_rows:
        ws2.row_dimensions[rn].height = 21.75
        lc = ws2.cell(row=rn, column=1, value=label)
        lc.font      = _font("Calibri", bold=True, size=13.0, color=C_BLUE)
        lc.fill      = fill(C_WHITE)
        lc.border    = Border(left=THIN_ANY)
        lc.alignment = A_LEFT_NW

        vc = ws2.cell(row=rn, column=3, value=value)
        vc.font      = _font("Calibri", bold=True, size=13.0, color=C_TEXT)
        vc.fill      = fill(C_WHITE)
        vc.alignment = A_LEFT_NW

        # Right-side border on col E
        ws2.cell(row=rn, column=5).border = Border(right=THIN_ANY)
        ws2.cell(row=rn, column=5).fill   = fill(C_WHITE)

    # ── Row 10: Spacer ────────────────────────────────────────────────────────
    ws2.row_dimensions[10].height = 9.75
    ws2.cell(row=10, column=1).border = Border(left=THIN_ANY)
    ws2.cell(row=10, column=5).border = Border(right=THIN_ANY)

    # ── Row 11: Table header ──────────────────────────────────────────────────
    # font: Calibri 10.5 bold white; fill: #003366; height: 36
    # A:  bL=thin, bR=medium/blue, bT=medium/blue, bBot=medium/gold
    # B-D: bL=medium/blue, bR=medium/blue, bT=medium/blue, bBot=medium/gold
    # E:  bL=medium/blue, bR=thin, bT=medium/blue, bBot=medium/gold
    ws2.row_dimensions[11].height = 36.0
    tbl_headers = [
        "No.", "Invoice ID\n(Paper Code)",
        "Beneficiary RAMA No.",
        "Amount Deducted (RWF)",
        "Explanation of Deduction",
    ]
    hdr_borders = [
        border_hdr_first(),  # A
        border_hdr_mid(),    # B
        border_hdr_mid(),    # C
        border_hdr_mid(),    # D
        border_hdr_last(),   # E
    ]
    for ci, (h, bdr) in enumerate(zip(tbl_headers, hdr_borders), 1):
        c = ws2.cell(row=11, column=ci, value=h)
        c.font      = _font("Calibri", bold=True, size=10.5, color=C_WHITE)
        c.fill      = fill(C_BLUE)
        c.border    = bdr
        c.alignment = A_CENTER

    # ── Rows 12+: Deduction data rows ─────────────────────────────────────────
    # font: Aptos Narrow 10.5 normal #333333; height: 18.0
    # zebra: i=0,2,4 → white; i=1,3,5 → grey
    # A: h=center v=center wrap, bL=thin/?, bR=thin/AAA, bT=thin/AAA, bBot=thin/AAA
    # B,C: h=left v=center wrap, thin/AAA all
    # D: h=right v=center no-wrap, color=red C0392B, numfmt=#,##0, thin/AAA all
    # E: h=left v=top wrap, bL=thin/AAA, bR=thin/?, bT=thin/AAA, bBot=thin/AAA
    n_ded      = len(deductions)
    n_rows     = n_ded          # rows grow with data — no padding
    data_start = 12

    FONT_D2 = _font("Aptos Narrow", bold=False, size=10.5, color=C_TEXT)
    FONT_DR = _font("Aptos Narrow", bold=False, size=10.5, color=C_RED)

    for i, d in enumerate(deductions):
        ri = data_start + i
        ws2.row_dimensions[ri].height = 18.0
        row_fill = fill(C_WHITE) if i % 2 == 0 else fill(C_GREY)

        vals = [
            i + 1,
            str(d.get("paper_code",  "")),
            str(d.get("rama_no",     "")),
            _safe_float(d.get("amount", 0)),
            str(d.get("explanation", "")),
        ]

        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.fill = row_fill

            if ci == 1:
                c.font      = FONT_D2
                c.border    = border_data_first()
                c.alignment = A_CENTER
            elif ci in (2, 3):
                c.font      = FONT_D2
                c.border    = border_data()
                c.alignment = A_LEFT
            elif ci == 4:
                c.font          = FONT_DR
                c.border        = border_data()
                c.number_format = "#,##0;-#,##0"
                c.alignment     = A_RIGHT
            elif ci == 5:
                c.font      = FONT_D2
                c.border    = border_data_last()
                c.alignment = A_LEFT_TOP

    # ── Total row ─────────────────────────────────────────────────────────────
    # font: Calibri 10.5 bold white; fill: #003366; height: 21.75
    # A: bL=thin, bBot=medium/blue
    # B: bL=medium/blue, bBot=medium/blue
    # C: h=right v=center, value="TOTAL AMOUNT DEDUCTED", bL=medium/blue, bBot=medium/blue
    # D: h=right v=center, numfmt=#,##0, formula, bL=medium/blue, bBot=medium/blue
    # E: bL=medium/blue, bBot=medium/blue
    tot_row = data_start + n_rows
    ws2.row_dimensions[tot_row].height = 21.75
    FONT_TOT2 = _font("Calibri", bold=True, size=10.5, color=C_WHITE)
    tot_borders = [
        Border(left=THIN_ANY,  bottom=MED_BLUE),  # A
        Border(left=MED_BLUE,  bottom=MED_BLUE),  # B
        Border(left=MED_BLUE,  bottom=MED_BLUE),  # C
        Border(left=MED_BLUE,  bottom=MED_BLUE),  # D
        Border(left=MED_BLUE,  bottom=MED_BLUE),  # E
    ]
    for ci, bdr in enumerate(tot_borders, 1):
        c = ws2.cell(row=tot_row, column=ci)
        c.font   = FONT_TOT2
        c.fill   = fill(C_BLUE)
        c.border = bdr

    ws2.cell(row=tot_row, column=3,
             value="TOTAL AMOUNT DEDUCTED").alignment = A_RIGHT
    tot_c = ws2.cell(row=tot_row, column=4,
                     value=f"=SUM(D{data_start}:D{tot_row - 1})")
    tot_c.number_format = "#,##0;-#,##0"
    tot_c.alignment     = A_RIGHT

    # ── Signature block ───────────────────────────────────────────────────────
    # Spacer row after total: height=12
    # Rows +1 to +5: cols A, C, E; height=19.5
    # Rows +6 to +10: height=15 (empty trailing rows)
    sig_start = tot_row + 2
    ws2.row_dimensions[sig_start - 1].height = 12.0

    for rn in range(sig_start, sig_start + 5):
        ws2.row_dimensions[rn].height = 19.5
    for rn in range(sig_start + 5, sig_start + 10):
        ws2.row_dimensions[rn].height = 15.0

    # Row 0 (+0): "Prepared by" / "Verified by" / "Approved By" — bold blue, underline
    sig_data = [
        # (offset, [(col, value, bold, color)])
        (0, [(1, "Prepared by",  True,  C_BLUE),
             (3, "Verified by",  True,  C_BLUE),
             (5, "Approved By",  True,  C_BLUE)]),
        (1, [(1, "Date:",        False, C_TEXT),
             (3, "Date:",        False, C_TEXT),
             (5, "Date:",        False, C_TEXT)]),
        (2, [(1, "Signature:",   False, C_TEXT),
             (3, "Signature:",   False, C_TEXT),
             (5, "Signature:",   False, C_TEXT)]),
        (3, [(1, "Names:",       False, C_TEXT),
             (3, f"Names:  {verified_by}", True, C_BLUE),
             (5, "Names:",       False, C_TEXT)]),
        (4, [(3, "Lead, Counter Verification", False, C_TEXT),
             (5, "Responsible of Pharmacy",    False, C_TEXT)]),
    ]
    for offset, cells in sig_data:
        rn = sig_start + offset
        for col, text, bold, color in cells:
            c = ws2.cell(row=rn, column=col, value=text)
            c.font      = _font("Calibri", bold=bold, size=10.5, color=color)
            c.alignment = A_LEFT_NW
            if offset == 0:
                c.border = Border(bottom=THIN_ANY)

    # ── Output ────────────────────────────────────────────────────────────────
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()




# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
<div style='margin-bottom:6px'>
  <div style='font-family:Syne,sans-serif;font-size:22px;font-weight:800;color:#e2e8f0;
       letter-spacing:-0.5px'>💊 Pharma<span style='color:#00e5a0'>Scan</span></div>
  <div style='font-size:11px;color:#475569;font-family:monospace;
       letter-spacing:.06em;text-transform:uppercase'>Voucher Intelligence · RSSB</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<hr style='border-color:#1e2a38;margin:10px 0 14px'>", unsafe_allow_html=True)

    # ── File uploader ─────────────────────────────────────────────────────────
    st.markdown("""<div style='font-size:11px;font-weight:700;color:#64748b;
    text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px'>
    📂 Source File</div>""", unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Upload voucher report",
        type=["csv", "xlsx", "xls", "ods"],
        help="Supports CSV, Excel (.xlsx/.xls), and ODS. Column mapping happens in the Data Prep tab.",
        label_visibility="collapsed",
    )

    st.markdown("<hr style='border-color:#1e2a38;margin:14px 0'>", unsafe_allow_html=True)

    # ── Data Lake status panel ────────────────────────────────────────────────
    st.markdown("""<div style='font-size:11px;font-weight:700;color:#64748b;
    text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px'>
    🗄️ Data Lake</div>""", unsafe_allow_html=True)

    _dl = st.session_state.get("data_lake", {})
    _dl_committed = _dl.get("committed", False)

    if _dl_committed:
        _dl_rows    = len(_dl["df"])
        _dl_cols    = len(_dl["df"].columns)
        _dl_fname   = _dl.get("filename", "—")[:28]
        _dl_ts      = _dl.get("committed_at", "—")
        _dl_fields  = _dl.get("mapped_fields", [])
        st.markdown(f"""
<div style='background:#031a0a;border:1px solid #14532d;border-radius:10px;
     padding:12px 14px;font-family:monospace;font-size:11px'>
  <div style='color:#22c55e;font-weight:700;margin-bottom:6px'>● ACTIVE</div>
  <div style='color:#64748b'>File: <span style='color:#e2e8f0'>{_dl_fname}</span></div>
  <div style='color:#64748b'>{_dl_rows:,} rows · {_dl_cols} columns</div>
  <div style='color:#64748b'>Committed: <span style='color:#94a3b8'>{_dl_ts}</span></div>
  <div style='margin-top:6px;display:flex;flex-wrap:wrap;gap:3px'>
    {''.join(f"<span style='background:rgba(0,229,160,.08);border:1px solid rgba(0,229,160,.2);border-radius:4px;padding:1px 6px;color:#00e5a0;font-size:10px'>{f}</span>" for f in _dl_fields[:8])}
  </div>
</div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Clear Data Lake", use_container_width=True, type="secondary"):
            for k in ["data_lake","dp_result","dp_map_used","raw_bytes","raw_filename"]:
                st.session_state.pop(k, None)
            st.rerun()
    else:
        st.markdown("""
<div style='background:#0d1117;border:1px dashed #1e2a38;border-radius:10px;
     padding:12px 14px;font-family:monospace;font-size:11px;text-align:center'>
  <div style='color:#475569;margin-bottom:4px'>● EMPTY</div>
  <div style='color:#334155;font-size:10px'>Upload a file and complete<br>Data Prep to activate</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<hr style='border-color:#1e2a38;margin:14px 0'>", unsafe_allow_html=True)

    # ── Quick settings (only when data lake active) ───────────────────────────
    if _dl_committed:
        st.markdown("""<div style='font-size:11px;font-weight:700;color:#64748b;
        text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px'>
        ⚙️ Quick Settings</div>""", unsafe_allow_html=True)
        show_raw = st.checkbox("Show raw column names", value=False)
        top_n    = st.slider("Top N for charts", 5, 25,
                             _dl.get("top_n", 15), key="sb_top_n")
    else:
        show_raw = False
        top_n    = 15

# ── Store raw bytes when file is uploaded ─────────────────────────────────────
if uploaded is not None:
    _raw_bytes_new = uploaded.read()
    _new_name = uploaded.name
    if st.session_state.get("raw_filename") != _new_name:
        st.session_state["raw_bytes"]    = _raw_bytes_new
        st.session_state["raw_filename"] = _new_name
        # New file → clear data lake so user re-runs Data Prep
        for _k in ["data_lake", "dp_result", "dp_map_used", "normalised_df",
                   "normalised_col", "normalised_map", "ann_df", "ann_detected"]:
            st.session_state.pop(_k, None)
        st.rerun()
    elif "raw_bytes" not in st.session_state:
        st.session_state["raw_bytes"]    = _raw_bytes_new
        st.session_state["raw_filename"] = _new_name

# ── Landing page (no file yet) ────────────────────────────────────────────────
if "raw_bytes" not in st.session_state:
    st.markdown("""
<div style='text-align:center;padding:60px 20px 30px'>
  <div style='font-size:52px;margin-bottom:14px'>💊</div>
  <div style='font-family:Syne,sans-serif;font-size:40px;font-weight:800;color:#e2e8f0;
       margin-bottom:6px;letter-spacing:-1px'>
    Pharma<span style='color:#00e5a0'>Scan</span>
  </div>
  <div style='color:#475569;font-size:14px;font-family:monospace;
       letter-spacing:.1em;text-transform:uppercase;margin-bottom:40px'>
    RSSB Pharmacy Voucher Intelligence
  </div>
</div>""", unsafe_allow_html=True)

    _f1, _f2, _f3, _f4 = st.columns(4)
    for _col, _icon, _title, _desc in zip(
        [_f1, _f2, _f3, _f4],
        ["🗂️", "🔍", "🕸️", "🛡️"],
        ["Data Prep & Lake", "Cross-Facility Match", "Network Graph", "Rules Engine"],
        [
            "Map columns from any file format once — all analysis reads from your committed data lake",
            "Fuzzy RAMA + name matching against clinic/hospital visit files to surface phantom claims",
            "Explore patient-doctor-facility relationships as an interactive network",
            "17 rules scoring every claim: quantity excess, diagnosis mismatches, rapid refills and more",
        ],
    ):
        with _col:
            st.markdown(f"""
<div style='background:#111720;border:1px solid #1e2a38;border-radius:14px;
            padding:22px 18px;text-align:center;min-height:160px;
            transition:border-color .2s'>
  <div style='font-size:30px;margin-bottom:10px'>{_icon}</div>
  <div style='font-weight:700;color:#e2e8f0;margin-bottom:6px;
       font-family:Syne,sans-serif;font-size:14px'>{_title}</div>
  <div style='font-size:11px;color:#475569;line-height:1.6'>{_desc}</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""
<div style='background:#030d1a;border:1px solid #1e3a5f;border-radius:12px;
     padding:16px 22px;text-align:center;font-family:monospace'>
  <span style='color:#38bdf8;font-size:13px'>
    👈 Upload a pharmacy voucher report in the sidebar to get started
  </span><br>
  <span style='color:#334155;font-size:11px'>
    Supports CSV · XLSX · XLS · ODS — column mapping adapts to any format
  </span>
</div>""", unsafe_allow_html=True)
    st.stop()

# ── Data lake unpack (available to all tabs after commit) ─────────────────────
_dl_committed = st.session_state.get("data_lake", {}).get("committed", False)

if _dl_committed:
    _dl        = st.session_state["data_lake"]
    df         = _dl["df"]
    s          = _dl["stats"]
    rapid      = _dl["rapid"]
    rapid_days = _dl["rapid_days"]
    repeat_groups  = _dl["repeat_groups"]
    repeat_detail  = _dl["repeat_detail"]
    col_map        = _dl["col_map"]
    top_n          = st.session_state.get("sb_top_n", _dl.get("top_n", 15))
else:
    # Minimal defaults so tabs don't crash if accessed before commit
    df = pd.DataFrame()
    s  = {}
    rapid = []
    rapid_days = 7
    repeat_groups = []
    repeat_detail = pd.DataFrame()
    col_map = {}

# ── Helper: render a locked-tab placeholder ───────────────────────────────────
def _render_tab_locked(tab_name: str = ""):
    st.markdown(f"""
<div style='display:flex;flex-direction:column;align-items:center;justify-content:center;
     padding:60px 20px;text-align:center'>
  <div style='font-size:40px;margin-bottom:16px;opacity:.4'>🗄️</div>
  <div style='font-family:Syne,sans-serif;font-size:20px;font-weight:700;
       color:#334155;margin-bottom:8px'>Data Lake Not Yet Active</div>
  <div style='font-size:13px;color:#1e3a5f;font-family:monospace;max-width:380px;
       line-height:1.7'>
    Complete the <b style='color:#0ea5e9'>Data Prep</b> wizard and click
    <b style='color:#00e5a0'>Commit to Data Lake</b> to unlock
    <b style='color:#e2e8f0'>{tab_name}</b> and all other analysis tabs.
  </div>
</div>""", unsafe_allow_html=True)

# ── Tabs ──────────────────────────────────────────────────────────────────────
(tab_dataprep, tab_summary, tab_records, tab_repeat,
 tab_network, tab_norm, tab_xfac, tab_cv) = st.tabs([
    "🗂️ Data Prep",
    "📊 Summary",
    "📋 All Records",
    f"🔁 Repeat Patients  {'🟡' if repeat_groups or rapid else '🟢'}  {len(repeat_groups)}",
    "🕸️ Network Graph",
    "✏️ Normalise Names",
    "🏥 Cross-Facility Match",
    "📄 Counter-Verification",
])

# ══ SUMMARY ══════════════════════════════════════════════════════════════════
with tab_summary:
  if not _dl_committed:
    _render_tab_locked("Summary")
  else:
    c = st.columns(4)
    c[0].metric("Total Records",         f"{s['total_rows']:,}")
    c[1].metric("Unique Patients",       f"{s['unique_patients']:,}" if "unique_patients" in s else "—")
    c[2].metric("Repeat Patients",       f"{s['repeat_patients']:,}" if "repeat_patients" in s else "—")
    c[3].metric("Rapid Revisits",        str(len(rapid)), delta=f"≤{rapid_days} day window")

    c2 = st.columns(4)
    c2[0].metric("Unique Practitioners", f"{s['unique_doctors']:,}" if "unique_doctors" in s else "—")
    c2[1].metric("Max Visits / Patient", str(s.get("max_visits", "—")))
    c2[2].metric("Total Cost (RWF)",     fmt_number(s["total_amount"]) if "total_amount" in s else "—")
    c2[3].metric("Avg Cost / Visit",     fmt_number(s["avg_amount"])   if "avg_amount"   in s else "—")

    if "date_min" in s:
        st.markdown(
            f'<p style="font-size:12px;color:#64748b;font-family:monospace;margin:8px 0 20px">'
            f'📅 {s["date_min"]} — {s["date_max"]}</p>',
            unsafe_allow_html=True,
        )

    fig_t = time_series_chart(df)
    if fig_t:
        st.pyplot(fig_t, use_container_width=True); plt.close(fig_t)

    left, right = st.columns(2)
    with left:
        if "top_patients" in s:
            td = s["top_patients"].head(top_n)
            colors = [DANGER if v >= 10 else WARN if v >= 5 else ACCENT for v in td["visits"]]
            fig = hbar_chart([str(x)[:22] for x in td["id"]], td["visits"].tolist(),
                             colors, "Top Patients (RAMA No.) by Visit Count", "Visits")
            st.pyplot(fig, use_container_width=True); plt.close(fig)
    with right:
        if "top_doctors" in s:
            td = s["top_doctors"].head(top_n)
            fig = hbar_chart([str(x)[:22] for x in td["doctor"]], td["visits"].tolist(),
                             ACCENT2, "Top Practitioners by Visit Volume", "Visits")
            st.pyplot(fig, use_container_width=True); plt.close(fig)

    # Practitioner type breakdown
    if "doctor_type" in df.columns:
        dt_vc = df["doctor_type"].value_counts().head(top_n)
        fig = hbar_chart(
            [str(x)[:30] for x in dt_vc.index],
            dt_vc.values.tolist(),
            PURPLE, "Visits by Practitioner Type", "Visits",
        )
        st.pyplot(fig, use_container_width=True); plt.close(fig)

    # Gender & patient type pie charts
    gc1, gc2 = st.columns(2)
    for target_col, label, target_col_obj in [
        (gc1, "Gender Breakdown",       "gender"),
        (gc2, "Patient Type Breakdown", "patient_type"),
    ]:
        if target_col_obj in df.columns:
            vc = df[target_col_obj].value_counts()
            fig, ax = plt.subplots(figsize=(4, 3))
            ax.pie(vc.values, labels=vc.index,
                   colors=[ACCENT, ACCENT2, PURPLE, WARN, DANGER][:len(vc)],
                   autopct="%1.1f%%", pctdistance=0.8,
                   textprops={"color": TEXT, "fontsize": 9},
                   wedgeprops={"linewidth": 1.5, "edgecolor": CARD})
            ax.set_title(label, fontsize=11, fontweight="bold", color=TEXT, pad=10)
            with target_col:
                st.pyplot(fig, use_container_width=True); plt.close(fig)

# ══ ALL RECORDS ═══════════════════════════════════════════════════════════════
with tab_records:
  if not _dl_committed:
    _render_tab_locked("All Records")
  else:
    st.markdown(f'<div class="sec-head">All Records — {len(df):,} rows</div>', unsafe_allow_html=True)
    search = st.text_input("🔍 Filter rows", key="rec_search", placeholder="Type to search any column…")
    display_df = df.copy()
    if not show_raw:
        display_df.columns = [c.replace("_", " ").title() for c in display_df.columns]
    if search:
        mask = display_df.apply(lambda col: col.astype(str).str.contains(search, case=False, na=False)).any(axis=1)
        display_df = display_df[mask]
        st.caption(f"{len(display_df):,} matching rows")
    st.dataframe(display_df, use_container_width=True, height=520)

# ══ REPEAT PATIENTS ══════════════════════════════════════════════════════════
# ══ REPEAT PATIENTS & RAPID REVISITS ═════════════════════════════════════════
with tab_repeat:
  if not _dl_committed:
    _render_tab_locked("Repeat Patients")
  else:
    # ── Sub-tab style: two internal sections via radio ─────────────────────
    _rp_section = st.radio(
        "View",
        ["🔁 Repeat Patients", "⚡ Rapid Revisits"],
        horizontal=True,
        label_visibility="collapsed",
        key="rp_section",
    )
    st.markdown("<hr style='border-color:#1e2a38;margin:8px 0 20px'>",
                unsafe_allow_html=True)

    # ────────────────────────────────────────────────────────────────────────
    # SECTION A — Repeat Patients
    # ────────────────────────────────────────────────────────────────────────
    if _rp_section == "🔁 Repeat Patients":
        if not repeat_groups:
            st.success("✅ No patients with multiple visits detected.")
        else:
            # KPI strip
            _rp1, _rp2, _rp3, _rp4 = st.columns(4)
            _rp1.metric("Repeat Patients",       len(repeat_groups))
            _rp2.metric("Max Visits / Patient",  s.get("max_visits", "—"))
            _heavy = sum(1 for g in repeat_groups if g["visits"] >= 5)
            _rp3.metric("High-frequency (≥5)",   _heavy)
            _very_heavy = sum(1 for g in repeat_groups if g["visits"] >= 10)
            _rp4.metric("Very high (≥10)",        _very_heavy,
                        delta=f"⚠ possible fraud" if _very_heavy else None,
                        delta_color="inverse" if _very_heavy else "normal")

            # Distribution chart
            _visit_counts = [g["visits"] for g in repeat_groups]
            if len(set(_visit_counts)) > 1:
                _fig_rp, _ax_rp = plt.subplots(figsize=(10, 3))
                _bins = list(range(2, max(_visit_counts) + 2))
                _, _bouts, _patches_rp = _ax_rp.hist(
                    _visit_counts, bins=_bins,
                    color=ACCENT, edgecolor=CARD, rwidth=0.8,
                )
                for _p, _l in zip(_patches_rp, _bouts):
                    if _l >= 10: _p.set_facecolor(DANGER)
                    elif _l >= 5: _p.set_facecolor(WARN)
                _ax_rp.set_xlabel("Visits per Patient")
                _ax_rp.set_ylabel("Patients")
                _ax_rp.set_title("Distribution of Repeat Visit Counts",
                                  fontsize=11, fontweight="bold", color=TEXT, pad=10)
                _ax_rp.spines[["top","right"]].set_visible(False)
                _ax_rp.grid(axis="y", alpha=0.3)
                _fig_rp.tight_layout()
                st.pyplot(_fig_rp, use_container_width=True)
                plt.close(_fig_rp)

            # Top offenders highlighted at the top
            _top_rp = sorted(repeat_groups, key=lambda x: -x["visits"])[:5]
            if _top_rp and _top_rp[0]["visits"] >= 5:
                st.markdown(
                    '<div class="sec-head">🚨 Top Repeat Visitors</div>',
                    unsafe_allow_html=True,
                )
                _tp_cols = st.columns(min(5, len(_top_rp)))
                for _tc, _tg in zip(_tp_cols, _top_rp):
                    _id_col = s.get("patient_col", "patient_id")
                    _tid  = str(_tg.get(_id_col, _tg.get("patient_id", "—")))[:16]
                    _tnam = str(_tg.get("patient_name", ""))[:20]
                    _tvis = _tg["visits"]
                    _tcol = DANGER if _tvis >= 10 else WARN if _tvis >= 5 else ACCENT
                    _tc.markdown(f"""
<div style='background:#111720;border:1px solid #1e2a38;
     border-left:3px solid {_tcol};border-radius:8px;padding:12px 14px;
     text-align:center'>
  <div style='font-size:22px;font-weight:800;color:{_tcol};
       font-family:Syne,sans-serif'>{_tvis}</div>
  <div style='font-size:10px;color:#64748b;font-family:monospace'>visits</div>
  <div style='font-size:11px;color:#e2e8f0;margin-top:4px;font-weight:600'>
    {_tnam or _tid}</div>
  <div style='font-size:9px;color:#475569;font-family:monospace'>{_tid}</div>
</div>""", unsafe_allow_html=True)

            # Searchable group table
            st.markdown(
                '<div class="sec-head">Patient Visit Groups</div>',
                unsafe_allow_html=True,
            )
            _grp_c1, _grp_c2, _grp_c3 = st.columns([3, 1, 1])
            with _grp_c1:
                _rp_srch = st.text_input(
                    "🔍 Filter", key="rep_search",
                    placeholder="Name, RAMA number, date…",
                )
            with _grp_c2:
                _rp_min_visits = st.number_input(
                    "Min visits", min_value=2, value=2, step=1, key="rep_min_vis",
                )
            with _grp_c3:
                _rp_max_show = st.number_input(
                    "Max rows", min_value=50, value=300, step=50, key="rep_max_rows",
                )

            _grp_df = pd.DataFrame(repeat_groups)
            if _rp_srch:
                _mask = _grp_df.apply(
                    lambda c: c.astype(str).str.contains(_rp_srch, case=False, na=False)
                ).any(axis=1)
                _grp_df = _grp_df[_mask]
            _grp_df = _grp_df[_grp_df["visits"] >= _rp_min_visits].head(_rp_max_show)

            def _highlight_v(val):
                try:
                    v = int(val)
                    if v >= 10: return "color:#ef4444;font-weight:bold"
                    if v >= 5:  return "color:#f59e0b;font-weight:bold"
                except Exception: pass
                return ""

            st.caption(f"{len(_grp_df):,} patient group(s) shown · "
                       f"≥5 visits: {sum(1 for g in repeat_groups if g['visits']>=5)} · "
                       f"≥10 visits: {sum(1 for g in repeat_groups if g['visits']>=10)}")
            st.dataframe(
                _grp_df.style.applymap(_highlight_v, subset=["visits"]),
                use_container_width=True, height=360,
            )

            # Detailed records expander
            with st.expander("📋 Detailed repeat visit records", expanded=False):
                st.dataframe(repeat_detail, use_container_width=True, height=400)

            # Prescriber breakdown
            if "doctor" in _grp_df.columns or "doctor_name" in df.columns:
                with st.expander("👨‍⚕️ Prescriber involvement in repeat visits",
                                 expanded=False):
                    _dcol = "doctor_name" if "doctor_name" in df.columns else None
                    _id_col_df = s.get("patient_col", "patient_id")
                    if _dcol and _id_col_df in df.columns:
                        _repeat_ids = set(
                            str(g.get(_id_col_df, g.get("patient_id","")))
                            for g in repeat_groups if g["visits"] >= 3
                        )
                        _rp_doc_df = df[df[_id_col_df].astype(str).isin(_repeat_ids)]
                        if not _rp_doc_df.empty:
                            _doc_counts = (
                                _rp_doc_df.groupby(_dcol)[_id_col_df]
                                .nunique()
                                .reset_index()
                                .rename(columns={_id_col_df: "Unique Patients",
                                                 _dcol: "Prescriber"})
                                .sort_values("Unique Patients", ascending=False)
                                .head(20)
                            )
                            st.dataframe(_doc_counts, use_container_width=True,
                                         height=280)

            # Download
            _dl_buf = repeat_detail.to_csv(index=False).encode()
            st.download_button(
                "⬇️ Download repeat visit records",
                data=_dl_buf,
                file_name="pharmascan_repeat_patients.csv",
                mime="text/csv",
                key="dl_repeat",
            )

    # ────────────────────────────────────────────────────────────────────────
    # SECTION B — Rapid Revisits
    # ────────────────────────────────────────────────────────────────────────
    else:
        if not rapid:
            st.success(f"✅ No rapid revisits detected within {rapid_days} days.")
        else:
            # KPI strip
            _rv1, _rv2, _rv3, _rv4 = st.columns(4)
            _rv1.metric("Total Rapid Cases",     len(rapid))
            _critical = sum(1 for r in rapid if r["days_apart"] <= 2)
            _rv2.metric("Critical (≤2 days)",   _critical,
                        delta="⚠ same-day / next-day" if _critical else None,
                        delta_color="inverse" if _critical else "normal")
            _avg_d = sum(r["days_apart"] for r in rapid) / len(rapid)
            _rv3.metric("Avg Days Apart",        f"{_avg_d:.1f}")
            _same_doc = sum(
                1 for r in rapid
                if r.get("doctor","—") not in ("—","")
                and len({r["doctor"]}) == 1
            )
            _rv4.metric("Rapid revisit window",  f"≤{rapid_days} days")

            # Histogram
            _fig_rh = rapid_histogram(rapid)
            if _fig_rh:
                st.pyplot(_fig_rh, use_container_width=True)
                plt.close(_fig_rh)

            # Filter controls
            _rfil1, _rfil2, _rfil3 = st.columns([3, 1, 1])
            with _rfil1:
                _rv_srch = st.text_input(
                    "🔍 Filter", key="rv_srch",
                    placeholder="Name, RAMA number, doctor…",
                )
            with _rfil2:
                _rv_max_days = st.number_input(
                    "Max days apart", min_value=1,
                    max_value=rapid_days, value=rapid_days, key="rv_maxd",
                )
            with _rfil3:
                _rv_sort = st.selectbox(
                    "Sort by", ["Days apart ↑","Days apart ↓","Patient name"],
                    key="rv_sort",
                )

            _rapid_filtered = [r for r in rapid if r["days_apart"] <= _rv_max_days]
            if _rv_srch:
                _rapid_filtered = [
                    r for r in _rapid_filtered
                    if _rv_srch.lower() in str(r).lower()
                ]
            if _rv_sort == "Days apart ↑":
                _rapid_filtered.sort(key=lambda x: x["days_apart"])
            elif _rv_sort == "Days apart ↓":
                _rapid_filtered.sort(key=lambda x: -x["days_apart"])
            else:
                _rapid_filtered.sort(key=lambda x: x["patient_name"])

            st.markdown(
                f'<div class="sec-head">⚠️ {len(_rapid_filtered):,} cases'
                f' — window ≤{rapid_days} days</div>',
                unsafe_allow_html=True,
            )

            # Card grid
            for _ri in range(0, min(len(_rapid_filtered), 30), 3):
                _batch = _rapid_filtered[_ri:_ri + 3]
                _bcols = st.columns(len(_batch))
                for _bcol, _r in zip(_bcols, _batch):
                    _crit = _r["days_apart"] <= 2
                    _day_col = DANGER if _crit else WARN
                    _bcol.markdown(f"""
<div style='background:#111720;border:1px solid #1e2a38;
     border-left:3px solid {_day_col};border-radius:10px;
     padding:12px 14px;margin-bottom:8px'>
  <div style='display:flex;justify-content:space-between;align-items:flex-start'>
    <div>
      <div style='font-size:13px;font-weight:600;color:#e2e8f0'>
        {_r["patient_name"][:24]}</div>
      <div style='font-size:11px;color:#64748b;font-family:monospace;margin-top:2px'>
        {_r["patient_id"]}</div>
    </div>
    <div style='font-size:24px;font-weight:800;color:{_day_col};
         font-family:Syne,sans-serif;line-height:1'>
      {_r["days_apart"]}<span style='font-size:11px;font-weight:400'> d</span>
    </div>
  </div>
  <div style='font-size:11px;color:#64748b;font-family:monospace;margin-top:6px'>
    📅 {_r["visit_1"]} → {_r["visit_2"]}</div>
  <div style='font-size:11px;color:#64748b;font-family:monospace;margin-top:2px'>
    👨‍⚕️ {str(_r["doctor"])[:28]}</div>
</div>""", unsafe_allow_html=True)

            if len(_rapid_filtered) > 30:
                st.caption(f"Showing first 30 cards. Full table below.")

            # Full table
            with st.expander("📋 Full rapid revisit table", expanded=False):
                _rv_df = pd.DataFrame(_rapid_filtered)
                st.dataframe(_rv_df, use_container_width=True, height=400)

            # Download
            _rv_dl = pd.DataFrame(_rapid_filtered).to_csv(index=False).encode()
            st.download_button(
                "⬇️ Download rapid revisit list",
                data=_rv_dl,
                file_name="pharmascan_rapid_revisits.csv",
                mime="text/csv",
                key="dl_rapid",
            )

# ══ NETWORK GRAPH ════════════════════════════════════════════════════════════
with tab_network:
  if not _dl_committed:
    _render_tab_locked("Network Graph")
  else:

    # All text / categorical columns available for network nodes
    cat_cols = [c for c in df.columns
                if df[c].dtype == object or str(df[c].dtype).startswith("string")]
    if not cat_cols:
        cat_cols = list(df.columns)

    def col_idx(candidates):
        for name in candidates:
            if name in cat_cols:
                return cat_cols.index(name)
        return 0

    st.markdown('<div class="sec-head">🔧 Network Configuration</div>', unsafe_allow_html=True)

    cfg1, cfg2, cfg3, cfg4, cfg5 = st.columns([2, 2, 1.8, 1.4, 1.4])

    with cfg1:
        col_a = st.selectbox(
            "◆ Node A (diamonds)",
            options=cat_cols,
            index=col_idx(["doctor_name", "doctor_type", "doctor_id"]),
            help="Each unique value in this column becomes a diamond-shaped node",
        )
    with cfg2:
        b_default = col_idx(["patient_name", "patient_id", "gender", "patient_type"])
        if cat_cols[b_default] == col_a and len(cat_cols) > 1:
            b_default = (b_default + 1) % len(cat_cols)
        col_b = st.selectbox(
            "● Node B (circles)",
            options=cat_cols,
            index=b_default,
            help="Each unique value in this column becomes a circle-shaped node",
        )
    with cfg3:
        physics_mode = st.selectbox(
            "Physics / layout",
            options=["Force Atlas 2", "Barnes-Hut", "Repulsion", "None (static)"],
            help="Force simulation used to position nodes",
        )
    with cfg4:
        max_nodes = st.slider("Max nodes", 20, 400, 150)
    with cfg5:
        min_edge  = st.slider("Min edge weight", 1, 10, 1,
                              help="Hide edges with fewer shared visits than this")

    net_height = st.slider("Canvas height (px)", 400, 900, 680, step=40)

    # Quick preset combos
    st.markdown("""
<div style='font-size:11px;color:#64748b;margin-bottom:14px;font-family:monospace'>
  💡 Try: <b style='color:#e2e8f0'>Practitioner Name ↔ Patient Name</b> &nbsp;·&nbsp;
  <b style='color:#e2e8f0'>Practitioner Type ↔ Patient Type</b> &nbsp;·&nbsp;
  <b style='color:#e2e8f0'>Practitioner Name ↔ Gender</b> &nbsp;·&nbsp;
  <b style='color:#e2e8f0'>Practitioner Name ↔ Patient Type</b>
</div>""", unsafe_allow_html=True)

    if col_a == col_b:
        st.warning("⚠️ Node A and Node B must be different columns.")
    else:
        card_a = df[col_a].nunique()
        card_b = df[col_b].nunique()
        st.markdown(
            f'<p style="font-size:12px;color:#64748b;font-family:monospace;margin-bottom:16px">'
            f'◆ <b style="color:#00e5a0">{col_a.replace("_"," ").title()}</b>: {card_a} unique values'
            f'&nbsp;|&nbsp;'
            f'● <b style="color:#0ea5e9">{col_b.replace("_"," ").title()}</b>: {card_b} unique values'
            f'&nbsp;|&nbsp; {len(df):,} rows total</p>',
            unsafe_allow_html=True,
        )

        with st.spinner("Building network…"):
            vis_nodes, vis_edges, stats = build_network_data(
                df, col_a, col_b, max_nodes, min_edge
            )

        if vis_nodes is None:
            st.warning("No edges found. Try lowering the minimum edge weight or choosing different columns.")
        else:
            # Stats row
            mc = st.columns(5)
            mc[0].metric(f"◆ {col_a.replace('_',' ').title()[:14]}", stats["nodes_a"])
            mc[1].metric(f"● {col_b.replace('_',' ').title()[:14]}", stats["nodes_b"])
            mc[2].metric("Edges",      stats["edges"])
            mc[3].metric("Avg Degree", stats["avg_degree"])
            mc[4].metric("Density",    stats["density"])

            # Controls hint
            st.markdown("""
<div style='font-size:11px;color:#64748b;margin:8px 0 6px;font-family:monospace'>
  🖱 <b style='color:#e2e8f0'>Drag</b> nodes to reposition &nbsp;·&nbsp;
  <b style='color:#e2e8f0'>Scroll</b> to zoom &nbsp;·&nbsp;
  <b style='color:#e2e8f0'>Click</b> a node to highlight its neighbours &nbsp;·&nbsp;
  <b style='color:#e2e8f0'>Double-click</b> to zoom in &nbsp;·&nbsp;
  <b style='color:#e2e8f0'>Search</b> any node by name
</div>""", unsafe_allow_html=True)

            # ── Render interactive vis.js network ──
            render_vis_network(vis_nodes, vis_edges, stats, physics_mode, height=net_height)

            # Top connected nodes tables below the canvas
            st.markdown("<br>", unsafe_allow_html=True)
            if stats["top_a"] or stats["top_b"]:
                ta, tb = st.columns(2)
                with ta:
                    st.markdown(
                        f'<div class="sec-head">◆ Top {col_a.replace("_"," ").title()} by connections</div>',
                        unsafe_allow_html=True,
                    )
                    st.dataframe(
                        pd.DataFrame(stats["top_a"], columns=["node", "connections"])
                          .style.background_gradient(subset=["connections"], cmap="Greens"),
                        use_container_width=True, height=320,
                    )
                with tb:
                    st.markdown(
                        f'<div class="sec-head">● Top {col_b.replace("_"," ").title()} by connections</div>',
                        unsafe_allow_html=True,
                    )
                    st.dataframe(
                        pd.DataFrame(stats["top_b"], columns=["node", "connections"])
                          .style.background_gradient(subset=["connections"], cmap="Blues"),
                        use_container_width=True, height=320,
                    )

# ══ NORMALISE NAMES TAB ══════════════════════════════════════════════════════
with tab_norm:
  if not _dl_committed:
    _render_tab_locked("Normalise Names")
  else:
    st.markdown('<div class="sec-head">✏️ Doctor / Practitioner Name Normalisation</div>',
                unsafe_allow_html=True)

    # Pick which column to normalise
    str_cols = [c for c in df.columns
                if df[c].dtype == object or str(df[c].dtype).startswith("string")]

    norm_col = st.selectbox(
        "Column to normalise",
        options=str_cols,
        index=str_cols.index("doctor_name") if "doctor_name" in str_cols else 0,
        help="Fuzzy-match similar names within this column and merge variants into one canonical form",
    )

    with st.spinner("Detecting name clusters…"):
        col_counts = df[norm_col].value_counts().to_dict()
        col_names  = sorted(df[norm_col].dropna().unique().tolist(), key=str)
        clusters   = detect_name_clusters(col_names, col_counts)

    n_clusters   = len(clusters)
    n_suspicious = sum(1 for c in clusters if c["suspicious"])
    n_variants   = sum(len(c["variants"]) for c in clusters)

    # ── Summary metrics ──
    mc = st.columns(4)
    mc[0].metric("Unique raw names",   len(col_names))
    mc[1].metric("Variant clusters",   n_clusters)
    mc[2].metric("Total variants",     n_variants,
                 delta=f"→ {len(col_names) - n_variants} after merge")
    mc[3].metric("⚠️ Needs review",    n_suspicious)

    st.markdown("""
<div style='font-size:11px;color:#64748b;margin:8px 0 16px;font-family:monospace'>
  Below are proposed name merges. <b style='color:#e2e8f0'>✅ Check</b> clusters to approve them,
  then click <b style='color:#00e5a0'>Apply Selected</b>. Suspicious clusters (no shared tokens)
  are shown in amber — review carefully before approving.
</div>""", unsafe_allow_html=True)

    if not clusters:
        st.success("✅ No similar names detected — the column looks clean.")
    else:
        # ── Quick select helpers ──
        qc1, qc2, qc3 = st.columns([1, 1, 4])
        with qc1:
            if st.button("✅ Select all"):
                for c in clusters:
                    st.session_state[f"norm_{c['canonical']}"] = True
        with qc2:
            if st.button("❌ Deselect all"):
                for c in clusters:
                    st.session_state[f"norm_{c['canonical']}"] = False

        # ── Cluster review cards ──
        approved = []
        for cluster in clusters:
            key      = f"norm_{cluster['canonical']}"
            default  = not cluster["suspicious"]    # auto-select non-suspicious
            checked  = st.session_state.get(key, default)

            conf_pct = int(cluster["confidence"] * 100)
            if cluster["suspicious"]:
                border = "#f59e0b"
                conf_badge = f'<span style="background:rgba(245,158,11,.15);color:#f59e0b;border-radius:5px;padding:2px 8px;font-size:10px">⚠️ {conf_pct}% — review</span>'
            elif conf_pct >= 90:
                border = "#00e5a0"
                conf_badge = f'<span style="background:rgba(0,229,160,.1);color:#00e5a0;border-radius:5px;padding:2px 8px;font-size:10px">✓ {conf_pct}%</span>'
            else:
                border = "#0ea5e9"
                conf_badge = f'<span style="background:rgba(14,165,233,.1);color:#0ea5e9;border-radius:5px;padding:2px 8px;font-size:10px">{conf_pct}%</span>'

            # Variant pills
            freq_total = sum(col_counts.get(v, 0) for v in cluster["variants"])
            canon_freq = col_counts.get(cluster["canonical"], 0)
            variant_pills = " ".join(
                f'<span style="background:#1e2a38;border-radius:5px;padding:2px 8px;font-size:11px;'
                f'font-family:monospace;color:#94a3b8;margin:2px">'
                f'{v} <span style="color:#64748b">×{col_counts.get(v,0)}</span></span>'
                for v in cluster["variants"]
            )

            st.markdown(f"""
<div style="background:#111720;border:1px solid {border};border-left:3px solid {border};
     border-radius:10px;padding:14px 16px;margin-bottom:10px">
  <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:12px;flex-wrap:wrap">
    <div style="flex:1;min-width:200px">
      <div style="font-size:13px;font-weight:700;color:#e2e8f0;margin-bottom:4px">
        → <span style="color:#00e5a0">{cluster['canonical']}</span>
        <span style="color:#64748b;font-size:11px;margin-left:6px">×{canon_freq} occurrences</span>
      </div>
      <div style="font-size:11px;color:#64748b;margin-bottom:8px;font-family:monospace">
        {len(cluster['variants'])} variant(s) · {freq_total} rows affected
      </div>
      <div style="display:flex;flex-wrap:wrap;gap:4px">{variant_pills}</div>
    </div>
    <div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px;flex-shrink:0">
      {conf_badge}
    </div>
  </div>
</div>""", unsafe_allow_html=True)

            checked = st.checkbox(
                f"Approve merge → {cluster['canonical']}",
                value=checked,
                key=key,
            )
            if checked:
                approved.append(cluster)

        # ── Apply button ──
        st.markdown("<br>", unsafe_allow_html=True)
        col_apply, col_preview = st.columns([1, 2])

        with col_apply:
            st.markdown(
                f'<p style="font-size:12px;color:#64748b;font-family:monospace">'
                f'{len(approved)} cluster(s) selected · '
                f'{sum(len(c["variants"]) for c in approved)} variant(s) will be merged</p>',
                unsafe_allow_html=True,
            )
            do_apply = st.button("⚡ Apply Selected Normalisations",
                                 type="primary", disabled=len(approved) == 0)

        if do_apply:
            st.session_state["normalised_df"]  = apply_name_normalisation(df, norm_col, approved)
            st.session_state["normalised_col"] = norm_col
            st.session_state["normalised_map"] = {
                v: c["canonical"]
                for c in approved for v in c["variants"]
            }

        # ── Show result ──
        if "normalised_df" in st.session_state and st.session_state.get("normalised_col") == norm_col:
            ndf = st.session_state["normalised_df"]
            nmap = st.session_state["normalised_map"]

            st.success(f"✅ Applied! {len(nmap)} variant names merged in column **{norm_col}**.")

            before_unique = df[norm_col].nunique()
            after_unique  = ndf[norm_col].nunique()

            rc1, rc2, rc3 = st.columns(3)
            rc1.metric("Before: unique names", before_unique)
            rc2.metric("After: unique names",  after_unique,
                       delta=f"−{before_unique - after_unique}")
            rc3.metric("Rows updated", len(nmap))

            # Rename map table
            st.markdown('<div class="sec-head">Applied Rename Map</div>', unsafe_allow_html=True)
            map_df = pd.DataFrame(
                [(k, v) for k, v in sorted(nmap.items())],
                columns=["Original variant", "→ Canonical name"]
            )
            st.dataframe(map_df, use_container_width=True, height=300)

            # Updated data preview
            st.markdown('<div class="sec-head">Updated Data Preview</div>', unsafe_allow_html=True)
            srch_n = st.text_input("🔍 Filter", key="norm_srch", placeholder="Search any column…")
            show_df = ndf.copy()
            if not show_raw:
                show_df.columns = [c.replace("_", " ").title() for c in show_df.columns]
            if srch_n:
                mask = show_df.apply(lambda c: c.astype(str).str.contains(srch_n, case=False, na=False)).any(axis=1)
                show_df = show_df[mask]
            st.dataframe(show_df, use_container_width=True, height=480)

            # Download
            csv_bytes = ndf.to_csv(index=False).encode()
            st.download_button(
                "⬇️ Download normalised CSV",
                data=csv_bytes,
                file_name=f"pharmascan_normalised_{norm_col}.csv",
                mime="text/csv",
            )

# ══ CROSS-FACILITY FRAUD DETECTION TAB ═══════════════════════════════════════
with tab_xfac:
  if not _dl_committed:
    _render_tab_locked("Cross-Facility Match")
  else:
    import math

    # ── CSS for fraud tab cards ───────────────────────────────────────────────
    st.markdown("""
<style>
.fraud-card{background:#111720;border:1px solid #1e2a38;border-radius:12px;
            padding:18px 22px;margin-bottom:14px}
.fraud-card-red{background:#1a0505;border-color:#7f1d1d}
.fraud-card-amber{background:#1a1000;border-color:#78350f}
.fraud-card-blue{background:#030d1a;border-color:#1e3a5f}
.fraud-card-green{background:#031a0a;border-color:#14532d}
.badge{display:inline-block;padding:2px 10px;border-radius:20px;
       font-size:11px;font-weight:700;font-family:monospace}
.badge-red{background:#7f1d1d;color:#fca5a5}
.badge-amber{background:#78350f;color:#fde68a}
.badge-green{background:#14532d;color:#86efac}
.badge-blue{background:#1e3a5f;color:#93c5fd}
.badge-purple{background:#3b1f7a;color:#c4b5fd}
.risk-bar-wrap{background:#1e2a38;border-radius:6px;height:10px;width:100%;margin:4px 0}
.risk-bar{height:10px;border-radius:6px}
</style>""", unsafe_allow_html=True)

    st.markdown("""
<div style='font-family:Syne,sans-serif;font-size:26px;font-weight:800;
     color:#e2e8f0;margin-bottom:4px'>
  🔍 Pharmacy Fraud Detection
</div>
<div style='color:#64748b;font-size:13px;margin-bottom:20px;font-family:monospace'>
  Identifies pharmacy claims where medicine was dispensed to patients
  with <b style='color:#ef4444'>no verifiable hospital or clinic visit record</b>
  — a primary indicator of fraudulent RSSB reimbursement claims.
</div>""", unsafe_allow_html=True)

    # ── How it works banner ────────────────────────────────────────────────────
    st.markdown("""
<div class='fraud-card' style='border-color:#0ea5e9;background:#020e1a;margin-bottom:22px'>
<div style='font-size:13px;font-weight:700;color:#38bdf8;margin-bottom:10px'>
  📋 Patient Journey & Fraud Logic
</div>
<div style='display:flex;gap:8px;align-items:center;flex-wrap:wrap;font-size:12px;
     font-family:monospace;color:#94a3b8'>
  <span style='background:#1e3a5f;color:#93c5fd;padding:4px 10px;border-radius:6px'>
    🏥 Clinic Visit
  </span>
  <span style='color:#475569'>→</span>
  <span style='background:#1e3a5f;color:#93c5fd;padding:4px 10px;border-radius:6px'>
    👨‍⚕️ Doctor consults
  </span>
  <span style='color:#475569'>→</span>
  <span style='background:#1e3a5f;color:#93c5fd;padding:4px 10px;border-radius:6px'>
    📝 Prescription + Voucher
  </span>
  <span style='color:#475569'>→</span>
  <span style='background:#1e3a5f;color:#93c5fd;padding:4px 10px;border-radius:6px'>
    💰 Patient pays 15%
  </span>
  <span style='color:#475569'>→</span>
  <span style='background:#14532d;color:#86efac;padding:4px 10px;border-radius:6px'>
    💊 Pharmacy dispenses
  </span>
  <span style='color:#475569'>→</span>
  <span style='background:#14532d;color:#86efac;padding:4px 10px;border-radius:6px'>
    📤 Pharmacy claims 85% from RSSB
  </span>
</div>
<div style='margin-top:10px;font-size:11px;color:#64748b;font-family:monospace'>
  <b style='color:#ef4444'>⚠️ Fraud signal:</b> Pharmacy claims 85% reimbursement for a patient
  who has <i>no hospital/clinic visit record</i> — meaning there was no legitimate
  consultation, no doctor, and likely no real prescription.
  The more facility files uploaded, the more accurate this detection becomes.
</div>
</div>""", unsafe_allow_html=True)

    # ── STEP 1: Upload facility files ─────────────────────────────────────────
    st.markdown('<div class="sec-head">📂 Step 1 — Upload Hospital & Clinic Visit Files</div>',
                unsafe_allow_html=True)
    st.markdown("""
<div style='font-size:11px;color:#64748b;font-family:monospace;margin-bottom:10px'>
  Upload one or more hospital/clinic Excel files for the <b>same period</b> as the
  pharmacy report. The app auto-detects the verified sheet and all relevant columns.
  <b style='color:#e2e8f0'>The more facilities you upload, the smaller the unverified group becomes.</b>
</div>""", unsafe_allow_html=True)

    xf_uploads = st.file_uploader(
        "Upload facility files (Excel / CSV) — multiple allowed",
        type=["xlsx","xls","csv"],
        accept_multiple_files=True,
        key="xf_uploads",
    )

    # ── Parser ────────────────────────────────────────────────────────────────
    def _parse_facility(raw_bytes, filename):
        fname = filename.lower()
        try:
            if fname.endswith(".csv"):
                raw = pd.read_csv(io.BytesIO(raw_bytes), encoding="utf-8", on_bad_lines="skip")
                chosen_name, chosen_df = "main", raw
                header_row = 0
            else:
                xl = pd.ExcelFile(io.BytesIO(raw_bytes))
                chosen_name = None; chosen_df = None
                for priority in ["after","verified","clean","before","data","invoice report","invoice"]:
                    for sn in xl.sheet_names:
                        if priority in sn.lower():
                            chosen_name = sn
                            chosen_df   = xl.parse(sn, header=None)
                            break
                    if chosen_name: break
                if chosen_df is None:
                    sizes = {}
                    for sn in xl.sheet_names:
                        try: sizes[sn] = xl.parse(sn,header=None).shape[0]
                        except: pass
                    chosen_name = max(sizes, key=sizes.get)
                    chosen_df   = xl.parse(chosen_name, header=None)

                # Find header row
                header_row = 0
                for i, row in chosen_df.head(20).iterrows():
                    joined = " ".join(str(v).lower() for v in row if pd.notna(v))
                    if any(k in joined for k in ["affil","rama","beneficiary","patient name","voucher"]):
                        header_row = i; break

            if header_row > 0:
                chosen_df = pd.read_excel(io.BytesIO(raw_bytes),
                                          sheet_name=chosen_name, header=header_row)
            else:
                if not isinstance(chosen_df.columns[0], str) or chosen_df.columns[0] == 0:
                    chosen_df.columns = chosen_df.iloc[0]
                    chosen_df = chosen_df.iloc[1:].reset_index(drop=True)

            chosen_df.columns = [str(c).strip() for c in chosen_df.columns]

            def _find(patterns):
                for pat in patterns:
                    for c in chosen_df.columns:
                        if re.search(pat, str(c).lower()): return c
                return None

            rama_c  = _find([r"affil", r"rama", r"member.*no"])
            name_c  = _find([r"benefi.*name", r"patient.*name", r"client.*name"])
            date_c  = _find([r"^date$", r"visit.*date", r"dispensing.*date", r"service.*date"])
            vou_c   = _find([r"voucher.*id", r"voucher.*ident", r"paper.*code", r"invoice.*no"])
            total_c = _find([r"total.*amount", r"total.*cost", r"^total$"])
            doc_c   = _find([r"practitioner", r"prescrib", r"doctor", r"physician", r"medecin"])

            if not rama_c:
                return None, f"No RAMA/Affiliation column found in {chosen_name!r}"

            # Drop footer/total rows
            no_col = chosen_df.columns[0]
            chosen_df = chosen_df[pd.to_numeric(chosen_df[no_col], errors="coerce").notna()].copy()

            out = pd.DataFrame()
            out["_rama"]       = chosen_df[rama_c].astype(str).str.strip().str.upper()
            out["_name"]       = chosen_df[name_c].fillna("").astype(str).str.strip() if name_c else ""
            out["_date"]       = pd.to_datetime(chosen_df[date_c], errors="coerce") if date_c else pd.NaT
            out["voucher_id"]  = chosen_df[vou_c].astype(str).str.strip() if vou_c else ""
            out["total"]       = pd.to_numeric(chosen_df[total_c], errors="coerce").fillna(0) if total_c else 0
            out["doctor"]      = chosen_df[doc_c].fillna("").astype(str).str.strip() if doc_c else ""
            out["_source"]     = filename
            out["_sheet"]      = chosen_name
            out = out[out["_rama"].str.len() > 2].reset_index(drop=True)
            return out, None
        except Exception as e:
            import traceback
            return None, f"{e}\n{traceback.format_exc()}"

    # ── Parse uploaded files ──────────────────────────────────────────────────
    if xf_uploads:
        new_frames = []
        for uf in xf_uploads:
            raw_bytes = uf.read()
            parsed, err = _parse_facility(raw_bytes, uf.name)
            if parsed is not None and len(parsed) > 0:
                new_frames.append(parsed)
                st.success(f"✅ **{uf.name}** — {len(parsed):,} visit records loaded "
                           f"(sheet: *{parsed['_sheet'].iloc[0]}*)")
            else:
                st.error(f"❌ **{uf.name}** — {err}")
        if new_frames:
            st.session_state["fd_facility"] = pd.concat(new_frames, ignore_index=True)

    if "fd_facility" not in st.session_state:
        st.info("👆 Upload at least one hospital or clinic file to run fraud detection.")
        st.stop()

    fac_df = st.session_state["fd_facility"]
    fac_ramas = set(fac_df["_rama"].tolist())

    # File summary
    source_summary = fac_df.groupby("_source").size().reset_index(name="records")
    cols_fsum = st.columns(min(len(source_summary), 4))
    for i, row in source_summary.iterrows():
        cols_fsum[i % len(cols_fsum)].metric(
            row["_source"][:35], f"{row['records']:,} records"
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── STEP 2: Matching config ────────────────────────────────────────────────
    st.markdown('<div class="sec-head">⚙️ Step 2 — Detection Settings</div>',
                unsafe_allow_html=True)

    cfg1, cfg2, cfg3 = st.columns(3)
    with cfg1:
        date_window = st.slider("Date window (±days)", 0, 30, 7,
            help="How many days before/after a pharmacy dispensing date to look for a matching hospital visit")
    with cfg2:
        name_thresh = st.slider("Name similarity (0–1)", 0.0, 1.0, 0.4, 0.05,
            help="Token overlap score between pharmacy and hospital patient names. Lower = more permissive.")
    with cfg3:
        require_name = st.checkbox("Require name match", value=True,
            help="If OFF, matches on RAMA number alone (catches name typos)")

    # ── Build pharmacy working set ────────────────────────────────────────────
    def _ph_col(*keys):
        for k in keys:
            if k in df.columns: return k
        return None

    vid_c = _ph_col("voucher_id","paper_code","Paper Code")
    pnm_c = _ph_col("patient_name","Patient Name")
    rma_c = _ph_col("patient_id","rama_number","RAMA Number")
    dt_c  = _ph_col("visit_date","dispensing_date","Dispensing Date")
    ins_c = _ph_col("insurance_copay","Insurance Co-payment")
    tot_c = _ph_col("amount","total_cost","Total Cost")
    doc_c = _ph_col("doctor_name","practitioner_name","Practitioner Name")
    dpt_c = _ph_col("practitioner_type","Practitioner Type")

    ph_work = df.copy()
    ph_work["_rama"]  = ph_work[rma_c].astype(str).str.strip().str.upper() if rma_c else ""
    ph_work["_name"]  = ph_work[pnm_c].fillna("").astype(str).str.strip()  if pnm_c else ""
    ph_work["_date"]  = pd.to_datetime(ph_work[dt_c], errors="coerce")     if dt_c  else pd.NaT
    ph_work["_vou"]   = ph_work[vid_c].astype(str).str.strip()             if vid_c else ""
    ph_work["_ins"]   = pd.to_numeric(ph_work[ins_c], errors="coerce").fillna(0) if ins_c else 0
    ph_work["_tot"]   = pd.to_numeric(ph_work[tot_c], errors="coerce").fillna(0) if tot_c else 0
    ph_work["_doc"]   = ph_work[doc_c].fillna("").astype(str)              if doc_c else ""
    ph_work["_dpt"]   = ph_work[dpt_c].fillna("").astype(str)              if dpt_c else ""

    def _tok(a, b):
        ta = set(str(a).upper().split()); tb = set(str(b).upper().split())
        return len(ta & tb) / len(ta | tb) if ta and tb else 0.0

    # ── Core matching ─────────────────────────────────────────────────────────
    # For each pharmacy row, find best facility match by RAMA + name + date
    results = []
    for _, pr in ph_work.iterrows():
        rama     = pr["_rama"]
        ph_date  = pr["_date"]
        ph_name  = pr["_name"]

        fac_rows = fac_df[fac_df["_rama"] == rama]

        if fac_rows.empty:
            # NO facility record for this RAMA at all
            results.append({
                "status": "NO_RECORD",
                "ph_voucher": pr["_vou"],
                "ph_patient": ph_name,
                "ph_rama":    rama,
                "ph_date":    ph_date,
                "ph_ins":     pr["_ins"],
                "ph_total":   pr["_tot"],
                "ph_doctor":  pr["_doc"],
                "ph_dept":    pr["_dpt"],
                "fac_voucher": None, "fac_name": None,
                "fac_date":    None, "fac_source": None,
                "days_apart":  None, "name_score": None,
            })
            continue

        # RAMA exists — check name + date
        best = None; best_delta = 9999; best_score = 0
        for _, fr in fac_rows.iterrows():
            fac_date = fr["_date"]
            nscore   = _tok(ph_name, fr["_name"])
            delta    = abs((ph_date - fac_date).days) if pd.notna(ph_date) and pd.notna(fac_date) else 9999
            name_ok  = (nscore >= name_thresh) if require_name else True
            if name_ok and delta <= date_window:
                if delta < best_delta or (delta == best_delta and nscore > best_score):
                    best_delta = delta; best_score = nscore; best = fr

        if best is not None:
            # MATCHED — legitimate dispensing with traced visit
            results.append({
                "status":      "MATCHED",
                "ph_voucher":  pr["_vou"],   "ph_patient": ph_name,
                "ph_rama":     rama,          "ph_date":    ph_date,
                "ph_ins":      pr["_ins"],    "ph_total":   pr["_tot"],
                "ph_doctor":   pr["_doc"],    "ph_dept":    pr["_dpt"],
                "fac_voucher": best["voucher_id"], "fac_name": best["_name"],
                "fac_date":    best["_date"],      "fac_source": best["_source"],
                "days_apart":  best_delta,    "name_score": round(best_score, 2),
            })
        else:
            # RAMA EXISTS but date/name mismatch — partial flag
            fac_dates = fac_rows["_date"].dropna()
            nearest_d = None
            if not fac_dates.empty and pd.notna(ph_date):
                deltas = (fac_dates - ph_date).abs()
                nearest_d = int(deltas.min().days)
            best_fr = fac_rows.iloc[0]
            results.append({
                "status":      "UNLINKED",
                "ph_voucher":  pr["_vou"],   "ph_patient": ph_name,
                "ph_rama":     rama,          "ph_date":    ph_date,
                "ph_ins":      pr["_ins"],    "ph_total":   pr["_tot"],
                "ph_doctor":   pr["_doc"],    "ph_dept":    pr["_dpt"],
                "fac_voucher": best_fr["voucher_id"], "fac_name": best_fr["_name"],
                "fac_date":    best_fr["_date"],      "fac_source": best_fr["_source"],
                "days_apart":  nearest_d,     "name_score": round(_tok(ph_name, best_fr["_name"]),2),
            })

    res_df = pd.DataFrame(results)

    no_rec    = res_df[res_df["status"]=="NO_RECORD"]
    unlinked  = res_df[res_df["status"]=="UNLINKED"]
    matched   = res_df[res_df["status"]=="MATCHED"]

    total_ins       = res_df["ph_ins"].sum()
    no_rec_ins      = no_rec["ph_ins"].sum()
    unlinked_ins    = unlinked["ph_ins"].sum()
    matched_ins     = matched["ph_ins"].sum()
    at_risk_ins     = no_rec_ins + unlinked_ins
    fac_count       = fac_df["_source"].nunique()
    coverage_pct    = 100 * matched_ins / total_ins if total_ins else 0

    # ── STEP 3: Dashboard ─────────────────────────────────────────────────────
    st.markdown('<div class="sec-head">📊 Step 3 — Fraud Detection Dashboard</div>',
                unsafe_allow_html=True)

    # Top KPI strip
    k1,k2,k3,k4,k5 = st.columns(5)
    k1.metric("Total pharmacy vouchers", f"{len(ph_work):,}")
    k2.metric("✅ Verified (visit found)",
              f"{len(matched):,}",
              f"{100*len(matched)/len(ph_work):.1f}%")
    k3.metric("🔴 No facility record",
              f"{len(no_rec):,}",
              f"-{100*len(no_rec)/len(ph_work):.1f}%",
              delta_color="inverse")
    k4.metric("🟡 RAMA found, visit unlinked",
              f"{len(unlinked):,}",
              f"-{100*len(unlinked)/len(ph_work):.1f}%",
              delta_color="inverse")
    k5.metric("Facilities loaded", f"{fac_count}")

    st.markdown("<br>", unsafe_allow_html=True)

    # Insurance risk strip
    r1,r2,r3 = st.columns(3)
    r1.metric("Total RSSB claims (85%)",      f"RWF {total_ins:,.0f}")
    r2.metric("🔴 At-risk amount (no record)", f"RWF {no_rec_ins:,.0f}",
              f"{100*no_rec_ins/total_ins:.1f}% of total",
              delta_color="inverse")
    r3.metric("🟡 Partially unlinked",         f"RWF {unlinked_ins:,.0f}",
              f"{100*unlinked_ins/total_ins:.1f}% of total",
              delta_color="inverse")

    # Risk bar
    bar_matched  = 100 * matched_ins  / total_ins if total_ins else 0
    bar_unlinked = 100 * unlinked_ins / total_ins if total_ins else 0
    bar_norec    = 100 * no_rec_ins   / total_ins if total_ins else 0
    st.markdown(f"""
<div style='margin:18px 0 8px'>
  <div style='font-size:11px;color:#64748b;font-family:monospace;margin-bottom:5px'>
    Insurance amount breakdown by verification status
  </div>
  <div style='display:flex;height:16px;border-radius:8px;overflow:hidden;width:100%'>
    <div style='width:{bar_matched:.1f}%;background:#22c55e' title='Verified: RWF {matched_ins:,.0f}'></div>
    <div style='width:{bar_unlinked:.1f}%;background:#f59e0b' title='Unlinked: RWF {unlinked_ins:,.0f}'></div>
    <div style='width:{bar_norec:.1f}%;background:#ef4444' title='No record: RWF {no_rec_ins:,.0f}'></div>
  </div>
  <div style='display:flex;gap:18px;margin-top:5px;font-size:11px;font-family:monospace'>
    <span style='color:#22c55e'>■ Verified {bar_matched:.1f}%</span>
    <span style='color:#f59e0b'>■ Unlinked {bar_unlinked:.1f}%</span>
    <span style='color:#ef4444'>■ No record {bar_norec:.1f}%</span>
  </div>
</div>""", unsafe_allow_html=True)

    # Coverage note
    st.markdown(f"""
<div style='background:#030d1a;border:1px solid #1e3a5f;border-radius:8px;
     padding:10px 16px;font-size:11px;font-family:monospace;color:#64748b;margin-bottom:20px'>
  <b style='color:#38bdf8'>Coverage note:</b>
  {fac_count} facility file(s) loaded covering {len(fac_ramas):,} unique RAMA numbers.
  Pharmacy serves patients from many facilities — patients in the
  <span style='color:#ef4444'>"No record"</span> group may have visited clinics
  <b>not yet uploaded</b>. Upload more facility files to reduce false positives.
</div>""", unsafe_allow_html=True)

    st.markdown("<hr style='border-color:#1e2a38;margin:4px 0 24px'>", unsafe_allow_html=True)

    # ── TABLE 1: No Hospital Record ───────────────────────────────────────────
    st.markdown(f"""
<div class='fraud-card fraud-card-red'>
  <div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px'>
    <div>
      <span style='font-size:16px;font-weight:800;color:#f87171;font-family:Syne,sans-serif'>
        🔴 Table 1 — No Hospital / Clinic Visit Record
      </span>
      <span class='badge badge-red' style='margin-left:10px'>{len(no_rec):,} vouchers</span>
      <span class='badge badge-red' style='margin-left:6px'>RWF {no_rec_ins:,.0f}</span>
    </div>
    <span style='font-size:11px;color:#991b1b;font-family:monospace'>
      Patient's RAMA number not found in ANY uploaded facility file
    </span>
  </div>
</div>""", unsafe_allow_html=True)

    if not no_rec.empty:
        # Sub-controls
        t1c1, t1c2, t1c3 = st.columns([2,1.5,1.5])
        with t1c1:
            t1_srch = st.text_input("🔍 Search", placeholder="Name, RAMA, voucher, doctor…",
                                     key="t1_srch")
        with t1c2:
            t1_doc = st.selectbox("Filter by Prescriber",
                ["All"] + sorted(no_rec["ph_doctor"].unique().tolist()),
                key="t1_doc")
        with t1c3:
            t1_min = st.number_input("Min insurance (RWF)", 0, value=0, step=5000, key="t1_min")

        t1_disp = no_rec.copy()
        if t1_srch:
            mask = t1_disp.apply(
                lambda c: c.astype(str).str.contains(t1_srch, case=False, na=False)
            ).any(axis=1)
            t1_disp = t1_disp[mask]
        if t1_doc != "All":
            t1_disp = t1_disp[t1_disp["ph_doctor"] == t1_doc]
        if t1_min > 0:
            t1_disp = t1_disp[t1_disp["ph_ins"] >= t1_min]

        # Build clean display table
        t1_show = t1_disp[[
            "ph_voucher","ph_patient","ph_rama","ph_date",
            "ph_ins","ph_total","ph_doctor","ph_dept"
        ]].copy()
        t1_show.columns = [
            "Pharmacy Voucher","Patient Name","RAMA Number","Dispensing Date",
            "Insurance Claim (RWF)","Total Cost (RWF)","Prescriber","Specialty"
        ]
        t1_show["Dispensing Date"] = pd.to_datetime(
            t1_show["Dispensing Date"], errors="coerce"
        ).dt.strftime("%d/%m/%Y").fillna("—")
        t1_show = t1_show.sort_values("Insurance Claim (RWF)", ascending=False)
        t1_show.index = range(1, len(t1_show)+1)

        st.markdown(
            f"<div style='font-size:11px;color:{MUTED};font-family:monospace;margin-bottom:6px'>"
            f"Showing <b style='color:#f87171'>{len(t1_show):,}</b> vouchers · "
            f"Insurance at risk: <b style='color:#ef4444'>RWF {t1_disp['ph_ins'].sum():,.0f}</b>"
            f"</div>", unsafe_allow_html=True
        )
        st.dataframe(t1_show, use_container_width=True, height=340)

        # Prescriber risk breakdown
        with st.expander("📊 Prescriber risk breakdown (Table 1)", expanded=False):
            doc_risk = (no_rec.groupby("ph_doctor")["ph_ins"]
                        .agg(Vouchers="count", Total_Claimed="sum")
                        .sort_values("Total_Claimed", ascending=False)
                        .reset_index())
            doc_risk.columns = ["Prescriber","Vouchers","Total Claimed (RWF)"]
            st.dataframe(doc_risk, use_container_width=True, height=280)

        # Download
        t1_buf = io.BytesIO()
        _t1_xl = no_rec.copy()
        _t1_xl["ph_date"] = pd.to_datetime(_t1_xl["ph_date"], errors="coerce").dt.strftime("%d/%m/%Y")
        with pd.ExcelWriter(t1_buf, engine="openpyxl") as xw:
            from openpyxl.styles import PatternFill as _PF, Font as _F, Alignment as _Al
            _t1_xl.to_excel(xw, index=False, sheet_name="No Facility Record")
            ws = xw.sheets["No Facility Record"]
            hf = _PF("solid", fgColor="7F1D1D")
            for cell in ws[1]:
                cell.fill = hf
                cell.font = _F(bold=True, color="FFFFFF", name="Arial", size=10)
                cell.alignment = _Al(horizontal="center", wrap_text=True)
            for i, r in enumerate(ws.iter_rows(min_row=2), 2):
                bg = "FFE4E4" if i%2==0 else "FFFFFF"
                for c in r: c.fill = _PF("solid", fgColor=bg)
        t1_buf.seek(0)
        st.download_button("⬇️ Download Table 1", t1_buf.getvalue(),
            "table1_no_facility_record.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_t1")

    st.markdown("<hr style='border-color:#1e2a38;margin:28px 0'>", unsafe_allow_html=True)

    # ── TABLE 2: UNLINKED (RAMA found, visit not linked) ──────────────────────
    st.markdown(f"""
<div class='fraud-card fraud-card-amber'>
  <div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px'>
    <div>
      <span style='font-size:16px;font-weight:800;color:#fbbf24;font-family:Syne,sans-serif'>
        🟡 Table 2 — RAMA Found, Visit Not Linked
      </span>
      <span class='badge badge-amber' style='margin-left:10px'>{len(unlinked):,} vouchers</span>
      <span class='badge badge-amber' style='margin-left:6px'>RWF {unlinked_ins:,.0f}</span>
    </div>
    <span style='font-size:11px;color:#92400e;font-family:monospace'>
      Patient exists in a facility file but dispensing date is outside the ±{date_window}-day window
    </span>
  </div>
</div>""", unsafe_allow_html=True)

    if not unlinked.empty:
        t2c1, t2c2 = st.columns([2,1])
        with t2c1:
            t2_srch = st.text_input("🔍 Search", placeholder="Name, RAMA…", key="t2_srch")
        with t2c2:
            t2_max_gap = st.number_input("Max days apart to show", 1, 365, 60, key="t2_gap")

        t2_disp = unlinked.copy()
        if t2_srch:
            mask = t2_disp.apply(
                lambda c: c.astype(str).str.contains(t2_srch, case=False, na=False)
            ).any(axis=1)
            t2_disp = t2_disp[mask]
        if t2_max_gap:
            t2_disp = t2_disp[
                t2_disp["days_apart"].isna() | (t2_disp["days_apart"] <= t2_max_gap)
            ]

        t2_show = t2_disp[[
            "ph_voucher","ph_patient","ph_rama","ph_date","ph_ins",
            "fac_name","fac_date","fac_source","days_apart","name_score"
        ]].copy()
        t2_show.columns = [
            "Pharmacy Voucher","Pharmacy Patient","RAMA","Pharmacy Date","Insurance (RWF)",
            "Facility Patient","Facility Visit Date","Facility","Days Apart","Name Score"
        ]
        for dcol in ["Pharmacy Date","Facility Visit Date"]:
            t2_show[dcol] = pd.to_datetime(t2_show[dcol], errors="coerce").dt.strftime("%d/%m/%Y").fillna("—")
        t2_show = t2_show.sort_values("Days Apart", na_position="last")
        t2_show.index = range(1, len(t2_show)+1)

        st.markdown(
            f"<div style='font-size:11px;color:{MUTED};font-family:monospace;margin-bottom:6px'>"
            f"Showing <b style='color:#fbbf24'>{len(t2_show):,}</b> vouchers · "
            f"Insurance: <b style='color:#f59e0b'>RWF {t2_disp['ph_ins'].sum():,.0f}</b>"
            f"</div>", unsafe_allow_html=True
        )
        st.dataframe(t2_show, use_container_width=True, height=300)

        t2_buf = io.BytesIO()
        _t2_xl = t2_disp.copy()
        _t2_xl["ph_date"]  = pd.to_datetime(_t2_xl["ph_date"],  errors="coerce").dt.strftime("%d/%m/%Y")
        _t2_xl["fac_date"] = pd.to_datetime(_t2_xl["fac_date"], errors="coerce").dt.strftime("%d/%m/%Y")
        with pd.ExcelWriter(t2_buf, engine="openpyxl") as xw:
            from openpyxl.styles import PatternFill as _PF, Font as _F, Alignment as _Al
            _t2_xl.to_excel(xw, index=False, sheet_name="Unlinked Visits")
            ws = xw.sheets["Unlinked Visits"]
            for cell in ws[1]:
                cell.fill = _PF("solid", fgColor="78350F")
                cell.font = _F(bold=True, color="FFFFFF", name="Arial", size=10)
                cell.alignment = _Al(horizontal="center", wrap_text=True)
            for i, r in enumerate(ws.iter_rows(min_row=2), 2):
                bg = "FEF3C7" if i%2==0 else "FFFFFF"
                for c in r: c.fill = _PF("solid", fgColor=bg)
        t2_buf.seek(0)
        st.download_button("⬇️ Download Table 2", t2_buf.getvalue(),
            "table2_unlinked_visits.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_t2")

    st.markdown("<hr style='border-color:#1e2a38;margin:28px 0'>", unsafe_allow_html=True)

    # ── TABLE 3: MATCHED (verified, informational) ────────────────────────────
    st.markdown(f"""
<div class='fraud-card fraud-card-green'>
  <div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px'>
    <div>
      <span style='font-size:16px;font-weight:800;color:#4ade80;font-family:Syne,sans-serif'>
        ✅ Table 3 — Verified: Hospital Visit + Pharmacy Dispensing Linked
      </span>
      <span class='badge badge-green' style='margin-left:10px'>{len(matched):,} vouchers</span>
      <span class='badge badge-green' style='margin-left:6px'>RWF {matched_ins:,.0f}</span>
    </div>
    <span style='font-size:11px;color:#14532d;font-family:monospace'>
      Legitimate patient journey confirmed — clinic visit → pharmacy dispensing
    </span>
  </div>
</div>""", unsafe_allow_html=True)

    with st.expander("View verified records (Table 3)", expanded=False):
        t3_search = st.text_input("🔍 Search verified records", key="t3_srch")
        t3_show = matched[[
            "ph_voucher","ph_patient","ph_rama","ph_date","ph_ins",
            "fac_voucher","fac_name","fac_date","fac_source","days_apart","name_score"
        ]].copy()
        t3_show.columns = [
            "Pharmacy Voucher","Pharmacy Patient","RAMA","Pharmacy Date","Insurance (RWF)",
            "Facility Voucher","Facility Patient","Facility Date","Facility","Days Apart","Name Score"
        ]
        for dcol in ["Pharmacy Date","Facility Date"]:
            t3_show[dcol] = pd.to_datetime(t3_show[dcol], errors="coerce").dt.strftime("%d/%m/%Y").fillna("—")
        if t3_search:
            mask = t3_show.apply(
                lambda c: c.astype(str).str.contains(t3_search, case=False, na=False)
            ).any(axis=1)
            t3_show = t3_show[mask]
        t3_show = t3_show.sort_values("Days Apart").reset_index(drop=True)
        t3_show.index = t3_show.index + 1
        st.dataframe(t3_show, use_container_width=True, height=300)

    st.markdown("<hr style='border-color:#1e2a38;margin:28px 0'>", unsafe_allow_html=True)

    # ── FULL REPORT DOWNLOAD ──────────────────────────────────────────────────
    st.markdown('<div class="sec-head">⬇️ Download Full Fraud Detection Report</div>',
                unsafe_allow_html=True)

    if st.button("📊 Generate Full Report (4 sheets)", type="primary", key="fd_gen"):
        from openpyxl import Workbook as _WB
        from openpyxl.styles import (PatternFill as _PF, Font as _F,
                                     Alignment as _Al, Border as _B, Side as _S)
        from openpyxl.utils import get_column_letter as _gcl

        wb = _WB(); wb.remove(wb.active)
        THIN = _S(border_style="thin", color="CCCCCC")
        BDR  = _B(left=THIN,right=THIN,top=THIN,bottom=THIN)

        def _make_sheet(wb, title, data_df, hdr_color, row_colors):
            ws = wb.create_sheet(title)
            for ci, col in enumerate(data_df.columns, 1):
                c = ws.cell(1, ci, col)
                c.fill = _PF("solid", fgColor=hdr_color)
                c.font = _F(bold=True, color="FFFFFF", name="Arial", size=10)
                c.alignment = _Al(horizontal="center", wrap_text=True)
                c.border = BDR
                ws.column_dimensions[_gcl(ci)].width = max(14, min(len(str(col))+4, 35))
            ws.row_dimensions[1].height = 30
            ws.freeze_panes = "A2"
            for ri, (_, row) in enumerate(data_df.iterrows(), 2):
                bg = row_colors[ri % len(row_colors)]
                for ci, val in enumerate(row, 1):
                    v = "" if (isinstance(val, float) and math.isnan(val)) else val
                    c = ws.cell(ri, ci, v)
                    c.font = _F(name="Arial", size=10)
                    c.fill = _PF("solid", fgColor=bg)
                    c.border = BDR
                    c.alignment = _Al(horizontal="left")
            return ws

        # Sheet 0: Summary
        ws0 = wb.create_sheet("Summary")
        ws0.sheet_view.showGridLines = False
        summary_data = [
            ("Pharmacy Report Period", "January 2025"),
            ("Total Pharmacy Vouchers", len(ph_work)),
            ("Facilities Loaded", fac_count),
            ("", ""),
            ("✅ Verified (visit found)", len(matched)),
            ("   % of vouchers", f"{100*len(matched)/len(ph_work):.1f}%"),
            ("   Insurance amount (RWF)", matched_ins),
            ("", ""),
            ("🔴 No Facility Record", len(no_rec)),
            ("   % of vouchers", f"{100*len(no_rec)/len(ph_work):.1f}%"),
            ("   Insurance at risk (RWF)", no_rec_ins),
            ("", ""),
            ("🟡 RAMA Found, Visit Unlinked", len(unlinked)),
            ("   % of vouchers", f"{100*len(unlinked)/len(ph_work):.1f}%"),
            ("   Insurance (RWF)", unlinked_ins),
            ("", ""),
            ("TOTAL INSURANCE AT RISK (RWF)", at_risk_ins),
            ("As % of total claims", f"{100*at_risk_ins/total_ins:.1f}%"),
        ]
        for ri, (label, val) in enumerate(summary_data, 2):
            lc = ws0.cell(ri, 1, label)
            vc = ws0.cell(ri, 2, val)
            lc.font = _F(name="Arial", size=11, bold=("TOTAL" in str(label) or "%" not in str(label) and str(label).startswith(("✅","🔴","🟡","Pharmacy","Facilities"))))
            vc.font = _F(name="Arial", size=11, bold="TOTAL" in str(label))
            if "TOTAL" in str(label):
                lc.fill = _PF("solid", fgColor="7F1D1D")
                vc.fill = _PF("solid", fgColor="7F1D1D")
                lc.font = _F(name="Arial", size=12, bold=True, color="FFFFFF")
                vc.font = _F(name="Arial", size=12, bold=True, color="FFFFFF")
        ws0.column_dimensions["A"].width = 38
        ws0.column_dimensions["B"].width = 22

        # Sheet 1: No record
        def _fmt_date(s):
            return pd.to_datetime(s, errors="coerce").dt.strftime("%d/%m/%Y").fillna("")
        t1_xl = no_rec[["ph_voucher","ph_patient","ph_rama","ph_date","ph_ins","ph_total","ph_doctor","ph_dept"]].copy()
        t1_xl.columns = ["Voucher","Patient Name","RAMA Number","Dispensing Date","Insurance (RWF)","Total Cost (RWF)","Prescriber","Specialty"]
        t1_xl["Dispensing Date"] = _fmt_date(t1_xl["Dispensing Date"])
        _make_sheet(wb, "1 - No Facility Record", t1_xl, "7F1D1D", ["FFE4E4","FFFFFF"])

        # Sheet 2: Unlinked
        t2_xl = unlinked[["ph_voucher","ph_patient","ph_rama","ph_date","ph_ins","fac_name","fac_date","fac_source","days_apart","name_score"]].copy()
        t2_xl.columns = ["Voucher","Patient Name","RAMA","Pharmacy Date","Insurance (RWF)","Facility Patient","Facility Date","Facility","Days Apart","Name Score"]
        t2_xl["Pharmacy Date"] = _fmt_date(t2_xl["Pharmacy Date"])
        t2_xl["Facility Date"] = _fmt_date(t2_xl["Facility Date"])
        _make_sheet(wb, "2 - Unlinked Visits", t2_xl, "78350F", ["FEF3C7","FFFFFF"])

        # Sheet 3: Verified
        t3_xl = matched[["ph_voucher","ph_patient","ph_rama","ph_date","ph_ins","fac_voucher","fac_name","fac_date","fac_source","days_apart","name_score"]].copy()
        t3_xl.columns = ["Voucher","Patient Name","RAMA","Pharmacy Date","Insurance (RWF)","Facility Voucher","Facility Patient","Facility Date","Facility","Days Apart","Name Score"]
        t3_xl["Pharmacy Date"] = _fmt_date(t3_xl["Pharmacy Date"])
        t3_xl["Facility Date"] = _fmt_date(t3_xl["Facility Date"])
        _make_sheet(wb, "3 - Verified", t3_xl, "14532D", ["E7F5EC","FFFFFF"])

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        st.success("✅ Full report ready!")
        st.download_button(
            "⬇️ Download Full Fraud Detection Report (.xlsx)",
            buf.getvalue(), "fraud_detection_report.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_full_fd"
        )



# ══ COUNTER-VERIFICATION REPORT TAB ══════════════════════════════════════════
with tab_cv:
  if not _dl_committed:
    _render_tab_locked("Counter-Verification Report")
  else:
    st.markdown('<div class="sec-head">📄 Counter-Verification Report Generator</div>',
                unsafe_allow_html=True)

    st.markdown("""
<div class="info-banner">
  <b>How it works:</b><br>
  Upload the annotated voucher file — the same invoice report with two extra columns
  added by the verifier: <b style='color:#00e5a0'>Difference</b> (deduction amount in RWF)
  and <b style='color:#00e5a0'>Observation</b> (reason for deduction). Rows without a
  deduction are simply left blank in those columns.<br><br>
  The app reads those columns, matches records by Paper Code, and generates the
  official two-sheet counter-verification Excel report automatically.
</div>""", unsafe_allow_html=True)

    # ── STEP 1 — Upload annotated file ────────────────────────────────────
    st.markdown('<div class="sec-head">📂 Step 1 — Upload Annotated Voucher File</div>',
                unsafe_allow_html=True)

    cv_upload = st.file_uploader(
        "Upload annotated file (Excel / CSV)",
        type=["xlsx", "xls", "csv"],
        key="cv_upload",
        help="Must be the invoice report with Difference and Observation columns filled in for deducted rows.",
    )

    # ── Parse uploaded file ────────────────────────────────────────────────
    ann_df = None
    deduction_list = []

    if cv_upload is not None:
        try:
            raw_bytes = cv_upload.read()
            fname = cv_upload.name.lower()
            if fname.endswith(".csv"):
                raw_ann = pd.read_csv(io.BytesIO(raw_bytes), encoding="utf-8",
                                      on_bad_lines="skip")
            else:
                # Try each sheet — prefer one named Sheet1, or the one with most rows
                xl_ann = pd.ExcelFile(io.BytesIO(raw_bytes))
                sheet_candidates = []
                for sn in xl_ann.sheet_names:
                    try:
                        tmp = xl_ann.parse(sn, header=0)
                        sheet_candidates.append((sn, tmp))
                    except Exception:
                        pass

                # Prefer sheet whose columns contain "difference" or "observation"
                # AND whose difference column contains meaningful deduction amounts
                # (not just tiny rounding artifacts < 100 RWF).
                def score_sheet(df_):
                    cols = " ".join(df_.columns.astype(str).str.lower())
                    if not ("diff" in cols or "observ" in cols or "remark" in cols):
                        return False
                    # Validate: difference column must have at least one value > 100
                    # to distinguish real deductions from rounding artifacts
                    for col in df_.columns:
                        if re.search(r"diff|deduct|deduit", str(col).lower()):
                            vals = pd.to_numeric(df_[col], errors="coerce").dropna()
                            if len(vals) > 0 and vals.abs().max() > 100:
                                return True
                    # Has "observ"/"remark" column but no meaningful difference values —
                    # could still be a valid annotated sheet if observation is non-empty
                    for col in df_.columns:
                        if re.search(r"observ|remark|reason|explan", str(col).lower()):
                            non_blank = df_[col].dropna()
                            non_blank = non_blank[non_blank.astype(str).str.strip() != ""]
                            if len(non_blank) > 0:
                                return True
                    return False

                preferred = next(
                    ((sn, d) for sn, d in sheet_candidates if score_sheet(d)),
                    None
                )
                # If no sheet passed validation, fall back to largest sheet by row count
                if preferred is None:
                    preferred = max(sheet_candidates, key=lambda x: len(x[1]),
                                    default=(None, None))
                chosen_sheet, raw_ann = preferred

            # ── Auto-detect columns ──────────────────────────────────────
            def _normalise_col(c):
                import re as _re
                return _re.sub(r"[^a-z0-9]", "_",
                               str(c).lower().strip()).strip("_")

            col_norm = {_normalise_col(c): c for c in raw_ann.columns}

            def _find_col(patterns):
                for pat in patterns:
                    for norm, orig in col_norm.items():
                        if re.search(pat, norm):
                            return orig
                return None

            detected = {
                "paper_code":  _find_col([r"paper_?code", r"voucher_?no", r"invoice_?no",
                                          r"invoice_?id", r"id_?invoice", r"claim_?no",
                                          r"ref_?no", r"receipt_?no", r"doc_?no"]),
                "rama":        _find_col([r"rama", r"affil", r"member", r"benef"]),
                "patient":     _find_col([r"patient_?name", r"beneficiary_?name",
                                          r"client_?name", r"nom"]),
                "difference":  _find_col([r"diff", r"deduct", r"deduit", r"montant_ded",
                                          r"amount_ded", r"ded_amount"]),
                "observation": _find_col([r"observ", r"remark", r"reason", r"explan",
                                          r"comment", r"note", r"justif", r"motif"]),
                "ins_copay":   _find_col([r"insurance_co", r"ins_cop", r"couverture",
                                          r"rama_amount"]),
                "total_cost":  _find_col([r"total_cost", r"total", r"amount", r"cout"]),
                "visit_date":  _find_col([r"dispensing", r"visit_date", r"date"]),
                "doctor":      _find_col([r"practitioner", r"prescriber", r"doctor",
                                          r"medecin"]),
            }

            # ── Warn if file looks like a raw (un-annotated) voucher report ──
            _diff_col = detected.get("difference")
            _obs_col  = detected.get("observation")
            _has_real_diffs = False
            _has_real_obs   = False
            if _diff_col and _diff_col in raw_ann.columns:
                _diff_vals = pd.to_numeric(raw_ann[_diff_col], errors="coerce").dropna()
                _has_real_diffs = len(_diff_vals) > 0 and _diff_vals.abs().max() > 100
            if _obs_col and _obs_col in raw_ann.columns:
                _obs_vals = raw_ann[_obs_col].dropna()
                _obs_vals = _obs_vals[_obs_vals.astype(str).str.strip() != ""]
                _has_real_obs = len(_obs_vals) > 0

            if not _diff_col and not _obs_col:
                st.warning(
                    "⚠️ No **Difference** or **Observation** columns detected in this file. "
                    "This looks like a raw voucher report that hasn't been annotated yet. "
                    "Please add a **Difference** column (deduction amount in RWF) and an "
                    "**Observation** column (reason for deduction) to each deducted row, "
                    "then re-upload."
                )
            elif _diff_col and not _has_real_diffs:
                st.warning(
                    f"⚠️ The detected **Difference** column (*{_diff_col}*) contains only "
                    f"very small values (< 100 RWF). This is likely a rounding artifact, "
                    f"not real counter-verification deductions. "
                    f"Please fill in the actual deduction amounts (e.g. 20,000 RWF) "
                    f"in that column before generating the report."
                )
            if _diff_col and _has_real_diffs and not _has_real_obs:
                st.warning(
                    f"⚠️ The **Observation** column (*{_obs_col or 'not found'}*) is empty. "
                    f"No deduction reasons will appear in the report. "
                    f"Please fill in the reason for each deduction."
                )

            st.session_state["ann_df"]       = raw_ann
            st.session_state["ann_detected"] = detected

        except Exception as e:
            st.error(f"❌ Could not read file: {e}")
            import traceback; st.code(traceback.format_exc())

    # Load from session state if already uploaded
    if "ann_df" in st.session_state:
        raw_ann  = st.session_state["ann_df"]
        detected = st.session_state["ann_detected"]

        # ── STEP 2 — Confirm / override column mapping ─────────────────────
        st.markdown('<div class="sec-head">🔗 Step 2 — Confirm Column Mapping</div>',
                    unsafe_allow_html=True)
        st.markdown("""
<div style='font-size:11px;color:#64748b;margin-bottom:10px;font-family:monospace'>
  Columns detected automatically. Correct any mismatches — only
  <b style='color:#e2e8f0'>Paper Code</b>, <b style='color:#e2e8f0'>Difference</b>
  and <b style='color:#e2e8f0'>Observation</b> are required.
</div>""", unsafe_allow_html=True)

        col_opts = ["(none)"] + raw_ann.columns.tolist()
        def _sel_idx(col_name):
            return col_opts.index(col_name) if col_name in col_opts else 0

        cc1, cc2, cc3 = st.columns(3)
        with cc1:
            sel_pc  = st.selectbox("📄 Paper Code",
                                   col_opts, index=_sel_idx(detected["paper_code"]),
                                   key="cv_sel_pc")
            sel_obs = st.selectbox("💬 Observation / Reason",
                                   col_opts, index=_sel_idx(detected["observation"]),
                                   key="cv_sel_obs")
        with cc2:
            sel_dif = st.selectbox("💰 Difference (amount deducted)",
                                   col_opts, index=_sel_idx(detected["difference"]),
                                   key="cv_sel_dif")
            sel_ins = st.selectbox("🏥 Insurance Co-payment",
                                   col_opts, index=_sel_idx(detected["ins_copay"]),
                                   key="cv_sel_ins")
        with cc3:
            sel_tot = st.selectbox("💵 Total Cost",
                                   col_opts, index=_sel_idx(detected["total_cost"]),
                                   key="cv_sel_tot")
            sel_pat = st.selectbox("👤 Patient Name",
                                   col_opts, index=_sel_idx(detected["patient"]),
                                   key="cv_sel_pat")

        # Preview
        with st.expander("🔍 Preview uploaded file", expanded=False):
            st.dataframe(raw_ann.head(30), use_container_width=True, height=280)

        # ── Build deduction list ─────────────────────────────────────────
        if sel_pc != "(none)" and sel_dif != "(none)":

            def _to_float(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return 0.0
                try:
                    return float(str(v).replace(",", "").replace(" ", ""))
                except ValueError:
                    return 0.0

            ann_work = raw_ann.copy()
            ann_work["_pc"]  = ann_work[sel_pc].astype(str).str.strip()
            ann_work["_dif"] = ann_work[sel_dif].apply(_to_float)
            ann_work["_obs"] = (ann_work[sel_obs].fillna("").astype(str).str.strip()
                                if sel_obs != "(none)" else "")
            ann_work["_ins"] = (ann_work[sel_ins].apply(_to_float)
                                if sel_ins != "(none)" else 0.0)
            ann_work["_tot"] = (ann_work[sel_tot].apply(_to_float)
                                if sel_tot != "(none)" else 0.0)
            ann_work["_pat"] = (ann_work[sel_pat].fillna("").astype(str).str.strip()
                                if sel_pat != "(none)" else "")

            # Identify real deduction rows.
            # Handles both sign conventions (positive or negative = deducted).
            # Keeps a row when EITHER:
            #   • |difference| >= 100 RWF  (meaningful amount even if no note), OR
            #   • observation is non-blank  (annotated by verifier, even if small)
            # This discards sub-100-RWF rounding artifacts that have no annotation.
            _has_obs   = ann_work["_obs"].str.strip() != ""
            _big_amt   = ann_work["_dif"].abs() >= 100
            deducted_rows = ann_work[_big_amt | _has_obs].copy()

            # Pull RAMA from main loaded df if possible
            rama_lookup = {}
            vid_main = "voucher_id" if "voucher_id" in df.columns else None
            pid_main = "patient_id" if "patient_id" in df.columns else None
            if vid_main and pid_main:
                for _, r in df[[vid_main, pid_main]].iterrows():
                    rama_lookup[str(r[vid_main]).strip()] = str(r[pid_main]).strip()

            deduction_list = []
            for _, r in deducted_rows.iterrows():
                pc   = r["_pc"]
                rama = rama_lookup.get(pc, "—")
                deduction_list.append({
                    "paper_code":  pc,
                    "rama_no":     rama,
                    "patient":     r["_pat"],
                    "amount":      -abs(r["_dif"]),  # always negative regardless of source sign
                    "explanation": r["_obs"],
                    "ins_copay":   r["_ins"],
                    "total_cost":  r["_tot"],
                })

            # ── STEP 3 — Deduction summary ─────────────────────────────
            st.markdown('<div class="sec-head">📊 Step 3 — Deduction Summary</div>',
                        unsafe_allow_html=True)

            total_ded    = sum(d["amount"]   for d in deduction_list)
            total_ins    = ann_work["_ins"].sum() if sel_ins != "(none)" else 0
            net_payable  = total_ins - total_ded

            sm1, sm2, sm3, sm4 = st.columns(4)
            sm1.metric("Rows with deductions",    f"{len(deduction_list):,}")
            sm2.metric("Total deducted (RWF)",    f"{total_ded:,.0f}")
            sm3.metric("Total insurance claims",  f"{total_ins:,.0f}")
            sm4.metric("Net payable (RWF)",       f"{net_payable:,.0f}")

            # Observation breakdown
            obs_counts = {}
            for d in deduction_list:
                key = d["explanation"].strip().lower()
                obs_counts[key] = obs_counts.get(key, 0) + 1
            if obs_counts:
                st.markdown("""
<div style='font-size:11px;color:#64748b;font-family:monospace;margin:8px 0 4px'>
  <b style='color:#e2e8f0'>Deduction reasons breakdown:</b>
</div>""", unsafe_allow_html=True)
                reason_cols = st.columns(min(len(obs_counts), 4))
                for i, (reason, cnt) in enumerate(
                        sorted(obs_counts.items(), key=lambda x: -x[1])):
                    reason_cols[i % len(reason_cols)].metric(
                        reason[:40] or "(blank)", cnt)

            # Deduction table
            ded_tbl = pd.DataFrame([{
                "#":             i + 1,
                "Paper Code":    d["paper_code"],
                "Patient":       d["patient"],
                "RAMA No.":      d["rama_no"],
                "Ins. Co-pay":   d["ins_copay"],
                "Deducted (RWF)":d["amount"],
                "Observation":   d["explanation"],
            } for i, d in enumerate(deduction_list)])
            st.dataframe(ded_tbl, use_container_width=True, height=320)

        else:
            st.info("👆 Assign at least the **Paper Code** and **Difference** columns above to continue.")

        # ── STEP 4 — Report metadata ───────────────────────────────────────
        if deduction_list:
            st.markdown('<div class="sec-head">📋 Step 4 — Report Metadata</div>',
                        unsafe_allow_html=True)

            mc1, mc2 = st.columns(2)
            with mc1:
                cv_province = st.text_input("Province",                value="WESTERN PROVINCE", key="cv_prov")
                cv_district = st.text_input("Administrative District", value="RUBAVU",            key="cv_dist")
                cv_pharmacy = st.text_input("Pharmacy",
                    value="PHARMACIE VINCA GISENYI LTD", key="cv_pharm")
            with mc2:
                if "date_min" in s and "date_max" in s:
                    default_period = f"{s['date_min']} to {s['date_max']}"
                else:
                    default_period = ""
                cv_period  = st.text_input("Period",          value=default_period, key="cv_per")
                cv_code    = st.text_input("Code",            value="",             key="cv_code")
                cv_prep    = st.text_input("Prepared by",     value="",             key="cv_prep")

            sc1, sc2 = st.columns(2)
            with sc1:
                cv_verified = st.text_input("Verified by",
                    value="Alphonsine MUKAKAYIBANDA", key="cv_verif")
            with sc2:
                cv_approved = st.text_input("Approved by", value="", key="cv_approv")

            # ── STEP 5 — Generate ──────────────────────────────────────────
            st.markdown('<div class="sec-head">⬇️ Step 5 — Generate Report</div>',
                        unsafe_allow_html=True)

            if st.button("📊 Generate Counter-Verification Excel Report",
                         type="primary", key="cv_gen_btn"):
                meta = {
                    "province": cv_province,
                    "district": cv_district,
                    "pharmacy": cv_pharmacy,
                    "period":   cv_period,
                    "code":     cv_code,
                }

                # Merge annotated file into df for Sheet1 (use uploaded file rows)
                # Build a combined df from ann_work so we have Difference + Observation
                ann_for_report = ann_work.copy()

                with st.spinner("Building Excel report…"):
                    try:
                        xlsx_bytes = generate_counter_verification_xlsx(
                            df          = ann_for_report,
                            deductions  = deduction_list,
                            meta        = meta,
                            prepared_by = cv_prep,
                            verified_by = cv_verified,
                            approved_by = cv_approved,
                            pc_col      = sel_pc,
                            ins_col     = sel_ins if sel_ins != "(none)" else None,
                            tot_col     = sel_tot if sel_tot != "(none)" else None,
                            obs_col     = sel_obs if sel_obs != "(none)" else None,
                            dif_col     = sel_dif,
                        )
                        st.session_state["cv_xlsx"]      = xlsx_bytes
                        st.session_state["cv_generated"] = True
                        st.session_state["cv_fname_meta"] = (cv_pharmacy, cv_period)
                    except Exception as e:
                        st.error(f"❌ Failed to generate report: {e}")
                        import traceback; st.code(traceback.format_exc())

            if st.session_state.get("cv_generated"):
                pharm_s, period_s = st.session_state.get("cv_fname_meta", ("pharmacy", "period"))
                fname = (f"counter_verification_"
                         f"{pharm_s.replace(' ','_')[:25]}_"
                         f"{period_s.replace(' ','_').replace('/','')[:15]}.xlsx")
                st.success("✅ Report ready!")
                st.download_button(
                    label     = "⬇️ Download Counter-Verification Report (.xlsx)",
                    data      = st.session_state["cv_xlsx"],
                    file_name = fname,
                    mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key       = "cv_download",
                )
                st.markdown("""
<div style='font-size:12px;color:#64748b;font-family:monospace;margin-top:8px'>
  <b style='color:#e2e8f0'>Sheet 1</b> — "After counter verification":
  full records table with 100% and 85% columns; deducted rows highlighted in amber.<br>
  <b style='color:#e2e8f0'>Sheet 2</b> — "Counter verification report":
  official summary with deduction list, totals, and signature block.
</div>""", unsafe_allow_html=True)

    else:
        st.info("👆 Upload your annotated voucher file to get started.")

# ══ DATA PREP TAB ════════════════════════════════════════════════════════════
with tab_dataprep:

    # ── Step progress bar helper ──────────────────────────────────────────────
    def _step_bar(active: int):
        steps = [
            ("1", "Load File"),
            ("2", "Map Columns"),
            ("3", "Review & Clean"),
            ("4", "Commit to Lake"),
        ]
        html = "<div style='display:flex;gap:0;margin-bottom:28px;align-items:stretch'>"
        for i, (num, label) in enumerate(steps, 1):
            is_done   = i < active
            is_active = i == active
            is_locked = i > active
            if is_done:
                bg, border, text_col, num_col = "#031a0a","#16a34a","#22c55e","#22c55e"
                icon = "✓"
            elif is_active:
                bg, border, text_col, num_col = "#030d1a","#0ea5e9","#38bdf8","#0ea5e9"
                icon = num
            else:
                bg, border, text_col, num_col = "#0d1117","#1e2a38","#334155","#334155"
                icon = num
            sep = "" if i == len(steps) else "<div style='width:2px;background:#1e2a38;flex-shrink:0'></div>"
            html += f"""
<div style='flex:1;background:{bg};border:1px solid {border};padding:12px 16px;
     {"border-radius:10px 0 0 10px" if i==1 else "border-radius:0 10px 10px 0" if i==len(steps) else ""};
     border-left:{"3px" if is_active else "1px"} solid {border}'>
  <div style='display:flex;align-items:center;gap:8px'>
    <div style='width:22px;height:22px;border-radius:50%;border:2px solid {num_col};
         display:flex;align-items:center;justify-content:center;
         font-size:11px;font-weight:700;color:{num_col};flex-shrink:0'>{icon}</div>
    <div style='font-size:12px;font-weight:{"700" if is_active else "500"};
         color:{text_col};font-family:Syne,sans-serif'>{label}</div>
  </div>
</div>{sep}"""
        html += "</div>"
        st.markdown(html, unsafe_allow_html=True)

    # ── Load raw file ─────────────────────────────────────────────────────────
    _raw_bytes = st.session_state.get("raw_bytes")
    _raw_fname = st.session_state.get("raw_filename", "")

    if not _raw_bytes:
        st.markdown("""
<div style='display:flex;flex-direction:column;align-items:center;justify-content:center;
     padding:60px 20px;text-align:center'>
  <div style='font-size:42px;margin-bottom:14px'>👈</div>
  <div style='font-family:Syne,sans-serif;font-size:20px;font-weight:700;
       color:#e2e8f0;margin-bottom:8px'>Upload a file first</div>
  <div style='font-size:13px;color:#475569;font-family:monospace'>
    Use the <b style='color:#0ea5e9'>Source File</b> uploader in the sidebar,
    then return here to map columns and commit to the data lake.
  </div>
</div>""", unsafe_allow_html=True)
        st.stop()

    # ── Parse raw file (cached on session key) ────────────────────────────────
    _dp_cache_key = f"dp_raw_{_raw_fname}"
    if _dp_cache_key not in st.session_state:
        with st.spinner(f"Loading {_raw_fname}…"):
            try:
                _fn = _raw_fname.lower()
                if _fn.endswith(".csv"):
                    _dp_df = pd.read_csv(io.BytesIO(_raw_bytes), encoding="utf-8",
                                         on_bad_lines="skip")
                elif _fn.endswith(".ods"):
                    _dp_df = pd.read_excel(io.BytesIO(_raw_bytes), engine="odf")
                else:
                    _xl = pd.ExcelFile(io.BytesIO(_raw_bytes))
                    # Best sheet = most columns; tie-break = most rows
                    _scores = {}
                    for _sn in _xl.sheet_names:
                        try:
                            _tmp = _xl.parse(_sn, nrows=5)
                            _scores[_sn] = len(_tmp.columns) * 1000 + len(_xl.parse(_sn, nrows=500))
                        except Exception:
                            _scores[_sn] = 0
                    _best_sh = max(_scores, key=_scores.get)
                    _dp_df = _xl.parse(_best_sh)
                    # Auto-detect header row
                    for _hi, _hrow in _dp_df.head(10).iterrows():
                        _joined = " ".join(str(v).lower() for v in _hrow if pd.notna(v))
                        if any(k in _joined for k in ["rama","patient","dispensing","voucher","date","affil"]):
                            if _hi > 0:
                                _dp_df = _xl.parse(_best_sh, header=_hi)
                            break
                st.session_state[_dp_cache_key] = _dp_df
            except Exception as _e:
                st.error(f"❌ Could not load file: {_e}")
                import traceback; st.code(traceback.format_exc())
                st.stop()
    _dp_df = st.session_state[_dp_cache_key]

    # ─────────────────────────────────────────────────────────────────────────
    # STEP 1 — File overview
    # ─────────────────────────────────────────────────────────────────────────
    _step_bar(1 if not st.session_state.get("dp_step1_done") else
              2 if not st.session_state.get("dp_step2_done") else
              3 if not st.session_state.get("dp_step3_done") else 4)

    with st.expander("📄 Step 1 — File Overview", expanded=not st.session_state.get("dp_step1_done", False)):
        _s1c1, _s1c2, _s1c3, _s1c4 = st.columns(4)
        _s1c1.metric("File", _raw_fname[:30])
        _s1c2.metric("Rows", f"{len(_dp_df):,}")
        _s1c3.metric("Columns", len(_dp_df.columns))
        _s1c4.metric("Est. size", f"{_dp_df.memory_usage(deep=True).sum() / 1024:.0f} KB")

        # Sheet selector if Excel multi-sheet
        _fn = _raw_fname.lower()
        if not _fn.endswith(".csv") and not _fn.endswith(".ods"):
            try:
                _xl2 = pd.ExcelFile(io.BytesIO(_raw_bytes))
                if len(_xl2.sheet_names) > 1:
                    _selected_sheet = st.selectbox(
                        "Sheet", _xl2.sheet_names, key="dp_sheet_sel",
                        help="Switch sheet — the app will re-profile")
                    if st.button("🔄 Load selected sheet", key="dp_reload_sheet"):
                        _dp_df2 = _xl2.parse(_selected_sheet)
                        st.session_state[_dp_cache_key] = _dp_df2
                        for _k in ["dp_step1_done","dp_step2_done","dp_step3_done","dp_mapping_confirmed"]:
                            st.session_state.pop(_k, None)
                        st.rerun()
            except Exception:
                pass

        st.markdown('<div style="font-size:11px;font-weight:700;color:#64748b;'
                    'text-transform:uppercase;letter-spacing:.06em;margin:14px 0 6px">'
                    'Column Preview</div>', unsafe_allow_html=True)

        # Column quality summary table
        _qual_rows = []
        for _c in _dp_df.columns:
            _s_col = _dp_df[_c]
            _null_pct = round(100 * _s_col.isna().sum() / max(len(_s_col), 1), 1)
            _uniq = int(_s_col.nunique())
            _samples = " · ".join(str(v) for v in _s_col.dropna().head(3).tolist())[:60]
            _fill_color = "#031a0a" if _null_pct < 5 else "#1a1000" if _null_pct < 30 else "#1a0505"
            _fill_text  = "#22c55e" if _null_pct < 5 else "#f59e0b" if _null_pct < 30 else "#ef4444"
            _qual_rows.append({
                "Column": str(_c)[:40],
                "Type": str(_s_col.dtype),
                "Fill %": f"{100-_null_pct:.0f}%",
                "Unique": _uniq,
                "Sample values": _samples,
            })
        _qual_df = pd.DataFrame(_qual_rows)

        def _color_fill(val):
            try:
                v = int(val.replace("%",""))
                if v >= 95: return "color:#22c55e"
                if v >= 70: return "color:#f59e0b"
                return "color:#ef4444;font-weight:bold"
            except Exception:
                return ""

        st.dataframe(
            _qual_df.style.applymap(_color_fill, subset=["Fill %"]),
            use_container_width=True, height=280,
        )

        if st.button("✅ Looks good — proceed to column mapping", type="primary", key="dp_s1_next"):
            st.session_state["dp_step1_done"] = True
            st.rerun()

    # ─────────────────────────────────────────────────────────────────────────
    # STEP 2 — Manual Column Mapping
    # ─────────────────────────────────────────────────────────────────────────
    if not st.session_state.get("dp_step1_done"):
        st.stop()

    with st.expander("🗺️ Step 2 — Map Columns to System Fields",
                     expanded=not st.session_state.get("dp_step2_done", False)):

        # ── Intro ─────────────────────────────────────────────────────────────
        st.markdown("""
<div style='background:#030d1a;border:1px solid #1e3a5f;border-radius:10px;
     padding:14px 18px;margin-bottom:20px'>
  <div style='font-size:13px;font-weight:700;color:#38bdf8;margin-bottom:8px'>
    How to use this step
  </div>
  <div style='font-size:12px;color:#64748b;font-family:monospace;line-height:1.9'>
    The left side lists every <b style='color:#e2e8f0'>system field</b> PharmaScan needs —
    grouped by category. For each one, use the dropdown on the right to pick which
    <b style='color:#e2e8f0'>column from your file</b> contains that data.<br>
    Fields marked <span style='color:#ef4444;font-weight:700'>REQUIRED</span> must be
    mapped for fraud detection to work.
    Leave optional fields as <code style='color:#475569'>— not in this file</code> if they
    don't exist in your data.
  </div>
</div>""", unsafe_allow_html=True)

        # ── Build source column options ────────────────────────────────────────
        # For each selectbox: blank option + every actual column in the file
        _src_cols     = list(_dp_df.columns)
        _col_none_opt = "— not in this file"
        _col_options  = [_col_none_opt] + _src_cols

        # Quick dtype + sample preview per source column (computed once)
        _col_meta = {}
        for _c in _src_cols:
            _s_c = _dp_df[_c]
            _dtype = str(_s_c.dtype)
            _uniq  = int(_s_c.nunique())
            _samp  = " · ".join(
                str(v) for v in _s_c.dropna().head(3).tolist()
            )[:55]
            _null_pct = round(100 * _s_c.isna().sum() / max(len(_s_c), 1), 1)
            _col_meta[_c] = {"dtype": _dtype, "unique": _uniq,
                             "sample": _samp, "null_pct": _null_pct}

        # ── Mapping selectors — one per system field, grouped by category ─────
        _user_mapping = {}   # field_name → source_col

        for _grp in _GROUPS_ORDER:
            _fields_in_grp = [(f, d) for f, d in _SYSTEM_FIELDS.items()
                              if d["group"] == _grp]
            if not _fields_in_grp:
                continue

            _gc = _GROUP_COLORS.get(_grp, "#64748b")

            # Section divider
            st.markdown(f"""
<div style='display:flex;align-items:center;gap:10px;margin:22px 0 14px'>
  <div style='height:1px;background:#1e2a38;flex:1'></div>
  <span style='font-size:11px;font-weight:700;color:{_gc};text-transform:uppercase;
       letter-spacing:.09em;background:#080c10;padding:0 12px'>{_grp}</span>
  <div style='height:1px;background:#1e2a38;flex:1'></div>
</div>""", unsafe_allow_html=True)

            for _fname, _fdef in _fields_in_grp:
                _is_req   = _fdef["required"]
                _req_pill = (
                    "<span style='background:#450a0a;color:#fca5a5;font-size:9px;"
                    "font-weight:700;padding:1px 6px;border-radius:4px;"
                    "letter-spacing:.04em;margin-left:6px'>REQUIRED</span>"
                ) if _is_req else (
                    "<span style='background:#0d1117;color:#334155;font-size:9px;"
                    "padding:1px 6px;border-radius:4px;margin-left:6px'>optional</span>"
                )

                _left, _right = st.columns([2, 3])

                with _left:
                    st.markdown(f"""
<div style='padding:10px 0 6px'>
  <div style='font-size:13px;font-weight:700;color:#e2e8f0;
       font-family:Syne,sans-serif;line-height:1.3'>
    {_fdef["label"]}{_req_pill}
  </div>
  <div style='font-size:10px;color:#475569;font-family:monospace;margin-top:4px'>
    system key: <code style='color:{_gc}'>{_fname}</code>
  </div>
</div>""", unsafe_allow_html=True)

                with _right:
                    _sel_col = st.selectbox(
                        label=_fname,
                        options=_col_options,
                        index=0,
                        key=f"dp2_field_{_fname}",
                        label_visibility="collapsed",
                        help=f"Pick the column in your file that contains {_fdef['label']}",
                    )

                    # Show inline preview of selected column
                    if _sel_col != _col_none_opt and _sel_col in _col_meta:
                        _m = _col_meta[_sel_col]
                        _fill_c = ("#22c55e" if _m["null_pct"] < 5
                                   else "#f59e0b" if _m["null_pct"] < 30
                                   else "#ef4444")
                        st.markdown(f"""
<div style='background:#111720;border:1px solid #1e2a38;border-radius:6px;
     padding:6px 10px;margin-top:2px;font-family:monospace;font-size:10px;
     display:flex;gap:16px;flex-wrap:wrap'>
  <span style='color:#475569'>type: <b style='color:#94a3b8'>{_m["dtype"]}</b></span>
  <span style='color:#475569'>unique: <b style='color:#94a3b8'>{_m["unique"]:,}</b></span>
  <span style='color:#475569'>fill: <b style='color:{_fill_c}'>{100-_m["null_pct"]:.0f}%</b></span>
  <span style='color:#334155;flex:1;overflow:hidden;text-overflow:ellipsis;
       white-space:nowrap'>{_m["sample"] or "—"}</span>
</div>""", unsafe_allow_html=True)
                    elif _sel_col == _col_none_opt and _is_req:
                        st.markdown(
                            "<div style='font-size:10px;color:#7f1d1d;font-family:monospace;"
                            "padding:4px 0'>⚠ Required — fraud detection will be limited</div>",
                            unsafe_allow_html=True,
                        )

                    if _sel_col != _col_none_opt:
                        _user_mapping[_fname] = _sel_col

                # Light separator between fields within the same group
                st.markdown(
                    "<div style='height:1px;background:#0d1117;margin:2px 0'></div>",
                    unsafe_allow_html=True,
                )

        # ── Extra columns — what to do with unmapped source cols ──────────────
        _mapped_src_cols = set(_user_mapping.values())
        _unmapped_src    = [c for c in _src_cols if c not in _mapped_src_cols]

        if _unmapped_src:
            st.markdown("""
<div style='display:flex;align-items:center;gap:10px;margin:22px 0 14px'>
  <div style='height:1px;background:#1e2a38;flex:1'></div>
  <span style='font-size:11px;font-weight:700;color:#475569;text-transform:uppercase;
       letter-spacing:.09em;background:#080c10;padding:0 12px'>Unmapped Columns</span>
  <div style='height:1px;background:#1e2a38;flex:1'></div>
</div>""", unsafe_allow_html=True)

            st.markdown(f"""
<div style='font-size:11px;color:#475569;font-family:monospace;margin-bottom:12px'>
  {len(_unmapped_src)} column(s) from your file were not assigned to any system field.
  Choose whether to keep each one (it will be stored with a
  <code style='color:#64748b'>raw_</code> prefix) or exclude it from the data lake entirely.
</div>""", unsafe_allow_html=True)

            _exclude_cols = set()
            _uc_rows = st.columns(3)
            for _ui, _uc in enumerate(_unmapped_src):
                _m = _col_meta.get(_uc, {})
                with _uc_rows[_ui % 3]:
                    _keep = st.checkbox(
                        f"Keep  `{str(_uc)[:32]}`",
                        value=True,
                        key=f"dp2_keep_{_uc}",
                        help=(
                            f"dtype: {_m.get('dtype','?')}  |  "
                            f"unique: {_m.get('unique','?')}  |  "
                            f"sample: {_m.get('sample','—')[:60]}"
                        ),
                    )
                    if not _keep:
                        _exclude_cols.add(_uc)
        else:
            _exclude_cols = set()

        # ── Validation + confirm ──────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)

        _missing_req = [f for f in _SYSTEM_FIELDS
                        if _SYSTEM_FIELDS[f]["required"] and f not in _user_mapping]
        _mapped_count = len(_user_mapping)

        # Status bar
        _sb_mapped   = _mapped_count
        _sb_optional = sum(1 for f in _SYSTEM_FIELDS
                           if not _SYSTEM_FIELDS[f]["required"] and f in _user_mapping)
        _sb_req_done = len(_SYSTEM_FIELDS) - len(_missing_req) - _sb_optional - \
                       (len(_SYSTEM_FIELDS) - sum(1 for d in _SYSTEM_FIELDS.values()
                                                  if d["required"]))

        _vm1, _vm2, _vm3 = st.columns(3)
        _vm1.metric("System fields mapped",   _mapped_count)
        _vm2.metric("Required fields filled",
                    f"{sum(1 for f in _SYSTEM_FIELDS if _SYSTEM_FIELDS[f]['required'] and f in _user_mapping)}"
                    f" / {sum(1 for d in _SYSTEM_FIELDS.values() if d['required'])}")
        _vm3.metric("Columns excluded",        len(_exclude_cols))

        if _missing_req:
            _mr_labels = "  ·  ".join(
                f"{_SYSTEM_FIELDS[f]['label']}" for f in _missing_req
            )
            st.warning(
                f"⚠️ Required fields not yet mapped: **{_mr_labels}**. "
                f"You can still proceed, but fraud detection modules that need "
                f"these fields will be unavailable."
            )

        _s2c1, _s2c2 = st.columns([1, 3])
        with _s2c1:
            if st.button("✅ Confirm mapping", type="primary", key="dp_s2_next"):
                st.session_state["dp_confirmed_mapping"] = dict(_user_mapping)
                st.session_state["dp_confirmed_exclude"] = set(_exclude_cols)
                st.session_state["dp_step2_done"]        = True
                st.rerun()

    # ─────────────────────────────────────────────────────────────────────────
    # STEP 3 — Review & Clean
    # ─────────────────────────────────────────────────────────────────────────
    if not st.session_state.get("dp_step2_done"):
        st.stop()

    _confirmed_map = st.session_state.get("dp_confirmed_mapping", {})
    _confirmed_exc = st.session_state.get("dp_confirmed_exclude", set())

    with st.expander("🔍 Step 3 — Review & Clean Data",
                     expanded=not st.session_state.get("dp_step3_done", False)):

        # Apply mapping to get preview df
        _prev_clean = apply_column_mapping(
            _dp_df.drop(columns=list(_confirmed_exc), errors="ignore"),
            _confirmed_map,
        )

        _p1, _p2, _p3 = st.columns(3)
        _p1.metric("Rows ready",           f"{len(_prev_clean):,}")
        _p2.metric("System fields mapped", len(_confirmed_map))
        _p3.metric("Columns excluded",     len(_confirmed_exc))

        # ── Rapid-revisit window (needed for processing) ───────────────────
        st.markdown('<div style="font-size:11px;font-weight:700;color:#64748b;'
                    'text-transform:uppercase;letter-spacing:.06em;margin:14px 0 6px">'
                    'Analysis Settings</div>', unsafe_allow_html=True)
        _dp_r1, _dp_r2, _dp_r3 = st.columns(3)
        with _dp_r1:
            _dp_rapid_days = st.slider("Rapid revisit window (days)", 1, 30, 7,
                                       key="dp_rapid_days",
                                       help="Days between visits to flag as suspicious")
        with _dp_r2:
            _dp_top_n = st.slider("Top N for charts", 5, 25, 15, key="dp_top_n")

        # ── Missing value summary ──────────────────────────────────────────
        st.markdown('<div style="font-size:11px;font-weight:700;color:#64748b;'
                    'text-transform:uppercase;letter-spacing:.06em;margin:14px 0 6px">'
                    'Data Quality Check</div>', unsafe_allow_html=True)

        _qual2 = []
        for _c in _prev_clean.columns:
            _null_n = int(_prev_clean[_c].isna().sum())
            _null_p = round(100 * _null_n / max(len(_prev_clean), 1), 1)
            _status = "✅ OK" if _null_p < 5 else "⚠️ Sparse" if _null_p < 30 else "🔴 High nulls"
            _qual2.append({"Column": _c, "Nulls": _null_n,
                           "Null %": f"{_null_p:.1f}%", "Status": _status})
        st.dataframe(pd.DataFrame(_qual2), use_container_width=True, height=220)

        # ── Data preview ──────────────────────────────────────────────────
        st.markdown('<div style="font-size:11px;font-weight:700;color:#64748b;'
                    'text-transform:uppercase;letter-spacing:.06em;margin:14px 0 6px">'
                    'Prepared Data Preview (first 50 rows)</div>', unsafe_allow_html=True)
        st.dataframe(_prev_clean.head(50), use_container_width=True, height=280)

        # ── Doctor / Prescriber name normalisation ────────────────────────
        _doc_col_dp = next(
            (c for c in ["doctor_name", "practitioner_name"]
             if c in _prev_clean.columns),
            None,
        )
        if _doc_col_dp:
            st.markdown(
                '<div style="font-size:11px;font-weight:700;color:#a78bfa;'
                'text-transform:uppercase;letter-spacing:.06em;margin:18px 0 6px">'
                '✏️ Prescriber Name Normalisation</div>',
                unsafe_allow_html=True,
            )
            st.markdown(f"""
<div style='background:#0d0714;border:1px solid #3b1f7a;border-radius:8px;
     padding:10px 14px;margin-bottom:12px;font-size:11px;
     font-family:monospace;color:#64748b'>
  Found <b style='color:#e2e8f0'>{_prev_clean[_doc_col_dp].nunique():,}</b> unique
  prescriber names in column
  <code style='color:#a78bfa'>{_doc_col_dp}</code>.
  The engine below detects likely duplicates (typos, reordering, initials)
  and suggests merges. Approve the ones you want before committing to the lake.
</div>""", unsafe_allow_html=True)

            _norm_thresh_dp = st.slider(
                "Detection sensitivity",
                0.70, 0.99, 0.82, 0.01,
                key="dp3_norm_thresh",
                help="Higher = stricter matching; fewer but more certain suggestions",
            )

            _run_norm = st.button(
                "🔍 Detect name variants",
                key="dp3_run_norm",
                type="secondary",
            )
            if _run_norm or st.session_state.get("dp3_clusters"):
                if _run_norm:
                    with st.spinner("Clustering names…"):
                        _doc_names = _prev_clean[_doc_col_dp].dropna().unique().tolist()
                        _doc_counts = (
                            _prev_clean[_doc_col_dp]
                            .value_counts()
                            .to_dict()
                        )
                        _clusters = detect_name_clusters(_doc_names, _doc_counts)
                    st.session_state["dp3_clusters"] = _clusters

                                   _clusters = st.session_state.get("dp3_clusters", [])

            # 🛠️ FIX: Recalculate counts here so they're available for the display loop
            _doc_counts = _prev_clean[_doc_col_dp].value_counts().to_dict()

            if not _clusters:
                    st.success("✅ No duplicate name variants detected.")
                else:
                    st.markdown(
                        f'<p style="font-size:11px;color:#a78bfa;font-family:monospace">'
                        f'{len(_clusters)} cluster(s) detected — approve merges below</p>',
                        unsafe_allow_html=True,
                    )
                    _approved_dp = []
                    for _ci_dp, _cl in enumerate(_clusters[:40]):
                        _canon  = _cl["canonical"]
                        _vars   = _cl["variants"]
                        _conf   = _cl["confidence"]
                        _susp   = _cl.get("suspicious", False)
                        _conf_c = (DANGER if _susp else
                                   WARN if _conf < 0.85 else ACCENT)
                        _freq_total = sum(
                            _doc_counts.get(v, 0) for v in _vars
                        )
                        _variant_pills = "".join(
                            f'<span style="background:#1e2a38;color:#94a3b8;'
                            f'font-size:10px;font-family:monospace;padding:2px 8px;'
                            f'border-radius:4px;margin:2px;display:inline-block">'
                            f'{v[:32]} <span style="color:#475569">×{_doc_counts.get(v,0)}</span>'
                            f'</span>'
                            for v in _vars[:8]
                        )
                        _key_dp = f"dp3_cl_{_ci_dp}"
                        _checked = st.session_state.get(_key_dp, _conf >= 0.88 and not _susp)
                        st.markdown(f"""
<div style='background:#111720;border:1px solid #1e2a38;
     border-left:3px solid {_conf_c};border-radius:8px;
     padding:10px 14px;margin-bottom:8px'>
  <div style='display:flex;justify-content:space-between;align-items:flex-start;
       flex-wrap:wrap;gap:6px'>
    <div>
      <span style='font-size:13px;font-weight:700;color:#e2e8f0'>
        {_canon}
      </span>
      <span style='font-size:10px;color:#475569;font-family:monospace;margin-left:8px'>
        canonical · {_doc_counts.get(_canon,0)} occurrences
      </span>
    </div>
    <span style='font-size:10px;font-family:monospace;color:{_conf_c}'>
      {'⚠ review' if _susp else f'{_conf*100:.0f}% confidence'}
    </span>
  </div>
  <div style='font-size:11px;color:#64748b;font-family:monospace;margin:4px 0'>
    {len(_vars)} variant(s) · {_freq_total} rows affected
  </div>
  <div style='display:flex;flex-wrap:wrap;gap:3px;margin-top:4px'>
    {_variant_pills}
  </div>
</div>""", unsafe_allow_html=True)
                        _checked = st.checkbox(
                            f"Merge → {_canon[:40]}",
                            value=_checked,
                            key=_key_dp,
                        )
                        if _checked:
                            _approved_dp.append(_cl)

                    if _approved_dp:
                        st.markdown(
                            f'<p style="font-size:11px;color:#64748b;'
                            f'font-family:monospace">'
                            f'{len(_approved_dp)} cluster(s) selected · '
                            f'{sum(len(c["variants"]) for c in _approved_dp)} '
                            f'variant names will be merged</p>',
                            unsafe_allow_html=True,
                        )
                        if st.button(
                            "⚡ Apply name normalisation",
                            key="dp3_apply_norm",
                            type="secondary",
                        ):
                            _prev_clean = apply_name_normalisation(
                                _prev_clean, _doc_col_dp, _approved_dp
                            )
                            st.session_state["dp_preview_clean"] = _prev_clean
                            st.session_state.pop("dp3_clusters", None)
                            _n_merged = sum(len(c["variants"])
                                            for c in _approved_dp)
                            st.success(
                                f"✅ Merged {_n_merged} variant name(s) into "
                                f"{len(_approved_dp)} canonical form(s) in "
                                f"`{_doc_col_dp}`."
                            )
                            st.rerun()
        else:
            st.markdown("""
<div style='font-size:11px;color:#334155;font-family:monospace;
     padding:8px 0;margin-top:6px'>
  ℹ️ No prescriber name column detected in the prepared data.
  Map a <code>doctor_name</code> field in Step 2 to enable name normalisation.
</div>""", unsafe_allow_html=True)

      # ── Column rename overrides ────────────────────────────────────────
        st.markdown('<div class="sec-head">🔧 Optional: rename raw_ columns</div>', unsafe_allow_html=True)
        
        st.markdown("""<div style='font-size:11px;color:#64748b;
        font-family:monospace;margin-bottom:8px'>
        Columns that were not mapped to a system field are kept with a
        <code>raw_</code> prefix. You can give them friendlier names below.
        Leave blank to keep the auto-generated name.</div>""",
        unsafe_allow_html=True)
        
        _raw_cols = [c for c in _prev_clean.columns if c.startswith("raw_")]
        _rename_overrides = {}
        if _raw_cols:
            _rnc = st.columns(min(3, len(_raw_cols)))
            for _ri, _rc in enumerate(_raw_cols):
                _new_name = _rnc[_ri % 3].text_input(
                    _rc, value="", placeholder=f"{_rc}", key=f"dp_rename_{_rc}",
                    label_visibility="visible",
                )
                if _new_name.strip():
                    _rename_overrides[_rc] = _new_name.strip()
        
        if _rename_overrides:
            _prev_clean = _prev_clean.rename(columns=_rename_overrides)
            st.session_state["dp_rename_overrides"] = _rename_overrides
    # ─────────────────────────────────────────────────────────────────────────
    # STEP 4 — Commit to Data Lake
    # ─────────────────────────────────────────────────────────────────────────
    if not st.session_state.get("dp_step3_done"):
        st.stop()

    _final_df        = st.session_state.get("dp_preview_clean",
                           apply_column_mapping(
                               _dp_df.drop(columns=list(_confirmed_exc), errors="ignore"),
                               _confirmed_map))
    _final_rapid     = st.session_state.get("dp_rapid_days_val", 7)
    _final_top_n     = st.session_state.get("dp_top_n_val", 15)

    # Apply any rename overrides
    _renames = st.session_state.get("dp_rename_overrides", {})
    if _renames:
        _final_df = _final_df.rename(columns=_renames)

    with st.expander("🚀 Step 4 — Commit to Data Lake", expanded=True):

        # Final summary before commit
        _f1, _f2, _f3, _f4 = st.columns(4)
        _f1.metric("Rows to commit",      f"{len(_final_df):,}")
        _f2.metric("System fields",       len(_confirmed_map))
        _f3.metric("Total columns",       len(_final_df.columns))
        _f4.metric("Rapid window",        f"{_final_rapid} days")

        # Mapping summary chips
        _chip_parts = []
        for _fld, _orig in _confirmed_map.items():
            _gc2 = _GROUP_COLORS.get(_SYSTEM_FIELDS.get(_fld, {}).get("group",""), "#64748b")
            _chip_parts.append(
                f'<span style="background:rgba(14,165,233,.07);border:1px solid #1e3a5f;'
                f'border-radius:6px;padding:3px 9px;font-size:10px;'
                f'font-family:monospace;color:{_gc2};margin:2px">'
                f'{_fld} ← <span style="color:#475569">{str(_orig)[:22]}</span></span>'
            )
        if _chip_parts:
            st.markdown(
                f'<div style="display:flex;flex-wrap:wrap;gap:3px;margin:10px 0 16px">'
                f'{"".join(_chip_parts)}</div>',
                unsafe_allow_html=True,
            )

        # Download preview before committing
        _dpc1, _dpc2 = st.columns(2)
        with _dpc1:
            _csv_prev = _final_df.to_csv(index=False).encode()
            st.download_button("⬇️ Export prepared CSV first",
                               data=_csv_prev,
                               file_name=f"pharmascan_prepared_{_raw_fname.rsplit('.',1)[0]}.csv",
                               mime="text/csv",
                               key="dp_export_csv")
        with _dpc2:
            _map_preview = pd.DataFrame(
                [(f, c, _SYSTEM_FIELDS.get(f, {}).get("label", "—"))
                 for f, c in _confirmed_map.items()],
                columns=["System Field", "Original Column", "Description"]
            )
            st.download_button("⬇️ Export mapping report",
                               data=_map_preview.to_csv(index=False).encode(),
                               file_name="column_mapping_report.csv",
                               mime="text/csv",
                               key="dp_export_map")

        st.markdown("<br>", unsafe_allow_html=True)

        # ── THE COMMIT BUTTON ─────────────────────────────────────────────
        st.markdown("""
<div style='background:#030d1a;border:1px solid #1e3a5f;border-radius:12px;
     padding:16px 20px;margin-bottom:16px'>
  <div style='font-size:13px;font-weight:700;color:#38bdf8;margin-bottom:6px'>
    ℹ️ What happens when you commit
  </div>
  <div style='font-size:11px;color:#64748b;font-family:monospace;line-height:1.9'>
    • The prepared dataframe is stored in a <b style='color:#e2e8f0'>session-level data lake</b><br>
    • All other tabs (Summary, Fraud Detection, Rules Engine, etc.) read exclusively from this lake<br>
    • You can re-run Data Prep at any time to update the lake with a different mapping<br>
    • Clearing the sidebar file upload or clicking <b style='color:#ef4444'>Clear Data Lake</b> resets everything
  </div>
</div>""", unsafe_allow_html=True)

        _commit_col, _reset_col = st.columns([2, 1])
        with _commit_col:
            _do_commit = st.button(
                "🚀 Commit to Data Lake",
                type="primary",
                use_container_width=True,
                key="dp_commit",
            )
        with _reset_col:
            if st.button("↩ Re-do column mapping", use_container_width=True, key="dp_redo"):
                for _k in ["dp_step1_done","dp_step2_done","dp_step3_done",
                           "dp_confirmed_mapping","dp_confirmed_exclude",
                           "dp_preview_clean","dp_rename_overrides"]:
                    st.session_state.pop(_k, None)
                st.rerun()

        if _do_commit:
            # Run load_and_process on the clean prepared data
            with st.spinner("Processing data lake…"):
                try:
                    # Serialise the clean df to CSV bytes and reparse (ensures consistent types)
                    _lake_bytes = _final_df.to_csv(index=False).encode()
                    _lake_df, _lake_colmap, _lake_s, _lake_rg, _lake_rd, _lake_rapid = \
                        load_and_process(_lake_bytes, "lake_data.csv", _final_rapid)

                    from datetime import datetime as _dt
                    st.session_state["data_lake"] = {
                        "committed":      True,
                        "df":             _lake_df,
                        "stats":          _lake_s,
                        "rapid":          _lake_rapid,
                        "rapid_days":     _final_rapid,
                        "repeat_groups":  _lake_rg,
                        "repeat_detail":  _lake_rd,
                        "col_map":        _lake_colmap,
                        "top_n":          _final_top_n,
                        "filename":       _raw_fname,
                        "mapped_fields":  list(_confirmed_map.keys()),
                        "committed_at":   _dt.now().strftime("%d/%m/%Y %H:%M"),
                        "source_rows":    len(_final_df),
                    }
                    st.success(
                        f"✅ Data lake committed! "
                        f"**{len(_lake_df):,} rows** · **{len(_confirmed_map)} system fields** · "
                        f"**{len(_lake_rapid)} rapid revisit pairs** detected."
                    )
                    st.balloons()
                    st.markdown("""
<div style='background:#031a0a;border:1px solid #16a34a;border-radius:10px;
     padding:14px 18px;margin-top:12px;font-family:monospace;font-size:12px'>
  <b style='color:#22c55e'>🎉 Data lake is now active!</b><br>
  <span style='color:#16a34a'>Switch to any analysis tab — they all read from your committed data.</span>
</div>""", unsafe_allow_html=True)
                except Exception as _ce:
                    st.error(f"❌ Commit failed: {_ce}")
                    import traceback; st.code(traceback.format_exc())

        # If already committed, show status
        elif _dl_committed:
            st.markdown(f"""
<div style='background:#031a0a;border:1px solid #16a34a;border-radius:10px;
     padding:14px 18px;font-family:monospace;font-size:12px'>
  <b style='color:#22c55e'>✅ Data Lake is ACTIVE</b><br>
  <span style='color:#64748b'>{_dl.get("source_rows",0):,} rows committed on
  {_dl.get("committed_at","—")} from <b style='color:#e2e8f0'>{_dl.get("filename","—")}</b></span>
</div>""", unsafe_allow_html=True)
