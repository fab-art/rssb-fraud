import io
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_counter_verification_xlsx(
    df: pd.DataFrame,
    deductions: list[dict],
    meta: dict,
    prepared_by: str,
    verified_by: str,
    approved_by: str,
    pc_col:  str | None = None,
    ins_col: str | None = None,
    tot_col: str | None = None,
    obs_col: str | None = None,
    dif_col: str | None = None,
) -> bytes:
    # --- Palette ---
    C_BLUE      = "003366"
    C_GOLD      = "FFCC00"
    C_GREY      = "F4F4F4"
    C_WHITE     = "FFFFFF"
    C_TEXT      = "333333"
    C_RED       = "C0392B"
    C_TITLE_BG  = "E8F0F7"
    C_FILL_AMBER= "FFF3CD"

    def fill(hex_col): return PatternFill("solid", fgColor=hex_col)
    def _font(name="Calibri", bold=False, size=11.0, color=C_TEXT): return Font(name=name, bold=bold, size=size, color=color)
    def side(style, color="000000"): return Side(border_style=style, color=color)

    THIN_GREY = side("thin", "AAAAAA"); MED_BLUE = side("medium", C_BLUE); MED_GOLD = side("medium", C_GOLD); THIN_ANY = side("thin")

    def border_hdr_mid(): return Border(left=MED_BLUE, right=MED_BLUE, top=MED_BLUE, bottom=MED_GOLD)
    def border_hdr_first(): return Border(left=THIN_ANY, right=MED_BLUE, top=MED_BLUE, bottom=MED_GOLD)
    def border_hdr_last(): return Border(left=MED_BLUE, right=THIN_ANY, top=MED_BLUE, bottom=MED_GOLD)
    def border_data(): return Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)
    def border_data_first(): return Border(left=THIN_ANY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)
    def border_data_last(): return Border(left=THIN_GREY, right=THIN_ANY, top=THIN_GREY, bottom=THIN_GREY)

    A_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    A_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    A_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)

    def _safe_float(v):
        if v is None: return 0.0
        try:
            if pd.isna(v): return 0.0
        except: pass
        try: return float(str(v).replace(",", "").replace(" ", ""))
        except: return 0.0

    wb = Workbook(); wb.remove(wb.active)
    ws1 = wb.create_sheet("After counter verification")
    ws1.sheet_view.showGridLines = False

    headers = ["No.", "Invoice ID (Paper Code)", "Beneficiary RAMA No.", "Beneficiary Name", "Dispensing Date", "Practitioner Name", "Practitioner Type", "Health Facility", "Total Cost (100%)", "Insurance Co-payment (85%)", "Patient Co-payment (15%)", "Difference", "Verified 85% Approved Amount", "Observation"]
    widths = [6.86, 17.65, 14.86, 33.45, 11.24, 25.86, 22.86, 21.04, 13.86, 13.86, 13.86, 13.86, 13.86, 45.45]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
        c = ws1.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment = fill(C_BLUE), _font(bold=True, color=C_WHITE, size=10.5), A_CENTER
        c.border = border_hdr_first() if ci == 1 else border_hdr_last() if ci == 14 else border_hdr_mid()

    ded_map = {str(d["paper_code"]).strip(): d for d in deductions}
    for ri_idx, (_, row) in enumerate(df.iterrows(), 2):
        pc = str(row.get(pc_col, "")).strip(); ded = ded_map.get(pc); is_ded = ded is not None
        ins_val = _safe_float(row.get(ins_col)); diff = _safe_float(ded["amount"]) if is_ded else 0.0
        vals = [ri_idx-1, pc, str(row.get("rama", "")), str(row.get("patient", "")), str(row.get("visit_date", "")), str(row.get("doctor", "")), "", "", _safe_float(row.get(tot_col)), ins_val, _safe_float(row.get(tot_col, 0)) - ins_val, diff, ins_val + diff, str(ded["explanation"]) if is_ded else ""]
        row_fill = fill(C_FILL_AMBER) if is_ded else fill(C_WHITE) if ri_idx % 2 == 0 else fill(C_GREY)
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(row=ri_idx, column=ci, value=val); c.fill = row_fill; c.font = _font("Calibri", size=10.5)
            c.border = border_data_first() if ci == 1 else border_data_last() if ci == 14 else border_data()
            if ci in (9, 10, 11, 12, 13): c.number_format = "#,##0"; c.alignment = A_RIGHT
            elif ci == 1: c.alignment = A_CENTER
            else: c.alignment = A_LEFT

    ws2 = wb.create_sheet("Counter verification report")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:E3"); t = ws2["A1"]; t.value = "RSSB - COUNTER VERIFICATION REPORT"
    t.font, t.fill, t.alignment = _font(bold=True, size=36, color=C_BLUE), fill(C_TITLE_BG), A_CENTER
    ws2.row_dimensions[4].height = 3.75
    for ci in range(1, 6): ws2.cell(row=4, column=ci).fill = fill(C_GOLD)
    meta_rows = [(5, "PROVINCE:", meta.get("province", "")), (6, "ADMINISTRATIVE DISTRICT:", meta.get("district", "")), (7, "PHARMACY:", meta.get("pharmacy", "")), (8, "PERIOD:", meta.get("period", "")), (9, "CODE:", meta.get("code", ""))]
    for rn, label, value in meta_rows:
        ws2.cell(row=rn, column=1, value=label).font = _font(bold=True, size=13, color=C_BLUE)
        ws2.cell(row=rn, column=3, value=value).font = _font(bold=True, size=13)
    th = ["No.", "Invoice ID (Paper Code)", "Beneficiary RAMA No.", "Amount Deducted (RWF)", "Explanation of Deduction"]
    for ci, h in enumerate(th, 1):
        c = ws2.cell(row=11, column=ci, value=h); c.fill, c.font, c.alignment = fill(C_BLUE), _font(bold=True, color=C_WHITE, size=10.5), A_CENTER
    for i, d in enumerate(deductions):
        ri = 12 + i; row_fill = fill(C_WHITE) if i % 2 == 0 else fill(C_GREY)
        vals = [i+1, d["paper_code"], d["rama_no"], d["amount"], d["explanation"]]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=v); c.fill, c.border = row_fill, border_data()
            if ci == 4: c.font, c.number_format, c.alignment = _font(color=C_RED), "#,##0", A_RIGHT
            elif ci == 1: c.alignment = A_CENTER
            else: c.alignment = A_LEFT
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
