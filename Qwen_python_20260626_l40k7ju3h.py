import streamlit as st
import pandas as pd
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="PharmaScan Workbench", page_icon="💊", layout="wide")

# ─── HELPER FUNCTIONS ────────────────────────────────────────────────────────
def clean_and_load_data(file_bytes, filename):
    """Loads the pharmacy file and handles messy headers/dates."""
    if filename.endswith('.csv'):
        df = pd.read_csv(io.BytesIO(file_bytes))
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0)
    
    # Fix messy headers (drop "Invoice Report" if it's the first row)
    if df.columns[0].lower().strip() in ['#', 'invoice report', 'unnamed: 0']:
        df.columns = df.iloc[0]
        df = df[1:].reset_index(drop=True)
        
    df.columns = [str(c).strip() for c in df.columns]
    
    # Standardize column names for the app
    col_map = {
        'Paper Code': 'paper_code', 'Dispensing Date': 'dispensing_date',
        'Patient Name': 'patient_name', 'RAMA Number': 'rama_number',
        'Practitioner Name': 'doctor_name', 'Total Cost': 'total_cost',
        'Insurance Co-payment': 'insurance_copay', 'Patient Co-payment': 'patient_copay'
    }
    df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
    
    # Convert dates (handles Excel serial numbers like 45901)
    if 'dispensing_date' in df.columns:
        df['dispensing_date'] = pd.to_datetime(df['dispensing_date'], origin='1899-12-30', unit='D', errors='coerce')
        
    # Convert financial columns to numeric
    for col in ['total_cost', 'insurance_copay', 'patient_copay']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
    return df

def generate_rssb_excel(df, annotations, metadata):
    """Generates the exact 2-sheet RSSB Counter-Verification Excel format."""
    wb = Workbook()
    
    # --- STYLES ---
    blue_fill = PatternFill("solid", fgColor="003366")
    white_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ═══════════════════════════════════════════════════════════
    # SHEET 1: After counter verification
    # ═══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "After counter verification"
    
    headers_s1 = [
        "Paper Code", "Dispensing Date", "Patient Name", "RAMA Number",
        "Practitioner Name", "Health Facility", "Date of Treatment", "Verified",
        "Total Before Counter-V (RWF)", "85% After Counter-V (RWF)",
        "After Counter-V 100%", "After Counter-V 85%",
        "Amount Deducted (RWF)", "Explanation"
    ]
    
    # Write Headers
    for ci, h in enumerate(headers_s1, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Write Data
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        pc = str(row.get('paper_code', ''))
        ann = annotations.get(pc, {})
        
        is_deducted = ann.get('status') == 'Deduct'
        ded_amount = float(ann.get('amount', 0)) if is_deducted else 0
        explanation = ann.get('reason', '') if is_deducted else ''
        verified = "NO" if is_deducted else "YES"
        
        ins_co = float(row.get('insurance_copay', 0))
        total_85 = round(ins_co * 0.85, 2)
        after_100 = round(ins_co - abs(ded_amount), 2)
        after_85 = round(after_100 * 0.85, 2)
        
        date_str = row['dispensing_date'].strftime('%d/%m/%Y') if pd.notna(row.get('dispensing_date')) else ''
        
        vals = [
            pc, date_str, row.get('patient_name', ''), row.get('rama_number', ''),
            row.get('doctor_name', ''), metadata.get('pharmacy', ''), date_str, verified,
            ins_co, total_85, after_100, after_85, -abs(ded_amount), explanation
        ]
        
        for ci, val in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            if ci == 8: # Verified column
                cell.fill = PatternFill("solid", fgColor="D4EDDA" if verified == "YES" else "FFF3CD")
                cell.font = Font(bold=True, color="1E7E34" if verified == "YES" else "B8860B")
            if ci in [9, 10, 11, 12, 13]:
                cell.number_format = '#,##0.00'
                
    # Adjust column widths
    for ci in range(1, 15):
        ws1.column_dimensions[get_column_letter(ci)].width = 18

    # ═══════════════════════════════════════════════════════════
    # SHEET 2: Counter verification report
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Counter verification report")
    
    # Metadata
    ws2.merge_cells('A1:E1')
    ws2['A1'] = "RSSB - COUNTER VERIFICATION REPORT"
    ws2['A1'].font = Font(bold=True, size=20, color="003366")
    ws2['A1'].alignment = Alignment(horizontal="center")
    
    meta_data = [
        ("PROVINCE:", metadata.get('province', '')),
        ("DISTRICT:", metadata.get('district', '')),
        ("PHARMACY:", metadata.get('pharmacy', '')),
        ("PERIOD:", metadata.get('period', '')),
    ]
    for i, (label, val) in enumerate(meta_data, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True, color="003366")
        ws2.cell(row=i, column=2, value=val)
        
    # Deductions Table
    start_row = 9
    ws2.cell(row=start_row, column=1, value="No.").font = white_font
    ws2.cell(row=start_row, column=2, value="Paper Code").font = white_font
    ws2.cell(row=start_row, column=3, value="RAMA No.").font = white_font
    ws2.cell(row=start_row, column=4, value="Amount Deducted").font = white_font
    ws2.cell(row=start_row, column=5, value="Explanation").font = white_font
    for ci in range(1, 6):
        ws2.cell(row=start_row, column=ci).fill = blue_fill
        
    deductions = [(pc, ann) for pc, ann in annotations.items() if ann.get('status') == 'Deduct']
    for i, (pc, ann) in enumerate(deductions, 1):
        r = start_row + i
        ws2.cell(row=r, column=1, value=i)
        ws2.cell(row=r, column=2, value=pc)
        # Lookup RAMA from df
        rama = df[df['paper_code'].astype(str) == pc]['rama_number'].iloc[0] if pc in df['paper_code'].astype(str).values else ''
        ws2.cell(row=r, column=3, value=rama)
        ws2.cell(row=r, column=4, value=-abs(float(ann.get('amount', 0))))
        ws2.cell(row=r, column=4).number_format = '#,##0'
        ws2.cell(row=r, column=5, value=ann.get('reason', ''))
        
    # Signature Block
    sig_row = start_row + len(deductions) + 3
    ws2.cell(row=sig_row, column=1, value="Prepared by:").font = Font(bold=True, underline="single")
    ws2.cell(row=sig_row, column=3, value="Verified by:").font = Font(bold=True, underline="single")
    ws2.cell(row=sig_row+2, column=3, value=metadata.get('verifier', ''))
    
    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─── STREAMLIT APP ───────────────────────────────────────────────────────────
def main():
    st.title("💊 PharmaScan Verification Workbench")
    st.caption("Replace manual Excel editing with an interactive card-based audit interface.")

    # --- SIDEBAR ---
    with st.sidebar:
        st.header("📂 Data & Metadata")
        uploaded_file = st.file_uploader("Upload Pharmacy Voucher File", type=['xlsx', 'csv'])
        
        st.divider()
        st.subheader("📋 Report Metadata")
        meta_province = st.text_input("Province", "KIGALI")
        meta_district = st.text_input("District", "GASABO")
        meta_pharmacy = st.text_input("Pharmacy Name", "PHARMACIE VINCA GISENYI LTD")
        meta_period = st.text_input("Period", "MAY 2024")
        meta_verifier = st.text_input("Verified By", "Alphonsine MUKAKAYIBANDA")
        
        metadata = {
            'province': meta_province, 'district': meta_district,
            'pharmacy': meta_pharmacy, 'period': meta_period, 'verifier': meta_verifier
        }

    if not uploaded_file:
        st.info("👆 Upload a pharmacy voucher file in the sidebar to begin.")
        return

    # --- LOAD DATA ---
    if 'df' not in st.session_state or st.session_state.get('file_name') != uploaded_file.name:
        with st.spinner("Loading and cleaning data..."):
            st.session_state.df = clean_and_load_data(uploaded_file.read(), uploaded_file.name)
            st.session_state.file_name = uploaded_file.name
            st.session_state.annotations = {} # Reset annotations on new file
            
    df = st.session_state.df
    annotations = st.session_state.annotations

    # --- MAIN TABS ---
    tab_workbench, tab_summary = st.tabs(["📝 Verification Workbench", "📊 Summary & Export"])

    # ═══════════════════════════════════════════════════════════
    # TAB 1: WORKBENCH (Card Interface)
    # ═══════════════════════════════════════════════════════════
    with tab_workbench:
        st.subheader(f"Auditing {len(df)} Claims")
        
        # Filters
        f1, f2, f3 = st.columns([2, 1, 1])
        with f1:
            search_term = st.text_input("🔍 Search by Patient, RAMA, or Paper Code", "")
        with f2:
            filter_status = st.multiselect("Filter by Audit Status", 
                                           ["Pending", "Verified", "Deduct", "Ghost"], 
                                           default=["Pending", "Verified", "Deduct", "Ghost"])
        with f3:
            sort_by = st.selectbox("Sort by", ["Default", "Total Cost (High-Low)", "Insurance Cost (High-Low)"])

        # Apply filters
        filtered_df = df.copy()
        if search_term:
            mask = filtered_df.apply(lambda row: row.astype(str).str.contains(search_term, case=False).any(), axis=1)
            filtered_df = filtered_df[mask]
            
        if sort_by == "Total Cost (High-Low)" and 'total_cost' in filtered_df.columns:
            filtered_df = filtered_df.sort_values('total_cost', ascending=False)
        elif sort_by == "Insurance Cost (High-Low)" and 'insurance_copay' in filtered_df.columns:
            filtered_df = filtered_df.sort_values('insurance_copay', ascending=False)

        # Pagination
        page_size = 6
        total_pages = max(1, (len(filtered_df) + page_size - 1) // page_size)
        page = st.number_input("Page", min_value=1, max_value=total_pages, value=1, step=1)
        start_idx = (page - 1) * page_size
        end_idx = min(start_idx + page_size, len(filtered_df))
        page_df = filtered_df.iloc[start_idx:end_idx]

        # Render Cards
        for i in range(0, len(page_df), 2):
            cols = st.columns(2)
            for j, col in enumerate(cols):
                idx = start_idx + i + j
                if idx < len(filtered_df):
                    row = filtered_df.iloc[idx]
                    pc = str(row.get('paper_code', idx))
                    ann = annotations.get(pc, {'status': 'Pending', 'amount': 0, 'reason': ''})
                    
                    with col:
                        with st.container(border=True):
                            # Card Header
                            h1, h2 = st.columns([3, 1])
                            with h1:
                                st.markdown(f"**Paper Code:** `{pc}`")
                                st.caption(f"{row.get('patient_name', 'Unknown')} | RAMA: {row.get('rama_number', 'N/A')}")
                            with h2:
                                status = ann['status']
                                color = "gray" if status == "Pending" else "green" if status == "Verified" else "red" if status == "Deduct" else "orange"
                                st.markdown(f":{color}[**{status}**]")
                            
                            st.divider()
                            
                            # Card Details
                            c1, c2, c3 = st.columns(3)
                            c1.metric("Total Cost", f"{row.get('total_cost', 0):,.0f} RWF")
                            c2.metric("Insurance (85%)", f"{row.get('insurance_copay', 0):,.0f} RWF")
                            c3.metric("Date", row['dispensing_date'].strftime('%d/%m/%Y') if pd.notna(row.get('dispensing_date')) else 'N/A')
                            
                            st.caption(f"👨‍⚕️ {row.get('doctor_name', 'N/A')}")
                            
                            # Actions
                            st.markdown("**Audit Actions:**")
                            a1, a2 = st.columns([1, 1])
                            new_status = a1.selectbox("Decision", ["Pending", "Verified", "Deduct", "Ghost"], 
                                                      index=["Pending", "Verified", "Deduct", "Ghost"].index(status),
                                                      key=f"status_{pc}")
                            
                            new_amount = 0.0
                            new_reason = ""
                            if new_status == "Deduct":
                                new_amount = a2.number_input("Deduction (RWF)", min_value=0.0, value=float(ann['amount']), key=f"amt_{pc}")
                                new_reason = st.text_input("Reason / Observation", value=ann['reason'], key=f"reason_{pc}")
                            
                            if st.button("💾 Save Annotation", key=f"save_{pc}", use_container_width=True):
                                annotations[pc] = {
                                    'status': new_status,
                                    'amount': new_amount,
                                    'reason': new_reason
                                }
                                st.session_state.annotations = annotations
                                st.rerun()

    # ═══════════════════════════════════════════════════════════
    # TAB 2: SUMMARY & EXPORT
    # ═══════════════════════════════════════════════════════════
    with tab_summary:
        st.subheader("Audit Summary")
        
        # Calculate stats
        total_claims = len(df)
        verified = sum(1 for ann in annotations.values() if ann['status'] == 'Verified')
        deducted = sum(1 for ann in annotations.values() if ann['status'] == 'Deduct')
        ghosts = sum(1 for ann in annotations.values() if ann['status'] == 'Ghost')
        pending = total_claims - len(annotations)
        
        total_deducted_amount = sum(float(ann.get('amount', 0)) for ann in annotations.values() if ann['status'] == 'Deduct')
        
        s1, s2, s3, s4, s5 = st.columns(5)
        s1.metric("Total Claims", total_claims)
        s2.metric("✅ Verified", verified)
        s3.metric("💸 Deducted", deducted, help=f"Total saved: {total_deducted_amount:,.0f} RWF")
        s4.metric("👻 Ghost/No Record", ghosts)
        s5.metric("⏳ Pending", pending)
        
        st.divider()
        
        st.subheader("📥 Generate RSSB Counter-Verification Report")
        st.caption("This will generate the exact 2-sheet Excel format required by RSSB, including metadata, deductions, and signature blocks.")
        
        if st.button("🚀 Generate & Download Excel Report", type="primary", use_container_width=True):
            with st.spinner("Building Excel workbook..."):
                excel_bytes = generate_rssb_excel(df, annotations, metadata)
                
            st.success("✅ Report generated successfully!")
            st.download_button(
                label="⬇️ Download RSSB Report (.xlsx)",
                data=excel_bytes,
                file_name=f"RSSB_CV_{meta_pharmacy.replace(' ', '_')}_{meta_period.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

if __name__ == "__main__":
    main()